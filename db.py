"""
db.py — Persistent disk storage for the subscription bot.

All data lives under DISK_PATH (Render persistent disk mount).
Structure:
  /data/
    users/
      {user_id}.json        ← one file per user (profile + APIs + stats)
    sessions/
      {user_id}.json        ← volatile P2P session state (reset every 12h)
    upgrade_requests.json   ← pending upgrade requests
    stats.json              ← global stats

User JSON schema:
  {
    "user_id":       int,
    "username":      str,
    "display_name":  str,
    "plan":          "free" | "pro",
    "plan_expires":  null | "YYYY-MM-DD HH:MM:SS",
    "created_at":    "YYYY-MM-DD HH:MM:SS",
    "upgrade_pending": bool,
    "apis": {
      "bybit_key":    "",
      "bybit_secret": "",
      "flw_key":      "",
      "flw_secret":   "",
      "flw_hash":     "",
      "paga_principal":  "",
      "paga_credential": "",
      "paga_api_key":    ""
    },
    "stats": {
      "total_buy_orders":  0,
      "total_sell_orders": 0,
      "last_active":       ""
    }
  }
"""

import json
import os
import logging
from datetime import datetime, timedelta
from pathlib import Path
from threading import Lock

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────
# Disk path — Render persistent disk mounts at /data by default
# Override with DISK_PATH env var if needed
# ─────────────────────────────────────────
DISK_PATH      = Path(os.getenv("DISK_PATH", "/data"))
USERS_DIR      = DISK_PATH / "users"
SESSION_DIR    = DISK_PATH / "sessions"
UPGRADE_REQ    = DISK_PATH / "upgrade_requests.json"
STATS_FILE     = DISK_PATH / "stats.json"
REFERRALS_FILE   = DISK_PATH / "referrals.json"
WITHDRAWALS_FILE = DISK_PATH / "withdrawals.json"

_lock = Lock()


def _init_dirs():
    for d in [USERS_DIR, SESSION_DIR]:
        d.mkdir(parents=True, exist_ok=True)
    if not UPGRADE_REQ.exists():
        _write_json(UPGRADE_REQ, {})
    if not STATS_FILE.exists():
        _write_json(STATS_FILE, {"total_users": 0})
    if not REFERRALS_FILE.exists():
        _write_json(REFERRALS_FILE, {})
    if not WITHDRAWALS_FILE.exists():
        _write_json(WITHDRAWALS_FILE, {})

def _read_json(path: Path, default=None):
    """
    Safe JSON read with corruption recovery.
    - Returns default on missing file, empty file, or corrupted JSON.
    - On corruption: logs the error and renames the bad file for inspection.
    """
    _default = default if default is not None else {}
    try:
        if not path.exists():
            return _default
        content = path.read_text(encoding="utf-8").strip()
        if not content:
            logger.warning(f"[DB] Empty file: {path} — returning default")
            return _default
        return json.loads(content)
    except json.JSONDecodeError as e:
        logger.error(f"[DB] JSON corruption in {path}: {e} — backing up and returning default")
        try:
            backup = path.with_suffix(".corrupt")
            path.rename(backup)
            logger.warning(f"[DB] Corrupt file moved to {backup}")
        except Exception:
            pass
        return _default
    except Exception as e:
        logger.error(f"[DB] Read error {path}: {e}")
        return _default

def _write_json(path: Path, data):
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        tmp.replace(path)
    except Exception as e:
        logger.error(f"[DB] Write failed {path}: {e}")

def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ─────────────────────────────────────────
# User CRUD
# ─────────────────────────────────────────
def _user_path(user_id: int) -> Path:
    return USERS_DIR / f"{user_id}.json"

def _default_user(user_id: int, username: str, display_name: str) -> dict:
    return {
        "user_id":          user_id,
        "username":         username or "",
        "display_name":     display_name or "",
        "plan":             "free",
        "plan_expires":     None,
        "created_at":       _now(),
        "upgrade_pending":  False,
        "apis": {
            "bybit_key":       "",
            "bybit_secret":    "",
            "flw_key":         "",
            "flw_secret":      "",
            "flw_hash":        "",
            "paga_principal":  "",
            "paga_credential": "",
            "paga_api_key":    "",
        },
        "stats": {
            "total_buy_orders":  0,
            "total_sell_orders": 0,
            "last_active":       _now(),
        },
        "referral": {
            "code":                "",     # generated lazily on first access
            "referred_by":         None,   # user_id of referrer, set once, immutable
            "balance":             0,      # NGN available balance (admin can adjust)
            "pending_commission":  0,      # informational only — live value comes from referrals.json
            "total_earned":        0,      # lifetime NGN approved into balance
            "bank": {                      # last-used withdrawal bank details, for reference only
                "account_number": "",
                "account_name":   "",
                "bank_name":      "",
            },
        }
    }

def get_user(user_id: int) -> dict | None:
    path = _user_path(user_id)
    if not path.exists():
        return None
    return _read_json(path)

def get_or_create_user(user_id: int, username: str, display_name: str) -> tuple[dict, bool]:
    """Returns (user_dict, is_new)."""
    with _lock:
        path = _user_path(user_id)
        if path.exists():
            user = _read_json(path)
            # Update username/display_name in case they changed
            user["username"]     = username or user.get("username", "")
            user["display_name"] = display_name or user.get("display_name", "")
            user["stats"]["last_active"] = _now()
            _write_json(path, user)
            return user, False
        user = _default_user(user_id, username, display_name)
        _write_json(path, user)
        logger.info(f"[DB] New user created: {user_id} @{username}")
        return user, True

def save_user(user: dict):
    with _lock:
        _write_json(_user_path(user["user_id"]), user)

def get_all_users() -> list:
    users = []
    for f in USERS_DIR.glob("*.json"):
        u = _read_json(f)
        if u:
            users.append(u)
    return users


# ─────────────────────────────────────────
# P2P Settings persistence
# Each user's P2P settings (ad_id, mode, interval, UIDs, etc.) are stored
# inside their user JSON file under the "p2p_settings" key.
# This lets settings survive bot restarts and be fully isolated per user.
# ─────────────────────────────────────────
def save_settings(user_id: int, settings: dict):
    """Persist a user's P2P settings dict to disk."""
    with _lock:
        user = _read_json(_user_path(user_id))
        if not user:
            return
        user["p2p_settings"] = settings
        _write_json(_user_path(user_id), user)

def load_settings(user_id: int) -> dict:
    """Load a user's P2P settings from disk. Returns empty dict if none saved."""
    user = get_user(user_id)
    if not user:
        return {}
    return user.get("p2p_settings", {})


# ─────────────────────────────────────────
# API key management
# ─────────────────────────────────────────
def save_api(user_id: int, key: str, value: str):
    """Save a single API key for a user. key = 'bybit_key', 'flw_key', etc."""
    with _lock:
        user = get_user(user_id)
        if not user:
            return
        user["apis"][key] = value
        _write_json(_user_path(user_id), user)

def get_api(user_id: int, key: str) -> str:
    user = get_user(user_id)
    if not user:
        return ""
    return user.get("apis", {}).get(key, "")

def delete_all_apis(user_id: int):
    with _lock:
        user = get_user(user_id)
        if not user:
            return
        # Clear all API entries — covers both old single-key and new slot-based keys
        user["apis"] = {}
        _write_json(_user_path(user_id), user)
        logger.info(f"[DB] APIs deleted for user {user_id}")


# ─────────────────────────────────────────
# Subscription management
# ─────────────────────────────────────────
def is_pro(user_id: int) -> bool:
    user = get_user(user_id)
    if not user:
        return False
    if user.get("plan") != "pro":
        return False
    expires = user.get("plan_expires")
    if not expires:
        return True   # no expiry set = lifetime
    try:
        return datetime.strptime(expires, "%Y-%m-%d %H:%M:%S") > datetime.now()
    except Exception:
        return False

def upgrade_user(user_id: int, days: int) -> dict:
    """Set user to pro plan for `days` days. Returns updated user."""
    with _lock:
        user = get_user(user_id)
        if not user:
            return {}
        now    = datetime.now()
        # Extend from current expiry if still active
        current_exp = user.get("plan_expires")
        if current_exp:
            try:
                base = datetime.strptime(current_exp, "%Y-%m-%d %H:%M:%S")
                if base > now:
                    expires = base + timedelta(days=days)
                else:
                    expires = now + timedelta(days=days)
            except Exception:
                expires = now + timedelta(days=days)
        else:
            expires = now + timedelta(days=days)
        user["plan"]             = "pro"
        user["plan_expires"]     = expires.strftime("%Y-%m-%d %H:%M:%S")
        user["upgrade_pending"]  = False
        _write_json(_user_path(user_id), user)
        logger.info(f"[DB] Upgraded user {user_id} → pro until {user['plan_expires']}")
        return user

def downgrade_user(user_id: int) -> dict:
    with _lock:
        user = get_user(user_id)
        if not user:
            return {}
        user["plan"]         = "free"
        user["plan_expires"] = None
        _write_json(_user_path(user_id), user)
        logger.info(f"[DB] Downgraded user {user_id} → free")
        return user

def get_plan_expiry_str(user_id: int) -> str:
    user = get_user(user_id)
    if not user or user.get("plan") != "pro":
        return "Free plan"
    exp = user.get("plan_expires")
    if not exp:
        return "Pro (lifetime)"
    try:
        dt   = datetime.strptime(exp, "%Y-%m-%d %H:%M:%S")
        days = (dt - datetime.now()).days
        return f"Pro — expires {exp} ({days}d left)"
    except Exception:
        return f"Pro — expires {exp}"

def check_and_auto_downgrade(user_id: int) -> bool:
    """Returns True if user was auto-downgraded (plan expired)."""
    user = get_user(user_id)
    if not user or user.get("plan") != "pro":
        return False
    exp = user.get("plan_expires")
    if not exp:
        return False
    try:
        if datetime.strptime(exp, "%Y-%m-%d %H:%M:%S") <= datetime.now():
            downgrade_user(user_id)
            logger.info(f"[DB] Auto-downgraded expired user {user_id}")
            return True
    except Exception:
        pass
    return False


# ─────────────────────────────────────────
# Upgrade requests
# ─────────────────────────────────────────
def request_upgrade(user_id: int, username: str, display_name: str):
    with _lock:
        reqs = _read_json(UPGRADE_REQ, {})
        reqs[str(user_id)] = {
            "user_id":      user_id,
            "username":     username,
            "display_name": display_name,
            "requested_at": _now(),
        }
        _write_json(UPGRADE_REQ, reqs)
        # Mark on user profile — use _write_json directly (we already hold _lock,
        # calling save_user() would try to re-acquire it and deadlock)
        user = _read_json(_user_path(user_id))
        if user:
            user["upgrade_pending"] = True
            _write_json(_user_path(user_id), user)
        logger.info(f"[DB] Upgrade request saved for user {user_id}")

def get_pending_requests() -> list:
    reqs = _read_json(UPGRADE_REQ, {})
    return list(reqs.values())

def remove_upgrade_request(user_id: int):
    with _lock:
        reqs = _read_json(UPGRADE_REQ, {})
        reqs.pop(str(user_id), None)
        _write_json(UPGRADE_REQ, reqs)


# ─────────────────────────────────────────
# Stats
# ─────────────────────────────────────────
def increment_stat(user_id: int, stat: str, amount: int = 1):
    """stat = 'total_buy_orders' | 'total_sell_orders'"""
    with _lock:
        user = get_user(user_id)
        if not user:
            return
        user["stats"][stat] = user["stats"].get(stat, 0) + amount
        user["stats"]["last_active"] = _now()
        _write_json(_user_path(user_id), user)


def update_user_stats(user_id: int, **kwargs):
    """
    Update one or more stats fields for a user.
    Accepts any keyword argument that maps to a field in user['stats'].
    Common usage:
        update_user_stats(uid, last_active="2025-01-01 12:00:00")
        update_user_stats(uid, total_buy_orders=5)
    Always safe to call — silently returns if user not found.
    """
    with _lock:
        user = get_user(user_id)
        if not user:
            return
        if "stats" not in user or not isinstance(user["stats"], dict):
            user["stats"] = {
                "total_buy_orders":  0,
                "total_sell_orders": 0,
                "last_active":       "",
            }
        for key, value in kwargs.items():
            user["stats"][key] = value
        _write_json(_user_path(user_id), user)


# ─────────────────────────────────────────
# Session state (volatile P2P data — reset every 12h)
# ─────────────────────────────────────────
def _session_path(user_id: int) -> Path:
    return SESSION_DIR / f"{user_id}.json"

def load_session(user_id: int) -> dict:
    path = _session_path(user_id)
    if not path.exists():
        return {}
    data = _read_json(path, {})
    # Check if session is older than 12 hours — auto-reset
    ts = data.get("_saved_at", "")
    if ts:
        try:
            age = (datetime.now() - datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")).total_seconds()
            if age > 12 * 3600:
                logger.info(f"[DB] Session for {user_id} is {age/3600:.1f}h old — auto-resetting")
                clear_session(user_id)
                return {}
        except Exception:
            pass
    return data

def save_session(user_id: int, data: dict):
    data["_saved_at"] = _now()
    with _lock:
        _write_json(_session_path(user_id), data)

def clear_session(user_id: int):
    path = _session_path(user_id)
    try:
        if path.exists():
            path.unlink()
    except Exception as e:
        logger.error(f"[DB] clear_session {user_id}: {e}")

def clear_all_old_sessions():
    """Call periodically — clears sessions older than 12 hours."""
    count = 0
    for f in SESSION_DIR.glob("*.json"):
        data = _read_json(f, {})
        ts   = data.get("_saved_at", "")
        if ts:
            try:
                age = (datetime.now() - datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")).total_seconds()
                if age > 12 * 3600:
                    f.unlink()
                    count += 1
            except Exception:
                pass
    if count:
        logger.info(f"[DB] Cleared {count} stale sessions")
    return count


# ─────────────────────────────────────────
# Referral system
# ─────────────────────────────────────────
# Every user gets a referral code that's a deterministic, reversible
# encoding of their own Telegram user_id — so there's no separate code
# index to keep in sync, and codes can never collide.
#
# referrals.json holds one record per REFERRED user (keyed by their
# user_id), tracking who referred them and the state of their reward:
#   "none"     → joined via link, not upgraded to Pro yet — no reward owed
#   "pending"  → upgraded to Pro — commission owed, admin hasn't approved yet
#   "approved" → admin approved — amount has been moved into referrer's balance
# ─────────────────────────────────────────
_B36_ALPHABET = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def _b36_encode(n: int) -> str:
    if n == 0:
        return "0"
    n = abs(n)
    digits = []
    while n:
        n, r = divmod(n, 36)
        digits.append(_B36_ALPHABET[r])
    return "".join(reversed(digits))

def _b36_decode(s: str):
    try:
        return int(s, 36)
    except Exception:
        return None

def _ensure_referral_block(user: dict):
    """Backfill the 'referral' block on user dicts created before this feature existed."""
    ref = user.get("referral")
    if not isinstance(ref, dict):
        user["referral"] = {
            "code": "", "referred_by": None,
            "balance": 0, "pending_commission": 0, "total_earned": 0,
            "bank": {"account_number": "", "account_name": "", "bank_name": ""},
        }
    else:
        ref.setdefault("code", "")
        ref.setdefault("referred_by", None)
        ref.setdefault("balance", 0)
        ref.setdefault("pending_commission", 0)
        ref.setdefault("total_earned", 0)
        ref.setdefault("bank", {"account_number": "", "account_name": "", "bank_name": ""})

def get_or_create_referral_code(user_id: int) -> str:
    """Return this user's referral code, generating it on first call."""
    with _lock:
        user = _read_json(_user_path(user_id))
        if not user:
            return ""
        _ensure_referral_block(user)
        code = user["referral"]["code"]
        if not code:
            code = "R" + _b36_encode(user_id)
            user["referral"]["code"] = code
            _write_json(_user_path(user_id), user)
        return code

def resolve_referral_code(code: str):
    """Decode a referral code back into a user_id. Returns None if invalid
    or the encoded user doesn't actually exist in the DB."""
    if not code or not code.upper().startswith("R"):
        return None
    uid = _b36_decode(code[1:].upper())
    if uid is None or not _user_path(uid).exists():
        return None
    return uid

def _read_referrals() -> dict:
    return _read_json(REFERRALS_FILE, {})

def _write_referrals(data: dict):
    _write_json(REFERRALS_FILE, data)

def record_referral_join(new_user_id: int, referrer_id: int, username: str, display_name: str) -> bool:
    """
    Link a brand-new user to whoever referred them. Anti-cheat rules:
      - Only fires for accounts that don't already have a referrer (immutable, one-time).
      - Self-referral (referrer_id == new_user_id) is rejected.
      - The referrer must be a real, existing user.
    Returns True if the link was recorded.
    """
    if not referrer_id or referrer_id == new_user_id:
        return False
    with _lock:
        user = _read_json(_user_path(new_user_id))
        if not user:
            return False
        _ensure_referral_block(user)
        if user["referral"]["referred_by"]:
            return False
        if not _user_path(referrer_id).exists():
            return False

        user["referral"]["referred_by"] = referrer_id
        _write_json(_user_path(new_user_id), user)

        refs = _read_referrals()
        refs[str(new_user_id)] = {
            "referrer_id":           referrer_id,
            "referred_username":     username or "",
            "referred_display_name": display_name or "",
            "joined_at":             _now(),
            "reward_status":         "none",
            "reward_amount":         0,
            "approved_at":           None,
        }
        _write_referrals(refs)
        logger.info(f"[Referral] user {new_user_id} joined via referrer {referrer_id}")
        return True

def get_referrer(user_id: int):
    user = get_user(user_id)
    if not user:
        return None
    _ensure_referral_block(user)
    return user["referral"]["referred_by"]

def get_referral_record(referred_user_id: int):
    return _read_referrals().get(str(referred_user_id))

def get_referrals_for(referrer_id: int) -> list:
    """All referral records where this user is the referrer, newest first."""
    refs = _read_referrals()
    rows = [r for r in refs.values() if r.get("referrer_id") == referrer_id]
    rows.sort(key=lambda r: r.get("joined_at", ""), reverse=True)
    return rows

def mark_reward_pending(referred_user_id: int, reward_amount: int):
    """
    Call right after a referred user's FIRST Pro upgrade is approved.
    Flips 'none' -> 'pending' exactly once — later renewal upgrades for the
    same user do not re-trigger a reward. Returns the updated record, or
    None if there's no referral for this user or it was already handled.
    """
    with _lock:
        refs = _read_referrals()
        rec  = refs.get(str(referred_user_id))
        if not rec or rec.get("reward_status") != "none":
            return None
        rec["reward_status"] = "pending"
        rec["reward_amount"] = reward_amount
        refs[str(referred_user_id)] = rec
        _write_referrals(refs)
        logger.info(f"[Referral] reward pending: referrer={rec['referrer_id']} referred={referred_user_id} amount={reward_amount}")
        return rec

def approve_referral_reward(referred_user_id: int) -> dict:
    """
    Admin approves the commission for a referral: moves reward_amount from
    'pending' into the referrer's available balance (idempotent — can't be
    approved twice). Returns {"ok", "reason", "referrer_id", "amount"}.
    """
    with _lock:
        refs = _read_referrals()
        rec  = refs.get(str(referred_user_id))
        if not rec:
            return {"ok": False, "reason": "no_referral", "referrer_id": None, "amount": 0}
        if rec.get("reward_status") == "approved":
            return {"ok": False, "reason": "already_approved", "referrer_id": rec["referrer_id"], "amount": rec.get("reward_amount", 0)}
        if rec.get("reward_status") != "pending":
            return {"ok": False, "reason": "not_pending", "referrer_id": rec["referrer_id"], "amount": 0}

        referrer_id = rec["referrer_id"]
        amount      = rec.get("reward_amount", 0)
        referrer    = _read_json(_user_path(referrer_id))
        if not referrer:
            return {"ok": False, "reason": "referrer_missing", "referrer_id": referrer_id, "amount": amount}

        _ensure_referral_block(referrer)
        referrer["referral"]["balance"]      += amount
        referrer["referral"]["total_earned"] += amount
        _write_json(_user_path(referrer_id), referrer)

        rec["reward_status"] = "approved"
        rec["approved_at"]   = _now()
        refs[str(referred_user_id)] = rec
        _write_referrals(refs)

        logger.info(f"[Referral] approved ₦{amount} to referrer {referrer_id} (referred {referred_user_id})")
        return {"ok": True, "reason": "", "referrer_id": referrer_id, "amount": amount}

def get_referral_balance(user_id: int) -> dict:
    user = get_user(user_id)
    if not user:
        return {"balance": 0, "total_earned": 0}
    _ensure_referral_block(user)
    return {
        "balance":      user["referral"]["balance"],
        "total_earned": user["referral"]["total_earned"],
    }

def adjust_balance(user_id: int, delta: int):
    """
    Admin manually adds (positive delta) or deducts (negative delta) NGN
    from a user's referral balance — used after an off-bot bank payout, or
    to correct mistakes. Floored at 0. Returns the updated referral block,
    or None if the user doesn't exist.
    """
    with _lock:
        user = _read_json(_user_path(user_id))
        if not user:
            return None
        _ensure_referral_block(user)
        user["referral"]["balance"] = max(0, user["referral"]["balance"] + delta)
        _write_json(_user_path(user_id), user)
        logger.info(f"[Referral] balance for {user_id} adjusted by {delta:+d} -> ₦{user['referral']['balance']}")
        return user["referral"]

def get_referral_leaderboard(limit: int = 10) -> list:
    """Top referrers by total lifetime earnings, for admin analytics."""
    rows = []
    for u in get_all_users():
        _ensure_referral_block(u)
        count = len(get_referrals_for(u["user_id"]))
        if count == 0 and not u["referral"]["total_earned"]:
            continue
        rows.append({
            "user_id":      u["user_id"],
            "username":     u.get("username", ""),
            "referrals":    count,
            "balance":      u["referral"]["balance"],
            "total_earned": u["referral"]["total_earned"],
        })
    rows.sort(key=lambda r: r["total_earned"], reverse=True)
    return rows[:limit]

def get_referral_summary() -> dict:
    """Global referral stats for the admin /referrals overview."""
    refs = _read_referrals()
    pending  = [r for r in refs.values() if r.get("reward_status") == "pending"]
    approved = [r for r in refs.values() if r.get("reward_status") == "approved"]
    total_balance_owed = sum(u.get("referral", {}).get("balance", 0) for u in get_all_users())
    return {
        "total_referred":      len(refs),
        "pending_count":       len(pending),
        "pending_amount":      sum(r.get("reward_amount", 0) for r in pending),
        "approved_count":      len(approved),
        "approved_amount":     sum(r.get("reward_amount", 0) for r in approved),
        "outstanding_balance": total_balance_owed,
    }


# ─────────────────────────────────────────
# Withdrawals
# ─────────────────────────────────────────
# One record per withdrawal request, keyed by a short withdrawal_id (W000001,
# W000002, ...). Status lifecycle:
#   "pending"   → submitted, amount already deducted from the user's balance
#   "completed" → admin approved — balance stays deducted, nothing more happens
#   "rejected"  → admin rejected — amount is refunded back to the user's balance
def set_bank_details(user_id: int, account_number: str, account_name: str, bank_name: str):
    """Save the user's most recently used withdrawal bank details (for reference/history)."""
    with _lock:
        user = _read_json(_user_path(user_id))
        if not user:
            return False
        _ensure_referral_block(user)
        user["referral"]["bank"] = {
            "account_number": account_number.strip(),
            "account_name":   account_name.strip(),
            "bank_name":      bank_name.strip(),
        }
        _write_json(_user_path(user_id), user)
        return True

def get_bank_details(user_id: int) -> dict:
    user = get_user(user_id)
    if not user:
        return {"account_number": "", "account_name": "", "bank_name": ""}
    _ensure_referral_block(user)
    return user["referral"]["bank"]

def _read_withdrawals() -> dict:
    return _read_json(WITHDRAWALS_FILE, {})

def _write_withdrawals(data: dict):
    _write_json(WITHDRAWALS_FILE, data)

def _next_withdrawal_id(existing: dict) -> str:
    n = len(existing) + 1
    wid = f"W{n:06d}"
    while wid in existing:   # extremely unlikely, but guard against gaps/races
        n += 1
        wid = f"W{n:06d}"
    return wid

def create_withdrawal_request(user_id: int, amount: int, bank: dict, min_amount: int) -> dict:
    """
    Validate and submit a withdrawal request. On success, the amount is
    deducted from the user's available balance IMMEDIATELY (status starts
    as "pending") — it's only refunded if the admin later rejects it.
    Returns:
        {"ok": True,  "withdrawal_id": str, "new_balance": int}
        {"ok": False, "reason": "below_minimum" | "insufficient", "balance": int}
        {"ok": False, "reason": "user_missing"}
    """
    with _lock:
        user = _read_json(_user_path(user_id))
        if not user:
            return {"ok": False, "reason": "user_missing", "balance": 0}
        _ensure_referral_block(user)
        balance = user["referral"]["balance"]

        if amount < min_amount:
            return {"ok": False, "reason": "below_minimum", "balance": balance}
        if amount > balance:
            return {"ok": False, "reason": "insufficient", "balance": balance}

        user["referral"]["balance"] = balance - amount
        user["referral"]["bank"] = {
            "account_number": bank.get("account_number", "").strip(),
            "account_name":   bank.get("account_name", "").strip(),
            "bank_name":      bank.get("bank_name", "").strip(),
        }
        _write_json(_user_path(user_id), user)

        withdrawals = _read_withdrawals()
        wid = _next_withdrawal_id(withdrawals)
        withdrawals[wid] = {
            "withdrawal_id":  wid,
            "user_id":        user_id,
            "username":       user.get("username", ""),
            "display_name":   user.get("display_name", ""),
            "amount":         amount,
            "bank":           user["referral"]["bank"],
            "status":         "pending",
            "requested_at":   _now(),
            "resolved_at":    None,
            "reject_reason":  "",
        }
        _write_withdrawals(withdrawals)
        logger.info(f"[Withdraw] {wid}: user {user_id} requested ₦{amount} — balance now ₦{user['referral']['balance']}")
        return {"ok": True, "withdrawal_id": wid, "new_balance": user["referral"]["balance"]}

def get_withdrawal(withdrawal_id: str):
    return _read_withdrawals().get(withdrawal_id)

def get_withdrawals_for(user_id: int) -> list:
    """This user's full withdrawal history, most recent first."""
    rows = [w for w in _read_withdrawals().values() if w.get("user_id") == user_id]
    rows.sort(key=lambda w: w.get("requested_at", ""), reverse=True)
    return rows

def get_pending_withdrawals() -> list:
    """All withdrawals awaiting admin action, oldest first (fair queue order)."""
    rows = [w for w in _read_withdrawals().values() if w.get("status") == "pending"]
    rows.sort(key=lambda w: w.get("requested_at", ""))
    return rows

def approve_withdrawal(withdrawal_id: str) -> dict:
    """Mark a pending withdrawal completed. Balance was already deducted at
    request time, so nothing more moves — this just confirms it was paid."""
    with _lock:
        withdrawals = _read_withdrawals()
        w = withdrawals.get(withdrawal_id)
        if not w:
            return {"ok": False, "reason": "not_found"}
        if w.get("status") != "pending":
            return {"ok": False, "reason": "not_pending", "status": w.get("status")}
        w["status"]      = "completed"
        w["resolved_at"] = _now()
        withdrawals[withdrawal_id] = w
        _write_withdrawals(withdrawals)
        logger.info(f"[Withdraw] {withdrawal_id} approved/completed")
        return {"ok": True, "withdrawal": w}

def reject_withdrawal(withdrawal_id: str, reason: str = "") -> dict:
    """Reject a pending withdrawal and refund the amount back to the user's balance."""
    with _lock:
        withdrawals = _read_withdrawals()
        w = withdrawals.get(withdrawal_id)
        if not w:
            return {"ok": False, "reason": "not_found"}
        if w.get("status") != "pending":
            return {"ok": False, "reason": "not_pending", "status": w.get("status")}

        user = _read_json(_user_path(w["user_id"]))
        if user:
            _ensure_referral_block(user)
            user["referral"]["balance"] += w.get("amount", 0)
            _write_json(_user_path(w["user_id"]), user)

        w["status"]         = "rejected"
        w["resolved_at"]    = _now()
        w["reject_reason"]  = reason
        withdrawals[withdrawal_id] = w
        _write_withdrawals(withdrawals)
        logger.info(f"[Withdraw] {withdrawal_id} rejected — ₦{w.get('amount',0)} refunded to user {w['user_id']}")
        return {"ok": True, "withdrawal": w}


# ─────────────────────────────────────────
# Export for admin
# ─────────────────────────────────────────
def export_users_to_excel() -> bytes:
    """Return Excel file bytes with user stats table."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        headers = [
            "User ID", "Username", "Display Name", "Plan",
            "Plan Expires", "Upgrade Pending", "Created At",
            "Total Buy Orders", "Total Sell Orders", "Last Active",
            "Referred By", "Total Referrals", "Referral Balance (NGN)", "Referral Total Earned (NGN)"
        ]
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill  = header_fill
            cell.font  = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(15, len(h) + 4)

        for row, user in enumerate(get_all_users(), 2):
            stats = user.get("stats", {})
            _ensure_referral_block(user)
            ref = user["referral"]
            ws.cell(row=row, column=1,  value=user.get("user_id", ""))
            ws.cell(row=row, column=2,  value=user.get("username", ""))
            ws.cell(row=row, column=3,  value=user.get("display_name", ""))
            ws.cell(row=row, column=4,  value=user.get("plan", "free").upper())
            ws.cell(row=row, column=5,  value=user.get("plan_expires") or "—")
            ws.cell(row=row, column=6,  value="Yes" if user.get("upgrade_pending") else "No")
            ws.cell(row=row, column=7,  value=user.get("created_at", ""))
            ws.cell(row=row, column=8,  value=stats.get("total_buy_orders", 0))
            ws.cell(row=row, column=9,  value=stats.get("total_sell_orders", 0))
            ws.cell(row=row, column=10, value=stats.get("last_active", ""))
            ws.cell(row=row, column=11, value=ref.get("referred_by") or "—")
            ws.cell(row=row, column=12, value=len(get_referrals_for(user.get("user_id"))))
            ws.cell(row=row, column=13, value=ref.get("balance", 0))
            ws.cell(row=row, column=14, value=ref.get("total_earned", 0))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.error(f"[DB] export_users_to_excel error: {e}")
        return b""


# Initialise on import
_init_dirs()
