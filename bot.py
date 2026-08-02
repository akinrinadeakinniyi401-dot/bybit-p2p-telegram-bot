import asyncio
import random
import re
from concurrent.futures import ThreadPoolExecutor
from functools import partial
import logging
from decimal import Decimal, ROUND_HALF_UP, ROUND_FLOOR, ROUND_CEILING
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputMediaPhoto
from telegram.error import RetryAfter, Forbidden
from telegram.ext import (
    ApplicationBuilder, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters
)
from config import TELEGRAM_TOKEN, ADMIN_IDS
import bybit
from bybit import (
    get_ad_details, get_my_ads, modify_ad,
    get_btc_usdt_price, get_eth_usdt_price, get_token_usdt_price,
    get_max_float_pct, get_min_float_pct, currency_needs_ref,
    get_pending_orders, get_sell_orders, get_incoming_sell_orders, get_order_detail,
    get_counterparty_info, mark_order_paid,
    send_chat_message, get_payment_name, release_assets,
    set_active_account, get_active_account, get_all_accounts,
    get_chat_messages, post_new_ad, remove_ad,
    take_ad_offline, put_ad_online,
    get_user_payment_list,
    review_seller_cancel,
    validate_interval, validate_float_pct, MAX_ADS_PER_USER,
    get_min_price_gap,
)
from fraud_check import check_buyer_name, load_scammers, get_scammer_count, get_last_updated
import db
import subscription as sub
from admin_commands import (
    cmd_upgrade, cmd_downgrade, cmd_requests, cmd_listusers, cmd_userdata,
    cmd_awardref, cmd_addbalance, cmd_deductbalance, cmd_referrals,
    cmd_withdrawals, cmd_approvewithdraw, cmd_rejectwithdraw,
)
import help_agent
import media_downloader as mediadl
from config import REFERRAL_REWARD_NGN, BOT_OWNER_USERNAME, MIN_WITHDRAWAL_NGN, PUBLIC_BASE_URL

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────
# 🖼️ Welcome banner image
# ─────────────────────────────────────────
BANNER_URL = "https://raw.githubusercontent.com/akinrinadeakinniyi401-dot/bybit-p2p-telegram-bot/main/photo_6017280178934975538_x.jpg"


async def _get_current_ip() -> str:
    import requests as _r
    for svc in ["https://api.ipify.org", "https://ifconfig.me/ip"]:
        try:
            return _r.get(svc, timeout=4).text.strip()
        except Exception:
            continue
    return "unknown"


# ─────────────────────────────────────────
# 🧠 Per-user session state — replaces ALL globals
# ─────────────────────────────────────────
# All P2P state (settings, ad_data, orders, toggles, tasks) is now stored
# per user inside a SessionState object from user_session.py.
# Globals below are ONLY kept for:
#   - admin-level Paga queue (shared infra, not per-user state)
#   - _current_user_id / _current_plan_badge (display-only, refreshed per request)
from user_session import get_session, clear_session, get_all_sessions, SessionState

# Dedicated thread pool for ad modification calls (modify_ad via run_in_executor).
# Isolated from the default executor so order/chat monitor threads can never starve
# ad-update threads (and vice versa), which was causing Telegram timeouts under load.
_ad_executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="ad_modify")

# Admin-level Paga queue (shared worker, but each user's jobs are tagged with their uid)
import asyncio as _asyncio
_paga_queue: _asyncio.Queue = None
_paga_worker_task           = None

# Display-only — refreshed at the top of every button/command handler
_current_user_id    = 0
_current_plan_badge = "⚪ Free"

# Legacy admin-scope globals still used by admin-only features (single admin session)
# These are ONLY read/written when is_admin(uid) is True.
user_state: dict = {}   # admin input action state (non-admins use context.user_data)

# ── FLW Transfer Registry ──────────────────────────────────────────────────────
# Maps transfer_ref → {order_id, user_id, slot, amount, pay_term}
# Written at transfer initiation; read by the webhook handler to reconnect the
# webhook event back to the correct Telegram user and Bybit order.
_flw_transfer_registry: dict = {}   # {transfer_ref: {order_id, user_id, slot, amount, pay_term}}

# ── Order Final State Tracker ──────────────────────────────────────────────────
# Prevents re-use of buttons after a terminal action.
# States: "completed", "rejected", "warned", "failed", "expired", "skipped"
_order_final_states: dict = {}      # {(chat_id, order_id): state_str}

# ── Per-order Action Locks ─────────────────────────────────────────────────────
# Prevents concurrent auto-pay + manual tap on the same order.
_order_action_locks: dict = {}      # {(chat_id, order_id): asyncio.Lock}

def _s(uid: int) -> SessionState:
    """Shorthand: get the per-user session for uid."""
    sess = get_session(uid)
    # Ensure per-user slot field exists (backfill for sessions created before this patch)
    if not hasattr(sess, "selected_slot"):
        sess.selected_slot = 0   # 0 = slot 1, 1 = slot 2 (matches bybit._active_index values)
    return sess


def _get_user_slot(uid: int) -> int:
    """Return the active account slot index (0-based) for this specific user."""
    return _s(uid).selected_slot


def _get_user_slot_str(uid: int) -> str:
    """Return slot as 1-based string: '1' or '2'."""
    return str(_s(uid).selected_slot + 1)


# ─────────────────────────────────────────
# Multi-ad slot accessors (up to 3 concurrent ads per user)
# ─────────────────────────────────────────
# NOTE: this "ad slot" (-1 / 0 / 1, meaning Ad 1 / Ad 2 / Ad 3 — up to 3
# ads running concurrently) is a completely different axis from the
# existing "_get_user_slot" (which Bybit ACCOUNT, 1 or 2, is active).
# Ads 2 and 3 always run against whichever account is currently active —
# they share that account's API keys and UID, same as Ad 1 does.
def _valid_slot(sess, slot_idx: int) -> int:
    """Clamp a possibly-stale slot index back to -1 (Ad 1) if it no longer
    exists — e.g. the slot was removed from another tab/click in between."""
    if slot_idx != -1 and slot_idx >= len(sess.extra_ad_slots):
        return -1
    return slot_idx

def _ad_settings(sess, slot_idx: int) -> dict:
    slot_idx = _valid_slot(sess, slot_idx)
    return sess.settings if slot_idx == -1 else sess.extra_ad_slots[slot_idx]["settings"]

def _ad_data_of(sess, slot_idx: int) -> dict:
    slot_idx = _valid_slot(sess, slot_idx)
    return sess.ad_data if slot_idx == -1 else sess.extra_ad_slots[slot_idx]["ad_data"]

def _ad_running(sess, slot_idx: int) -> bool:
    slot_idx = _valid_slot(sess, slot_idx)
    if slot_idx == -1:
        return sess.refresh_running
    return sess.extra_ad_slots[slot_idx]["running"]

def _set_ad_running(sess, slot_idx: int, val: bool):
    slot_idx = _valid_slot(sess, slot_idx)
    if slot_idx == -1:
        sess.refresh_running = val
    else:
        sess.extra_ad_slots[slot_idx]["running"] = val

def _set_ad_task(sess, slot_idx: int, task):
    slot_idx = _valid_slot(sess, slot_idx)
    if slot_idx == -1:
        sess.refresh_task = task
    else:
        sess.extra_ad_slots[slot_idx]["task"] = task

def _ad_current_price(sess, slot_idx: int) -> Decimal:
    slot_idx = _valid_slot(sess, slot_idx)
    return sess.current_price if slot_idx == -1 else sess.extra_ad_slots[slot_idx]["current_price"]

def _set_ad_current_price(sess, slot_idx: int, price):
    slot_idx = _valid_slot(sess, slot_idx)
    if slot_idx == -1:
        sess.current_price = price
    else:
        sess.extra_ad_slots[slot_idx]["current_price"] = price

def _ad_slot_label(slot_idx: int) -> str:
    return "Ad 1" if slot_idx == -1 else f"Ad {slot_idx + 2}"

def _increment_ad_failures(sess, slot_idx: int) -> int:
    slot_idx = _valid_slot(sess, slot_idx)
    if slot_idx == -1:
        sess.consecutive_failures += 1
        return sess.consecutive_failures
    slot = sess.extra_ad_slots[slot_idx]
    slot["consecutive_failures"] += 1
    return slot["consecutive_failures"]

def _reset_ad_failures(sess, slot_idx: int):
    slot_idx = _valid_slot(sess, slot_idx)
    if slot_idx == -1:
        sess.consecutive_failures = 0
    else:
        sess.extra_ad_slots[slot_idx]["consecutive_failures"] = 0


# ─────────────────────────────────────────
# Fast-chase modify budget (Ad 1, single-ad, floating mode only)
# ─────────────────────────────────────────
# Bybit's own limit: "a single advertisement can be modified no more than
# 10 times within 5 minutes." The scheduled cycle already spends part of
# that budget; the fast-chase 10-second price check spends from the SAME
# rolling window rather than its own separate count, so the two together
# can never add up to more than Bybit allows. Capped at 8, not 10, to
# leave headroom for a manual edit on Bybit's own site in the same window.
_FAST_CHASE_BUDGET       = 8
_FAST_CHASE_WINDOW_SECS  = 300

def _can_modify_ad1(sess) -> bool:
    now = datetime.now().timestamp()
    sess.modify_call_times = [t for t in sess.modify_call_times if now - t < _FAST_CHASE_WINDOW_SECS]
    return len(sess.modify_call_times) < _FAST_CHASE_BUDGET

def _record_modify_ad1(sess):
    sess.modify_call_times.append(datetime.now().timestamp())


def _resolve_price_collision(sess, slot_idx: int, currency_id: str, token_id: str, natural_price: Decimal) -> Decimal:
    """
    Multiple ads on the SAME (currency, coin) pair are allowed to use the
    same floating % (or fixed base) — Bybit doesn't reject on the % or the
    starting price, it rejects when the actual POSTED prices land too
    close together. So: if this ad's naturally-computed price is within
    get_min_price_gap() of another active ad's currently-known price on
    the same pair, step this one down below the lowest conflicting price
    by that gap — e.g. Ad 1 posts ₦84,908,465.23, Ad 2 (same pair, same
    float) would naturally compute the same number, so it gets pushed to
    ₦84,903,465.23 or lower instead.

    Fixed-amount gap for BTC/ETH pairs (₦5,000 / $100 etc — see
    MIN_PRICE_GAP in bybit.py). For USDT/USDC specifically, the gap is 1%
    of the actual price instead, since a flat amount would be the wrong
    scale for a stablecoin price (see get_min_price_gap in bybit.py).

    This only ever looks at OTHER slots — never changes which ad "wins"
    the natural price, it just moves the others out of the way. Resolves
    iteratively so a 3-way collision clears every conflict, not just the
    nearest one.
    """
    gap = get_min_price_gap(currency_id, token_id, natural_price)
    conflicting_prices = []
    for i in range(-1, len(sess.extra_ad_slots)):
        if i == slot_idx:
            continue
        other_ad_data = _ad_data_of(sess, i)
        if not other_ad_data:
            continue
        if (other_ad_data.get("currencyId","").upper() != currency_id.upper()
                or other_ad_data.get("tokenId","").upper() != token_id.upper()):
            continue
        other_price = _ad_current_price(sess, i)
        if other_price and other_price > 0:
            conflicting_prices.append(other_price)

    # Also guard against landing back on THIS ad's own currently-live price.
    # Bybit rejects a submission with retCode 90043 ("price differs from
    # your existing ad by less than 0%") whenever the new price rounds to
    # what this SAME ad already has posted — that's a per-ad comparison
    # Bybit makes internally, nothing to do with any other ad. Previously
    # this function only ever checked OTHER ads, so when the underlying
    # market hadn't moved between cycles, a collision nudge could push this
    # ad's price right back onto its own last-posted value and get
    # rejected — which is exactly what showed up as repeated 90043 failures
    # once a user had 2+ ads on the same pair/float %.
    own_price = _ad_current_price(sess, slot_idx)
    if own_price and own_price > 0:
        conflicting_prices.append(own_price)

    price = natural_price
    for _ in range(10):   # safety cap — converges in 1-2 passes in practice
        adjusted = False
        for cp in conflicting_prices:
            if abs(price - cp) < gap:
                price = cp - gap
                adjusted = True
        if not adjusted:
            break
    return price


def _settings(uid: int) -> dict:
    """Shorthand: get the mutable settings dict for uid."""
    return get_session(uid).settings

def _save_settings(uid: int):
    """Persist the user's current session settings to disk."""
    db.save_settings(uid, get_session(uid).settings)

def _load_settings_from_disk(uid: int):
    """Load persisted settings from disk into the user's session on first access.

    Also back-populates slot-keyed keys from generic keys (and vice versa)
    so that both ad_id_1/bybit_uid_1 and ad_id/bybit_uid are always in sync.
    This ensures UID and Ad ID survive /start, restarts, and slot switches.

    AD BOT settings (mode, increment, float_pct, local_usdt_ref, interval) are
    also stored per-slot and restored here for the user's active slot.
    """
    saved = db.load_settings(uid)
    if saved:
        sess = get_session(uid)
        for k, v in saved.items():
            sess.settings[k] = v
        # Back-fill: if only generic key exists, populate slot-keyed keys (slots 1 & 2)
        for field in ("ad_id", "bybit_uid"):
            generic_val = sess.settings.get(field, "")
            for slot in ("1", "2"):
                slot_key = f"{field}_{slot}"
                if not sess.settings.get(slot_key) and generic_val:
                    sess.settings[slot_key] = generic_val
        # Back-fill: if only slot-keyed keys exist, populate generic key from slot 1
        for field in ("ad_id", "bybit_uid"):
            if not sess.settings.get(field):
                slot1_val = sess.settings.get(f"{field}_1", "")
                if slot1_val:
                    sess.settings[field] = slot1_val

        # ── Restore active slot's AD BOT settings into generic keys ──
        # This ensures the correct slot's config is active after /start or restart.
        # Ensure selected_slot is set first (default to 0)
        if not hasattr(sess, "selected_slot"):
            sess.selected_slot = 0
        active_slot_str = str(sess.selected_slot + 1)
        for field, default in [("mode", "fixed"), ("increment", "0.05"),
                                ("float_pct", ""), ("local_usdt_ref", ""), ("interval", 2)]:
            slot_val = sess.settings.get(f"{field}_{active_slot_str}")
            if slot_val is not None and slot_val != "":
                sess.settings[field] = slot_val
            elif not sess.settings.get(field):
                sess.settings[field] = default

        logger.debug(f"[Settings] Loaded for user={uid}: ad_id={sess.settings.get('ad_id')!r} "
                     f"bybit_uid={sess.settings.get('bybit_uid')!r} "
                     f"ad_id_1={sess.settings.get('ad_id_1')!r} bybit_uid_1={sess.settings.get('bybit_uid_1')!r} "
                     f"mode={sess.settings.get('mode')!r} slot={active_slot_str}")

SELLER_WARN_MSG = (
    "Dear seller, your average release time is too long, I can't proceed with the payment. "
    "Kindly check your order page at the top right corner to request cancel. Thank you"
)

NO_ACCOUNT_WARN_MSG = (
    "Dear seller, your payment details (account name / account number) are incomplete. "
    "Kindly request a cancel on this order. Thank you."
)

def is_admin(uid): return uid in ADMIN_IDS

def _get_or_register_user(telegram_user):
    """Register user in DB on first access and update last_active. Returns (user_dict, is_new)."""
    uid   = telegram_user.id
    uname = telegram_user.username or ""
    dname = telegram_user.full_name or ""
    # get_or_create_user already writes last_active on every call (db.py line 160)
    return db.get_or_create_user(uid, uname, dname)

# Pre-populate admin chat IDs from environment config so upgrade notifications
# work even before the admin has sent /start in this deploy session.
_admin_chat_ids: set = set(ADMIN_IDS)  # seeded from config; updated on /start

def _get_admin_chat_ids() -> set:
    return _admin_chat_ids


# ─────────────────────────────────────────
# 📊 Setup progress checker (per-user)
# ─────────────────────────────────────────
def setup_progress(uid: int) -> tuple:
    s     = _settings(uid)
    sess  = _s(uid)
    slot  = _get_user_slot_str(uid)   # per-user slot — NOT global
    steps = [
        bool(s.get(f"ad_id_{slot}") or s.get("ad_id")),
        bool(s.get(f"bybit_uid_{slot}") or s.get("bybit_uid")),
        bool(sess.ad_data),
        bool(s.get("increment") or s.get("float_pct")),
        bool(s.get("interval")),
    ]
    done  = sum(steps)
    total = len(steps)
    bar   = "".join("✅" if s else "⬜" for s in steps)
    return done, total, bar


def next_setup_hint(uid: int) -> str:
    s    = _settings(uid)
    sess = _s(uid)
    slot = _get_user_slot_str(uid)   # per-user slot — NOT global
    ad_id    = s.get(f"ad_id_{slot}") or s.get("ad_id","")
    bybit_uid = s.get(f"bybit_uid_{slot}") or s.get("bybit_uid","")
    if not ad_id:
        return "👉 Start by tapping *🆔 Set Ad ID*"
    if not bybit_uid:
        return "👉 Next: tap *👤 Set UID* to set your Bybit user ID"
    if not sess.ad_data:
        return "👉 Next: tap *📋 Fetch Ad Details* to load your ad from Bybit"
    mode = s.get("mode", "fixed")
    if mode == "fixed" and not s.get("increment"):
        return "👉 Next: tap *➕ Set Increment* to set your price step"
    if mode == "floating" and not s.get("float_pct"):
        return "👉 Next: tap *📊 Set Float %* to set your market percentage"
    currency_upper = sess.ad_data.get("currencyId","").upper()
    needs_ref_cur  = currency_needs_ref(currency_upper) or currency_upper == "NGN"
    if mode == "floating" and needs_ref_cur and not s.get("local_usdt_ref"):
        return f"👉 Next: tap *💱 Set {currency_upper}/USDT Ref* to set the reference rate"
    return "✅ *All set!* Tap *🟢 Start Auto-Update* to begin"


# ─────────────────────────────────────────
# 🔑 Per-user credential helper
# ─────────────────────────────────────────
def get_user_creds(user_id: int, slot: int | None = None) -> dict | None:
    """
    Load Bybit credentials for a user using THEIR OWN per-user slot from DB.

    CRITICAL: Uses _s(user_id).selected_slot — NOT the global bybit._active_index.
    This ensures User A switching slots never affects User B.

    ALL users — including admins — now load from DB first.
    Admins fall back to env account ONLY if no DB key is saved for their slot,
    so the bot works with or without Render env keys.

    Args:
        user_id: Telegram user ID
        slot: Optional override (0-based index). If None, uses user's own selected_slot.

    Return values:
      - User/admin w/ DB key  → {"key": ..., "secret": ...}
      - Admin w/ no DB key    → None  (bybit._resolve_creds(None) uses env account if set)
      - Non-admin no DB key   → {"key": "", "secret": ""}  ← SENTINEL: no key saved
    """
    user_slot = slot if slot is not None else _get_user_slot(user_id)
    slot_str  = str(user_slot + 1)   # "1" or "2"

    key    = db.get_api(user_id, f"bybit_key_{slot_str}")
    secret = db.get_api(user_id, f"bybit_secret_{slot_str}")
    if key and secret:
        logger.debug(f"[Creds] User {user_id} slot {slot_str} — DB key found")
        return {"key": key, "secret": secret}

    # No DB key for this user/slot
    if is_admin(user_id):
        # Admin fallback: use env account (may also be empty if no env keys set)
        logger.info(f"[Creds] Admin {user_id} slot {slot_str} — no DB key, falling back to env account")
        return None   # bybit._resolve_creds(None) uses BYBIT_ACCOUNTS[_active_index] if available

    # Non-admin: return sentinel (empty strings) — callers show "No API set" error
    logger.info(f"[Creds] User {user_id} slot {slot_str} — NO API KEY SAVED")
    return {"key": "", "secret": ""}


# ─────────────────────────────────────────
# 🏠 MAIN MENU
# ─────────────────────────────────────────
def main_menu_keyboard(uid: int = 0):
    sess   = _s(uid) if uid else None
    o_icon = "🔔" if (sess and sess.order_monitor_running) else "🔕"
    p_icon = "💳✅" if (sess and (sess.auto_pay_enabled or sess.flw_pay_enabled)) else "💳"
    r_icon = "🟢" if (sess and sess.refresh_running) else "📊"
    all_ac = get_all_accounts()

    # ── Account slot buttons ──
    # Always show 2 slots regardless of whether env keys are set.
    # Labels come from BYBIT_ACCOUNTS if available; otherwise use "Account N" fallback.
    # This ensures the switcher is visible even in pure multi-user (no env keys) mode.
    _user_slot_idx = _s(uid).selected_slot if uid else 0
    _num_slots = max(len(all_ac), 2)   # always at least 2 slots
    _slot_row = []
    for i in range(_num_slots):
        label = all_ac[i]["label"] if i < len(all_ac) else f"Account {i + 1}"
        tick  = "✅ " if i == _user_slot_idx else ""
        _slot_row.append(InlineKeyboardButton(f"{tick}{label}", callback_data=f"switch_account_{i}"))
    kb = [_slot_row]

    kb += [
        [InlineKeyboardButton(f"{r_icon} AD PRICE BOT",  callback_data="section_ads"),
         InlineKeyboardButton(f"{o_icon} ORDER MONITOR", callback_data="section_orders")],
        [InlineKeyboardButton(f"{p_icon} AUTO-PAY",      callback_data="section_autopay"),
         InlineKeyboardButton("🔑 Set APIs",             callback_data="section_apis")],
        [InlineKeyboardButton("⬆️ Upgrade Plan",         callback_data="upgrade_plan"),
         InlineKeyboardButton("🎁 Referrals",            callback_data="referrals")],
        [InlineKeyboardButton("🎬 Video Downloader",     callback_data="video_downloader"),
         InlineKeyboardButton("💬 Contact Support",      callback_data="contact_support")],
        [InlineKeyboardButton("📡 Bot Status",           callback_data="bot_status"),
         InlineKeyboardButton("🌍 Get My IP",            callback_data="get_my_ip")],
        [InlineKeyboardButton("🔁 Reset Session",        callback_data="reset_confirm")],
    ]
    return InlineKeyboardMarkup(kb)


def main_menu_text(uid: int = 0) -> str:
    uid      = uid or _current_user_id
    sess     = _s(uid)
    done, total, bar = setup_progress(uid)
    o_status = "🔔 Active"  if sess.order_monitor_running else "🔕 Off"
    p_status = "💳 ON"      if sess.auto_pay_enabled       else "💳 OFF"
    r_status = "🟢 Running" if sess.refresh_running        else "🔴 Off"
    # Per-user active account — NOT global
    _uid_slot = _s(uid).selected_slot
    _all_ac   = get_all_accounts()
    if _all_ac and _uid_slot < len(_all_ac):
        acct = _all_ac[_uid_slot]
    elif _all_ac:
        acct = _all_ac[0]
    else:
        acct = {"label": f"Account {_uid_slot + 1}"}
    bp_status = f"🛡 ON ({sess.buyer_protection_mins}min)" if sess.buyer_protection_on else "🛡 OFF"
    nm_status = "🔍 ON"     if sess.name_match_enabled     else "🔍 OFF"
    badge     = sub.plan_badge(uid) if uid else _current_plan_badge

    return (
        "🤖 *P2P Auto Bot — Control Panel*\n\n"
        f"🆔 Your ID: <code>{uid}</code> | {badge}\n"
        f"🔑 Active Account: <b>{acct['label']}</b>\n"
        f"📋 Setup: {bar} <code>{done}/{total}</code>\n\n"
        f"┌ 📊 Price Bot: {r_status}\n"
        f"├ 📦 Orders: {o_status}\n"
        f"├ 💳 Auto-Pay: {p_status}\n"
        f"├ {bp_status} Buyer Protection\n"
        f"└ {nm_status} Name Match\n\n"
        "_Select a section below to get started:_"
    )


def back_main():
    return [[InlineKeyboardButton("🏠 Main Menu", callback_data="main_menu")]]


def back_section(section: str):
    labels = {
        "section_ads":     "📊 AD Price Bot",
        "section_orders":  "📦 Order Monitor",
        "section_autopay": "💳 Auto-Pay",
    }
    return [[InlineKeyboardButton(f"⬅️ Back — {labels.get(section,'Back')}", callback_data=section)]]


def back_manager():
    """Back button that returns to the Post/Remove Ad Manager."""
    return [[InlineKeyboardButton("⬅️ Back — 📢 Post/Remove Manager", callback_data="post_ad_prompt")]]


def back_prev(prev: str):
    """Back to previous section button — used after text input success."""
    labels = {
        "section_ads":     "📊 AD Price Bot",
        "section_orders":  "📦 Order Monitor",
        "section_autopay": "💳 Auto-Pay",
        "main_menu":       "🏠 Main Menu",
        "post_ad_prompt":  "📢 Post/Remove Manager",
    }
    label = labels.get(prev, "⬅️ Back")
    return InlineKeyboardMarkup([[InlineKeyboardButton(f"⬅️ Back to {label}", callback_data=prev)]])


# ─────────────────────────────────────────
# 📊 AD PRICE BOT SECTION
# ─────────────────────────────────────────
def ads_section_keyboard(uid: int = 0):
    sess       = _s(uid) if uid else None
    slot_idx   = sess.editing_slot if sess else -1
    s          = _ad_settings(sess, slot_idx) if sess else {}
    ad_data    = _ad_data_of(sess, slot_idx) if sess else {}
    mode       = s.get("mode", "fixed")
    mode_icon  = "💲" if mode == "fixed" else "📈"
    mode_label = f"{mode_icon} Mode: {mode.upper()}"
    ad_loaded  = bool(ad_data)
    running    = _ad_running(sess, slot_idx) if sess else False
    status     = "🟢 Stop Auto-Update" if running else "▶️ Start Auto-Update"
    total_ads  = sess.total_ad_slots() if sess else 1

    rows = []

    # ── Ad slot switcher — only shown once the user has more than 1 ad ──
    if total_ads > 1:
        switch_row = []
        for i in range(-1, total_ads - 1):
            label = _ad_slot_label(i)
            if i == slot_idx:
                label = f"• {label} •"
            switch_row.append(InlineKeyboardButton(label, callback_data=f"edit_ad_{i+2}"))
        rows.append(switch_row)
        rows.append([InlineKeyboardButton("🗂 All My Ads (Dashboard)", callback_data="ads_dashboard")])

    if slot_idx == -1:
        rows.append([
            InlineKeyboardButton("🆔 Set Ad ID",    callback_data="set_ad_id"),
            InlineKeyboardButton("👤 Set UID",      callback_data="set_uid"),
            InlineKeyboardButton("🗑 Del UID",      callback_data="delete_uid"),
        ])
    else:
        # Ads 2/3 share Ad 1's Bybit account + UID — only the Ad ID differs.
        rows.append([InlineKeyboardButton("🆔 Set Ad ID", callback_data="set_ad_id")])

    rows.append([
        InlineKeyboardButton("📋 Fetch Ad Details", callback_data="fetch_ad"),
        InlineKeyboardButton("📃 My Ads List",      callback_data="fetch_my_ads"),
    ])
    rows.append([
        InlineKeyboardButton(mode_label,        callback_data="switch_mode"),
        InlineKeyboardButton("⏱ Set Interval", callback_data="set_interval"),
    ])

    if mode == "fixed":
        rows.append([InlineKeyboardButton("➕ Set Increment", callback_data="set_increment")])
    else:
        rows.append([InlineKeyboardButton("📊 Set Float %",   callback_data="set_float_pct")])
        _cur = ad_data.get("currencyId","").upper()
        if currency_needs_ref(_cur) or _cur == "NGN":
            rows.append([InlineKeyboardButton(f"💱 Set {_cur}/USDT Ref", callback_data="set_ngn_ref")])

    # Update Once Now only makes sense — and is only offered — when the
    # user is running a single ad. With multiple ads active, a one-off
    # manual update on one of them can no longer be validated against the
    # others' prices at that exact instant, so it's disabled entirely.
    if ad_loaded and total_ads == 1:
        rows.append([InlineKeyboardButton("🔄 Update Once Now", callback_data="update_now")])

    if slot_idx == -1:
        rows.append([InlineKeyboardButton("📢 Post / Remove Ad", callback_data="post_ad_prompt")])

    rows.append([InlineKeyboardButton(status, callback_data="toggle_refresh")])

    if total_ads < MAX_ADS_PER_USER:
        rows.append([InlineKeyboardButton(f"➕ Add {_ad_slot_label(total_ads - 1)}", callback_data="add_ad_slot")])
    if slot_idx != -1:
        rows.append([InlineKeyboardButton(f"🗑 Remove {_ad_slot_label(slot_idx)}", callback_data="remove_ad_slot")])

    rows += back_main()
    return InlineKeyboardMarkup(rows)


def ads_section_text(uid: int = 0) -> str:
    uid      = uid or _current_user_id
    sess     = _s(uid)
    slot_idx = sess.editing_slot
    s        = _ad_settings(sess, slot_idx)
    ad_data  = _ad_data_of(sess, slot_idx)
    acct_slot = _get_user_slot_str(uid)   # per-user BYBIT ACCOUNT slot — NOT the ad slot

    if slot_idx == -1:
        ad_id     = s.get(f"ad_id_{acct_slot}") or s.get("ad_id","") or "❗ Not set"
        bybit_uid = s.get(f"bybit_uid_{acct_slot}") or s.get("bybit_uid","") or "❗ Not set"
    else:
        ad_id     = s.get("ad_id","") or "❗ Not set"
        # Ads 2/3 always use the same UID as the active account (Ad 1's UID).
        bybit_uid = (
            sess.settings.get(f"bybit_uid_{acct_slot}")
            or sess.settings.get("bybit_uid","")
            or "❗ Not set"
        )

    mode      = s.get("mode",           "fixed")
    interval  = s.get("interval",       2)
    increment = s.get("increment",      "0.05")
    float_pct = s.get("float_pct",     "") or "❗ Not set"
    local_ref = sess.shared_local_usdt_ref or s.get("local_usdt_ref","") or "❗ Not set"
    cur_label = ad_data.get("currencyId","NGN").upper() if ad_data else "NGN"
    cur       = str(_ad_current_price(sess, slot_idx)) if _ad_current_price(sess, slot_idx) else "—"
    status    = "🟢 Running" if _ad_running(sess, slot_idx) else "🔴 Stopped"

    if ad_data:
        price    = ad_data.get("price",        "—")
        min_amt  = ad_data.get("minAmount",    "—")
        max_amt  = ad_data.get("maxAmount",    "—")
        qty      = ad_data.get("lastQuantity", ad_data.get("quantity","—"))
        token    = ad_data.get("tokenId",      "—")
        currency = ad_data.get("currencyId",   "—")
        ad_stat  = {10:"🟢 Online",20:"🔴 Offline",30:"✅ Done"}.get(ad_data.get("status"),"?")
        max_pct  = get_max_float_pct(currency, token)
        ad_info  = (
            f"\n📋 <b>Loaded Ad:</b>\n"
            f"  💱 <code>{token}/{currency}</code> | 💲 <code>{price}</code>\n"
            f"  Min: <code>{min_amt}</code> | Max: <code>{max_amt}</code> | Qty: <code>{qty}</code>\n"
            f"  Status: {ad_stat} | Max float: <code>{max_pct}%</code>\n"
        )
    else:
        ad_info = "\n  ⚠️ No ad loaded yet\n"

    if mode == "fixed":
        mode_info = f"  ➕ Increment: `+{increment}` per cycle"
    else:
        mode_info = f"  📊 Float: `{float_pct}%`"
        if ad_data.get("currencyId","").upper() == "NGN":
            mode_info += f" | 💱 {cur_label}/USDT: `{local_ref}`"

    hint = next_setup_hint(uid) if slot_idx == -1 else "Set this ad's Ad ID, fetch its details, then set its mode."
    user_slot_idx = _get_user_slot(uid)
    acct_label = bybit.BYBIT_ACCOUNTS[user_slot_idx]["label"] if (bybit.BYBIT_ACCOUNTS and user_slot_idx < len(bybit.BYBIT_ACCOUNTS)) else f"Account {acct_slot}"
    slot_header = _ad_slot_label(slot_idx)
    multi_note = f" ({sess.total_ad_slots()} ads active)" if sess.total_ad_slots() > 1 else ""

    return (
        f"📊 <b>AD PRICE BOT — {slot_header}{multi_note}</b>\n"
        f"<i>{acct_label}</i>\n\n"
        f"🆔 Ad ID: <code>{ad_id}</code>\n"
        f"👤 UID (Acct {acct_slot}): <code>{bybit_uid}</code>\n"
        f"🔀 Mode: <code>{mode.upper()}</code> | ⏱ Every <code>{interval}</code> min\n"
        f"{mode_info}\n"
        f"{ad_info}\n"
        f"📈 Session price: <code>{cur}</code> | {status}\n\n"
        f"<i>{hint}</i>"
    )


def ads_dashboard_text(uid: int) -> str:
    """All-ads-at-a-glance view — coin/pair + status per ad, per user's request."""
    sess = _s(uid)
    lines = [f"🗂 <b>All My Ads ({sess.total_ad_slots()} active)</b>\n"]
    for i in range(-1, len(sess.extra_ad_slots)):
        ad_data = _ad_data_of(sess, i)
        running = _ad_running(sess, i)
        s       = _ad_settings(sess, i)
        icon    = "🟢" if running else "🔴"
        if ad_data:
            pair = f"{ad_data.get('tokenId','?')}/{ad_data.get('currencyId','?')}"
            price = ad_data.get("price", "—")
        else:
            pair, price = "not loaded yet", "—"
        mode = s.get("mode", "fixed").upper()
        lines.append(f"{icon} <b>{_ad_slot_label(i)}</b> — {pair} | 💲{price} | {mode} | {'Running' if running else 'Stopped'}")
    return "\n".join(lines)


def ads_dashboard_keyboard(uid: int) -> InlineKeyboardMarkup:
    sess = _s(uid)
    rows = []
    stop_row = []
    for i in range(-1, len(sess.extra_ad_slots)):
        if _ad_running(sess, i):
            stop_row.append(InlineKeyboardButton(f"⏹ Stop {_ad_slot_label(i)}", callback_data=f"stop_ad_{i+2}"))
    if stop_row:
        rows.append(stop_row)
    if sum(1 for i in range(-1, len(sess.extra_ad_slots)) if _ad_running(sess, i)) > 1:
        rows.append([InlineKeyboardButton("⏹ Stop All Ads", callback_data="stop_all_ads")])
    edit_row = [InlineKeyboardButton(f"✏️ Edit {_ad_slot_label(i)}", callback_data=f"edit_ad_{i+2}") for i in range(-1, len(sess.extra_ad_slots))]
    rows.append(edit_row)
    rows += back_main()
    return InlineKeyboardMarkup(rows)




# ─────────────────────────────────────────
# 📦 ORDER MONITOR SECTION
# ─────────────────────────────────────────
def orders_section_keyboard(uid: int = 0):
    sess     = _s(uid) if uid else None
    mon      = "🔔 Stop Monitoring" if (sess and sess.order_monitor_running) else "🔕 Start Monitoring"
    sell_tog = "✉️ Sell Msg: ON — tap to OFF" if (sess and sess.sell_msg_enabled) else "✉️ Sell Msg: OFF — tap to ON"
    chat_tog = "💬 Chat Monitor: ON ✅ — tap to OFF" if (sess and sess.chat_monitor_enabled) else "💬 Chat Monitor: OFF ❌ — tap to ON"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(mon,                        callback_data="toggle_order_monitor")],
        [InlineKeyboardButton(chat_tog,                   callback_data="toggle_chat_monitor")],
        [InlineKeyboardButton("📋 Check Orders Now",      callback_data="check_orders_now")],
        [InlineKeyboardButton("🗑 Clear Seen Orders",     callback_data="clear_seen_orders")],
        [InlineKeyboardButton(sell_tog,                   callback_data="toggle_sell_msg")],
        [InlineKeyboardButton("✏️ Set Sell Message",      callback_data="set_sell_msg")],
        [InlineKeyboardButton("🔢 Set Message Count",     callback_data="set_sell_msg_count")],
        *back_main()
    ])


def orders_section_text(uid: int = 0) -> str:
    uid  = uid or _current_user_id
    sess = _s(uid)
    status    = "🔔 Active — checking every 10 sec" if sess.order_monitor_running else "🔕 Stopped"
    seen_buy  = len(sess.seen_order_ids)
    seen_sell = len(sess.seen_sell_ids)
    paid      = len(sess.paid_order_ids)
    released  = len(sess.released_ids)
    ap_status = "💳 ON — auto marking orders paid" if sess.auto_pay_enabled else "💳 OFF — manual only"
    sm_status = f"✅ ON — sending {sess.sell_msg_count}x per order" if sess.sell_msg_enabled else "❌ OFF"
    msg_preview = sess.sell_custom_msg[:60] + "..." if len(sess.sell_custom_msg) > 60 else sess.sell_custom_msg
    chat_status = "💬 ON — forwarding messages every 8s" if sess.chat_monitor_enabled else "💬 OFF"
    return (
        "📦 *ORDER MONITOR*\n\n"
        f"Status: {status}\n"
        f"BUY orders seen: <code>{seen_buy}</code> | Marked paid: <code>{paid}</code>\n"
        f"SELL orders seen: <code>{seen_sell}</code> | Released: <code>{released}</code>\n\n"
        f"Auto-Pay (BUY): {ap_status}\n\n"
        f"💬 <b>Chat Monitor:</b> {chat_status}\n\n"
        f"✉️ <b>Sell Order Message: {sm_status}</b>\n"
        f"Message (<code>{sess.sell_msg_count}x</code>): _{msg_preview}_\n\n"
        "_BUY orders → Mark as Paid buttons_\n"
        "_SELL orders → Release Coin button_\n"
        "_Both show seller/buyer info + payment details_"
    )


# ─────────────────────────────────────────
# 💳 AUTO-PAY SECTION
# ─────────────────────────────────────────
def autopay_section_keyboard(uid: int = 0):
    sess    = _s(uid) if uid else None
    pay     = "💳 Disable Auto-Pay (Bybit)" if (sess and sess.auto_pay_enabled)  else "💳 Enable Auto-Pay (Bybit)"
    flw     = "🟢 Disable Flutterwave Pay ✅" if (sess and sess.flw_pay_enabled) else "🔴 Enable Flutterwave Pay"
    paga    = "🟡 Disable Paga Pay ✅" if (sess and sess.paga_pay_enabled) else "🟡 Enable Paga Pay"
    bp_tog  = f"🛡 Buyer Protection: {'ON ✅' if (sess and sess.buyer_protection_on) else 'OFF ❌'}"
    nm_tog  = f"🔍 Name Match: {'ON ✅' if (sess and sess.name_match_enabled) else 'OFF ❌'}"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(pay,  callback_data="toggle_auto_pay")],
        [InlineKeyboardButton(flw,  callback_data="toggle_flw_pay")],
        [InlineKeyboardButton(paga, callback_data="toggle_paga_pay")],
        [InlineKeyboardButton("✏️ Set My Sender Name",        callback_data="set_sender_name")],
        [InlineKeyboardButton("🛡 Buyer Protection Settings", callback_data="buyer_protection_menu")],
        [InlineKeyboardButton(bp_tog,                         callback_data="toggle_buyer_protection")],
        [InlineKeyboardButton(nm_tog,                         callback_data="toggle_name_match")],
        [InlineKeyboardButton("📋 View Unpaid Orders",        callback_data="view_unpaid_orders")],
        [InlineKeyboardButton("📊 Buy Volume (24h)",          callback_data="view_buy_volume")],
        [InlineKeyboardButton("ℹ️ How Auto-Pay Works",        callback_data="autopay_info")],
        [InlineKeyboardButton("ℹ️ How Flutterwave Pay Works", callback_data="flw_info")],
        [InlineKeyboardButton("ℹ️ How Paga Pay Works",        callback_data="paga_info")],
        *back_main()
    ])


def autopay_section_text(uid: int = 0) -> str:
    uid  = uid or _current_user_id
    sess = _s(uid)
    bybit_status = "✅ ENABLED" if sess.auto_pay_enabled  else "❌ DISABLED"
    flw_status   = "✅ ENABLED" if sess.flw_pay_enabled   else "❌ DISABLED"
    paga_status  = "✅ ENABLED" if sess.paga_pay_enabled  else "❌ DISABLED"
    # All API keys are per-user — stored in DB only
    # FLW only needs: PUBLIC_KEY, SECRET_HASH, SECRET_KEY (3 keys)
    flw_fully_set = all(db.get_api(uid, k) for k in (
        "flw_public_key", "flw_secret_hash", "flw_secret_key"
    ))
    paga_key  = db.get_api(uid, "paga_principal")
    flw_configured  = "✅ Configured (3/3 keys)" if flw_fully_set else "❌ Not configured"
    paga_configured = "✅ Configured" if paga_key else "❌ Not configured"
    sender_name  = sess.settings.get("sender_name", "Not set")
    unpaid_count = len(sess.unpaid_log)
    bp_status    = f"✅ ON — threshold: {sess.buyer_protection_mins} min" if sess.buyer_protection_on else "❌ OFF"
    nm_status    = "✅ ON — skips orders with missing account info" if sess.name_match_enabled else "❌ OFF"
    bv_lines     = sess.get_buy_volume_lines()
    bv_summary   = ", ".join(l.split("  (")[0] for l in bv_lines) if bv_lines else "No buy orders yet"
    return (
        f"💳 <b>AUTO-PAY</b>\n\n"
        f"Bybit Mark-Paid: <b>{bybit_status}</b>\n"
        f"Flutterwave Pay: <b>{flw_status}</b>\n"
        f"Paga Pay: <b>{paga_status}</b>\n\n"
        f"Flutterwave: {flw_configured}\n"
        f"Paga: {paga_configured}\n"
        f"✏️ Sender name: <code>{sender_name}</code>\n"
        f"📋 Unpaid orders this session: <code>{unpaid_count}</code>\n"
        f"📊 Buy volume (24h): <code>{bv_summary}</code>\n\n"
        f"🛡 <b>Buyer Protection:</b> {bp_status}\n"
        f"🔍 <b>Name Match:</b> {nm_status}\n\n"
        "⚠️ Enable only ONE of Bybit or Flutterwave at a time.\n"
        "Bybit marks the order paid without sending money.\n"
        "Flutterwave actually sends the money then marks paid.\n\n"
        "ℹ️ FLW Auto-Pay falls back to Bybit mark-paid + warning\n"
        "   if seller release time exceeds the Buyer Protection threshold."
    )


# ─────────────────────────────────────────
# 🛡 BUYER PROTECTION MENU
# ─────────────────────────────────────────
def buyer_protection_menu_keyboard(uid: int = 0):
    sess   = _s(uid) if uid else None
    bp_tog = f"🛡 Buyer Protection: {'ON ✅ — tap to OFF' if (sess and sess.buyer_protection_on) else 'OFF ❌ — tap to ON'}"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏱ 10 min", callback_data="bp_set_10"),
         InlineKeyboardButton("⏱ 15 min", callback_data="bp_set_15")],
        [InlineKeyboardButton("⏱ 20 min", callback_data="bp_set_20"),
         InlineKeyboardButton("⏱ 30 min", callback_data="bp_set_30")],
        [InlineKeyboardButton("✏️ Custom minutes", callback_data="bp_set_custom")],
        [InlineKeyboardButton(bp_tog, callback_data="toggle_buyer_protection")],
        *back_section("section_autopay"),
    ])


def buyer_protection_menu_text(uid: int = 0):
    sess   = _s(uid) if uid else None
    thresh = sess.buyer_protection_mins if sess else 30
    on     = bool(sess and sess.buyer_protection_on)
    status = f"✅ ON — threshold: *{thresh} min*" if on else "❌ OFF"
    return (
        "🛡 *Buyer Protection*\n\n"
        f"Current status: {status}\n\n"
        "When enabled, if a seller's average release time is at or above "
        "your chosen threshold, the bot will:\n\n"
        "  1️⃣ Mark the order as paid on Bybit\n"
        "  2️⃣ Send a warning message to the seller\n"
        "  3️⃣ Skip Flutterwave transfer (if FLW Pay is active)\n\n"
        f"⏱ <b>Choose your threshold time:</b>\n"
        f"  Current: <code>{thresh} min</code>\n\n"
        "_Tap a time button below or enter a custom value:_"
    )


# ─────────────────────────────────────────
# 💳 Payment helpers
# ─────────────────────────────────────────

# Per-user payment method name cache: {user_id: {paymentType_str: paymentName_str}}
# Populated lazily on first use; survives the session but is lightweight (just strings).
_payment_name_cache: dict = {}


def _fetch_user_payment_map(creds: dict) -> dict:
    """
    Call POST /v5/p2p/user/payment/list and return a dict mapping
    paymentType (str) -> paymentName (str).
    Falls back to empty dict on any error — callers always have the
    static PAYMENT_TYPE_MAP in bybit.get_payment_name() as last resort.
    """
    try:
        res = get_user_payment_list(creds=creds)
        if res.get("retCode", -1) != 0:
            return {}
        items = res.get("result", {})
        # result may be a list directly or wrapped in a key
        if isinstance(items, dict):
            items = items.get("paymentConfigVoList", items.get("list", []))
        if not isinstance(items, list):
            return {}
        mapping = {}
        for item in items:
            ptype = str(item.get("paymentType", "")).strip()
            pname = (
                item.get("paymentConfigVo", {}).get("paymentName", "").strip()
                or item.get("paymentName", "").strip()
            )
            if ptype and pname:
                mapping[ptype] = pname
        return mapping
    except Exception as e:
        logger.debug(f"[PaymentMap] Could not fetch user payment list: {e}")
        return {}


def _get_payment_name_for_user(ptype: str, uid: int) -> str:
    """
    Resolve a paymentType code to a human-readable name.
    Priority: user's own payment list cache → static map → 'Type XXX' fallback.
    """
    if not ptype:
        return "—"
    ptype_str = str(ptype)
    # 1. User's own cached map (most accurate — reflects their actual payment methods)
    user_map = _payment_name_cache.get(uid, {})
    if ptype_str in user_map:
        return user_map[ptype_str]
    # 2. Static built-in map in bybit.py
    static = get_payment_name(ptype_str)
    if not static.startswith("Type "):
        return static
    # 3. Fallback
    return f"Type {ptype_str}"


def _resolve_pay_term(order_detail: dict) -> dict:
    """
    Return the best available pay_term dict from an order.
    Merges confirmedPayTerm + paymentTermList[0] so missing fields in one
    can be filled from the other.
    """
    confirmed = order_detail.get("confirmedPayTerm", {}) or {}
    terms     = order_detail.get("paymentTermList",   []) or []
    fallback  = terms[0] if terms else {}
    if not confirmed:
        return dict(fallback)
    if not fallback:
        return dict(confirmed)
    # Merge: confirmed takes priority; fallback fills any blank fields
    merged = dict(fallback)
    merged.update({k: v for k, v in confirmed.items() if v not in (None, "", {}, [])})
    return merged


def _get_pay_name(pay_term: dict, uid: int = 0) -> str:
    """Resolve payment method name from a pay_term dict."""
    # 1. paymentConfig.paymentName (richest source — comes from Bybit's own config)
    cfg      = pay_term.get("paymentConfig", {}) or {}
    cfg_name = cfg.get("paymentName", "").strip()
    if cfg_name:
        return cfg_name
    # 2. bankName field (sometimes contains the method name for bank transfers)
    bank = pay_term.get("bankName", "").strip()
    # 3. paymentType code → resolved name
    ptype = str(pay_term.get("paymentType", "")).strip()
    if ptype:
        resolved = _get_payment_name_for_user(ptype, uid)
        if not resolved.startswith("Type "):
            return resolved
        # If we only have "Type XXX" and also have a bankName, prefer bankName
        if bank:
            return bank
        return resolved
    if bank:
        return bank
    return "—"


def _has_account_info(order_detail: dict) -> tuple:
    """
    Returns (has_info: bool, account_no: str, real_name: str).
    Uses merged pay_term from confirmedPayTerm + paymentTermList.
    """
    pay_term   = _resolve_pay_term(order_detail)
    account_no = pay_term.get("accountNo", "").strip()
    real_name  = (
        pay_term.get("realName", "").strip()
        or order_detail.get("sellerRealName", "").strip()
    )
    has_info = bool(account_no) and bool(real_name)
    return has_info, account_no, real_name


# ─────────────────────────────────────────
# 📦 FORMAT ORDER MESSAGES
# ─────────────────────────────────────────
def format_order_message(order_detail: dict, seller_info: dict, uid: int = 0) -> str:
    order_type = order_detail.get("orderType", "ORIGIN")
    quantity   = order_detail.get("quantity",  "—")
    amount     = order_detail.get("amount",    "—")
    currency   = order_detail.get("currencyId","—")
    price      = order_detail.get("price",     "—")
    order_id   = order_detail.get("id",        "—")
    token      = order_detail.get("tokenId",   "—")

    # Unified resolver: merges confirmedPayTerm + paymentTermList[0] so no field is lost
    pay_term   = _resolve_pay_term(order_detail)

    pay_name   = _get_pay_name(pay_term, uid)
    bank_name  = pay_term.get("bankName",  "").strip() or "—"
    real_name  = (
        pay_term.get("realName", "").strip()
        or order_detail.get("sellerRealName", "").strip()
        or "—"
    )
    account_no = pay_term.get("accountNo", "").strip() or "—"

    good_rate   = seller_info.get("goodAppraiseRate", "—")
    avg_release = seller_info.get("averageReleaseTime", "0")

    try:
        release_mins = float(avg_release)
        release_str  = f"{release_mins:.0f} min"
        _bp_thresh   = _s(uid).buyer_protection_mins if uid else 30
        slow_warn    = f"\n\n⚠️ *Seller release time too long!* ({release_mins:.0f} min)" \
                       if release_mins >= _bp_thresh else ""
    except (ValueError, TypeError):
        release_mins = 0
        release_str  = str(avg_release)
        slow_warn    = ""

    missing_warn = "\n\n❗ *Missing account info — Name Match will skip FLW transfer.*" \
                   if (account_no == "—" or real_name == "—") else ""

    return (
        f"{'─' * 28}\n"
        f"🆔 <code>{order_id}</code>\n"
        f"🔄 <code>{order_type}</code> | 🪙 <code>{token}</code>\n"
        f"📦 Qty: <code>{quantity}</code> | 💵 <code>{amount} {currency}</code>\n"
        f"💲 Price: <code>{price}</code>\n"
        f"{'─' * 28}\n"
        f"💳 Payment: <b>{pay_name}</b>\n"
        f"🏦 Bank: <code>{bank_name}</code>\n"
        f"👤 Seller Name: <code>{real_name}</code>\n"
        f"🔢 Account: <code>{account_no}</code>\n"
        f"{'─' * 28}\n"
        f"📊 Seller Rating: <code>{good_rate}%</code>\n"
        f"⏱ Avg Release: <code>{release_str}</code>"
        f"{slow_warn}"
        f"{missing_warn}"
    )


def format_sell_order_message(order_detail: dict, buyer_info: dict, uid: int = 0) -> str:
    quantity  = order_detail.get("quantity",  "—")
    amount    = order_detail.get("amount",    "—")
    currency  = order_detail.get("currencyId","—")
    price     = order_detail.get("price",     "—")
    order_id  = order_detail.get("id",        "—")
    token     = order_detail.get("tokenId",   "—")

    buyer_name = (
        order_detail.get("buyerRealName", "").strip()
        or buyer_info.get("realName", "").strip()
        or "—"
    )

    # For sell orders MY payment details are in paymentTermList (seller's own terms).
    # Use _resolve_pay_term to merge confirmedPayTerm + paymentTermList[0] so no field is lost.
    my_pay_term  = _resolve_pay_term(order_detail)
    my_pay_name  = _get_pay_name(my_pay_term, uid)
    my_bank      = my_pay_term.get("bankName",  "").strip() or "—"
    my_name      = (
        my_pay_term.get("realName", "").strip()
        or order_detail.get("sellerRealName", "").strip()
        or "—"
    )
    my_account   = my_pay_term.get("accountNo", "").strip() or "—"

    good_rate    = buyer_info.get("goodAppraiseRate",    "—")
    avg_transfer = buyer_info.get("averageTransferTime", "—")

    return (
        f"{'─' * 28}\n"
        f"🆔 <code>{order_id}</code>\n"
        f"🪙 Token: <code>{token}</code> | Qty: <code>{quantity}</code>\n"
        f"💵 Amount: <code>{amount} {currency}</code> | 💲 <code>{price}</code>\n"
        f"{'─' * 28}\n"
        f"👤 <b>Buyer Name:</b> <code>{buyer_name}</code>\n"
        f"📊 Buyer Rating: <code>{good_rate}%</code>\n"
        f"⏱ Avg Transfer Time: <code>{avg_transfer} min</code>\n"
        f"{'─' * 28}\n"
        f"🏦 <b>My Payment Details:</b>\n"
        f"💳 Method: <b>{my_pay_name}</b>\n"
        f"🏦 Bank: <code>{my_bank}</code>\n"
        f"👤 My Name: <code>{my_name}</code>\n"
        f"🔢 Account: <code>{my_account}</code>\n"
        f"{'─' * 28}"
    )


def order_buttons(order_id: str, autopay_failed: bool = False, uid: int = 0) -> InlineKeyboardMarkup | None:
    """
    BUY order buttons.
    - If auto-pay succeeded → return None (no buttons — order is handled)
    - If auto-pay failed or manual → show Mark Paid buttons
    """
    if not autopay_failed and uid and order_id in _s(uid).paid_order_ids:
        return None   # already paid — remove buttons
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Mark as Paid",            callback_data=f"pay_{order_id}")],
        [InlineKeyboardButton("⚠️ Paid + Warn Seller 🐌", callback_data=f"paywarn_{order_id}")],
    ])


def sell_order_buttons(order_id: str, uid: int = 0) -> InlineKeyboardMarkup | None:
    """SELL order buttons — disappear once coins are released."""
    if uid and order_id in _s(uid).released_ids:
        return None
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🪙 RELEASE COIN", callback_data=f"release_{order_id}")],
    ])


# ─────────────────────────────────────────
# 📦 ORDER MONITOR LOOP
# ─────────────────────────────────────────
async def _flw_autopay(bot, chat_id, order_id, order_detail):
    """
    Flutterwave auto-pay flow — fully NoneType-safe, correct order:
      STEP 1 → Name Match / Buyer Protection checks
      STEP 2 → Resolve bank code
      STEP 3 → Verify account
      STEP 4 → Initiate transfer
      STEP 5 → Poll transfer until SUCCESSFUL or FAILED
      STEP 6 → ONLY if SUCCESSFUL → mark Bybit order paid
      STEP 7 → Send confirmation to user

    Per-user isolated: uses chat_id to load FLW keys and Bybit creds.
    NEVER marks Bybit paid before transfer is confirmed SUCCESSFUL.
    """
    from flutterwave import match_bank_code, verify_account, send_transfer, get_transfer_status

    # ── Per-user FLW secret key (slot-independent — FLW is shared across slots) ──
    flw_secret_key = db.get_api(chat_id, "flw_secret_key")
    user_slot      = _get_user_slot_str(chat_id)

    # ── Abort if this order was already finalized (manual pay, webhook, etc.) ──
    if _is_order_finalized(chat_id, order_id):
        logger.info(f"[FLW] Order {order_id} already finalized — skipping autopay")
        return

    logger.info(
        f"[FLW] _flw_autopay START | user={chat_id} slot={user_slot} order={order_id}"
    )

    if not flw_secret_key:
        oid = _esc(order_id)
        logger.warning(f"[FLW] No FLW secret key for user={chat_id} — aborting order={order_id}")
        await bot.send_message(chat_id=chat_id,
            text=(
                f"❌ <b>FLW Auto-Pay</b> — Order <code>{oid}</code>\n\n"
                "No Flutterwave API configured.\n"
                "Go to 🔑 <b>Set APIs</b> → Set Flutterwave API first."
            ),
            parse_mode="HTML")
        return

    try:
        # ── STEP 1a: Name Match check ──
        if _s(chat_id).name_match_enabled:
            has_info, account_no_chk, real_name_chk = _has_account_info(order_detail)
            if not has_info:
                logger.info(f"[FLW][NameMatch] Missing account info on order={order_id} — marking paid + warn, skipping FLW")
                pay_term_nm = order_detail.get("confirmedPayTerm", {}) or {}
                if not pay_term_nm:
                    terms_nm    = order_detail.get("paymentTermList", [])
                    pay_term_nm = terms_nm[0] if terms_nm else {}
                pt  = str(pay_term_nm.get("paymentType", ""))
                pid = str(pay_term_nm.get("id", ""))
                if pt and pid:
                    await asyncio.get_event_loop().run_in_executor(
                        None, partial(mark_order_paid, order_id, pt, pid, creds=get_user_creds(chat_id))
                    )
                    _s(chat_id).paid_order_ids.add(order_id)
                    _track_buy_volume(chat_id, order_id, order_detail)
                await asyncio.get_event_loop().run_in_executor(
                    None, partial(send_chat_message, order_id, NO_ACCOUNT_WARN_MSG,
                                  creds=get_user_creds(chat_id))
                )
                _s(chat_id).expecting_cancel_ids.add(order_id)
                oid = _esc(order_id)
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"🔍 <b>Name Match — Missing Info</b>\n\n"
                        f"Order: <code>{oid}</code>\n"
                        f"Account details incomplete — FLW transfer skipped.\n"
                        f"Marked paid on Bybit + seller asked to cancel."
                    ),
                    parse_mode="HTML")
                return

        # ── Extract payment term details ──
        pay_term = order_detail.get("confirmedPayTerm", {}) or {}
        if not pay_term:
            terms    = order_detail.get("paymentTermList", [])
            pay_term = terms[0] if terms else {}

        account_no    = pay_term.get("accountNo", "").strip()
        bank_name     = pay_term.get("bankName",  "").strip()
        pay_cfg       = pay_term.get("paymentConfigVo", {}) or pay_term.get("paymentConfig", {}) or {}
        pay_type_name = pay_cfg.get("paymentName", "").strip()
        seller_name   = pay_term.get("realName", order_detail.get("sellerRealName", "Seller")).strip() or "Seller"

        # ── Amount: parse safely, format as float rounded to 2 dp ──
        try:
            amount = round(float(str(order_detail.get("amount", "0")).replace(",", "")), 2)
        except (ValueError, TypeError):
            amount = 0.0

        currency = str(order_detail.get("currencyId", "NGN")).upper()

        logger.info(
            f"[FLW] Payload preview | user={chat_id} slot={user_slot} order={order_id} "
            f"account_no={account_no!r} bank_name={bank_name!r} pay_type_name={pay_type_name!r} "
            f"amount={amount} currency={currency} seller={seller_name!r}"
        )

        if not account_no:
            oid = _esc(order_id)
            logger.warning(f"[FLW] No account_no for order={order_id} | user={chat_id}")
            await bot.send_message(chat_id=chat_id,
                text=f"❌ <b>FLW Auto-Pay</b> — Order <code>{oid}</code>\nNo account number found in order. Mark manually.",
                parse_mode="HTML")
            return

        if amount <= 0:
            oid = _esc(order_id)
            logger.warning(f"[FLW] Invalid amount={amount} for order={order_id} | user={chat_id}")
            await bot.send_message(chat_id=chat_id,
                text=f"❌ <b>FLW Auto-Pay</b> — Order <code>{oid}</code>\nInvalid order amount: <code>{amount}</code>. Mark manually.",
                parse_mode="HTML")
            return

        # ── STEP 1b: Buyer Protection — slow seller → skip FLW, mark paid + warn ──
        if _s(chat_id).buyer_protection_on:
            release_mins = 0.0
            try:
                release_mins = float(order_detail.get("_seller_release_mins", 0) or 0)
            except (ValueError, TypeError):
                release_mins = 0.0
            release_unknown = bool(order_detail.get("_seller_release_unknown", False))
            if release_unknown or release_mins >= _s(chat_id).buyer_protection_mins:
                reason = (
                    "Seller release time unknown (fetch failed) — flagged as high-risk"
                    if release_unknown else
                    f"Seller avg release time ({release_mins:.0f} min) ≥ threshold ({_s(chat_id).buyer_protection_mins} min)"
                )
                logger.info(f"[FLW][BuyerProtection] Skipping FLW — {reason} | order={order_id} user={chat_id}")
                _s(chat_id).unpaid_log.append({
                    "order_id":   order_id,
                    "account_no": account_no,
                    "bank":       bank_name or pay_type_name,
                    "amount":     amount,
                    "reason":     reason,
                    "timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })
                pay_type   = str(pay_term.get("paymentType", ""))
                payment_id = str(pay_term.get("id", ""))
                if pay_type and payment_id:
                    await asyncio.get_event_loop().run_in_executor(
                        None, partial(mark_order_paid, order_id, pay_type, payment_id, creds=get_user_creds(chat_id))
                    )
                    _s(chat_id).paid_order_ids.add(order_id)
                    _track_buy_volume(chat_id, order_id, order_detail)
                await asyncio.get_event_loop().run_in_executor(
                    None, partial(send_chat_message, order_id, SELLER_WARN_MSG,
                                  creds=get_user_creds(chat_id))
                )
                _s(chat_id).expecting_cancel_ids.add(order_id)
                oid    = _esc(order_id)
                thresh = _s(chat_id).buyer_protection_mins
                # ── Update order message: remove buttons, show ⏭ Skipped badge ──
                await _update_order_message_final(bot, chat_id, order_id, "BP Triggered", "skipped")
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"🛡 <b>Buyer Protection Triggered</b> — Order <code>{oid}</code>\n\n"
                        f"Seller release time: <code>{release_mins:.0f} min</code> ≥ <code>{thresh} min</code>\n"
                        f"✅ Marked paid on Bybit + warning sent to seller.\n"
                        f"FLW transfer was skipped."
                    ),
                    parse_mode="HTML")
                return

        acct_safe = _esc(account_no)
        bank_safe = _esc(bank_name or pay_type_name)
        oid       = _esc(order_id)

        # ── STEP 2: Resolve bank code ──
        bank_code = match_bank_code(bank_name, pay_type_name, secret_key=flw_secret_key)
        logger.info(f"[FLW] Bank resolve | user={chat_id} order={order_id} bank_name={bank_name!r} pay_type={pay_type_name!r} → bank_code={bank_code!r}")
        if not bank_code:
            logger.warning(f"[FLW] Unknown bank for order={order_id} | user={chat_id} | bank={bank_name!r} type={pay_type_name!r}")
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"❌ <b>FLW Auto-Pay</b> — Order <code>{oid}</code>\n"
                    f"Unknown bank: <code>{bank_safe}</code>\n"
                    f"Cannot resolve bank code — mark this order manually."
                ),
                parse_mode="HTML")
            return

        # ── STEP 3: Verify account ──
        await bot.send_message(chat_id=chat_id,
            text=(
                f"⏳ <b>FLW</b> — Order <code>{oid}</code>\n"
                f"Verifying account <code>{acct_safe}</code> ({bank_safe})...\n"
                f"Amount: <b>{amount:,.2f} {currency}</b>"
            ),
            parse_mode="HTML")

        verify = await asyncio.get_event_loop().run_in_executor(
            None, verify_account, account_no, bank_code, flw_secret_key
        )

        # Safely extract verify data — data may be null
        verify_data   = verify.get("data") or {}
        verify_status = verify.get("status", "")
        verify_error  = verify.get("message", verify.get("error", ""))

        logger.info(
            f"[FLW] Account verify result | user={chat_id} order={order_id} "
            f"status={verify_status!r} data={verify_data} error={verify_error!r}"
        )

        if verify_status != "success" or "error" in verify:
            err = _esc(str(verify_error or "Unknown verification error")[:200])
            _s(chat_id).unpaid_log.append({
                "order_id":   order_id,
                "account_no": account_no,
                "bank":       bank_name or pay_type_name,
                "amount":     amount,
                "reason":     f"Account verification failed: {verify_error}",
                "timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"❌ <b>FLW Account Invalid</b> — Order <code>{oid}</code>\n\n"
                    f"Account <code>{acct_safe}</code> @ {bank_safe}\n"
                    f"Reason: <code>{err}</code>\n\n"
                    f"Transfer aborted. Mark order manually."
                ),
                parse_mode="HTML")
            return

        verified_name = (verify_data.get("account_name") or seller_name or "Seller").strip()
        working_code  = verify.get("_working_bank_code") or bank_code
        vname_safe    = _esc(verified_name)

        await bot.send_message(chat_id=chat_id,
            text=(
                f"✅ <b>Account Verified:</b> {vname_safe}\n"
                f"Account: <code>{acct_safe}</code> ({bank_safe})\n\n"
                f"⏳ Initiating transfer of <b>{amount:,.2f} {currency}</b>..."
            ),
            parse_mode="HTML")

        # ── STEP 4: Initiate transfer ──
        sender_name = (_s(chat_id).settings.get("sender_name") or "P2P Bot").strip()
        narration   = f"{sender_name} payment to {verified_name}"[:100]
        ref         = f"p2p{order_id[-12:]}"

        transfer_payload = {
            "account_no":    account_no,
            "bank_code":     working_code,
            "amount":        amount,
            "narration":     narration,
            "reference":     ref,
            "currency":      currency,
            "beneficiary":   verified_name,
        }
        logger.info(
            f"[FLW] Transfer payload | user={chat_id} slot={user_slot} order={order_id} "
            f"account={account_no} bank_code={working_code} amount={amount} "
            f"currency={currency} ref={ref!r} narration={narration!r}"
        )

        result = await asyncio.get_event_loop().run_in_executor(
            None, send_transfer, account_no, working_code, amount,
            narration, ref, flw_secret_key
        )

        # ── Sanitised response log (never log full response to avoid key leaks) ──
        result_status = result.get("status", "")
        result_msg    = result.get("message", "")
        result_error  = result.get("error", "")
        # Safely get data — Flutterwave sometimes returns "data": null on errors
        result_data   = result.get("data") or {}
        logger.info(
            f"[FLW] Transfer response | user={chat_id} slot={user_slot} order={order_id} "
            f"status={result_status!r} message={result_msg!r} error={result_error!r} "
            f"data_keys={list(result_data.keys()) if result_data else 'null'}"
        )

        # ── Handle hard error key ──
        if result_error:
            err_msg  = str(result_error)
            ip       = await _get_current_ip()
            err_safe = _esc(err_msg[:250])
            ip_safe  = _esc(ip)
            logger.error(f"[FLW] Transfer error | user={chat_id} order={order_id} | {err_msg}")
            _s(chat_id).unpaid_log.append({
                "order_id": order_id, "account_no": account_no,
                "bank": bank_name or pay_type_name, "amount": amount,
                "reason": err_msg[:300],
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            if "Empty response" in err_msg or "401" in err_msg or "403" in err_msg:
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"❌ <b>FLW Blocked</b> — Order <code>{oid}</code>\n\n"
                        f"<code>{err_safe}</code>\n\n"
                        f"👉 Add <code>{ip_safe}</code> to Flutterwave IP Whitelist.\n"
                        f"Mark order manually."
                    ),
                    parse_mode="HTML")
            else:
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"❌ <b>FLW Transfer Error</b> — Order <code>{oid}</code>\n\n"
                        f"<code>{err_safe}</code>\n\n"
                        f"Mark order manually."
                    ),
                    parse_mode="HTML")
            return

        # ── Handle API-level error status (e.g. "data": null + "status": "error") ──
        if result_status == "error" or (not result_data and result_status != "success"):
            api_err  = _esc((result_msg or "Flutterwave rejected the transfer request")[:300])
            logger.error(
                f"[FLW] API-level error | user={chat_id} slot={user_slot} order={order_id} "
                f"message={result_msg!r} data=null"
            )
            _s(chat_id).unpaid_log.append({
                "order_id": order_id, "account_no": account_no,
                "bank": bank_name or pay_type_name, "amount": amount,
                "reason": result_msg or "FLW API error — data null",
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            # Check for the specific "contact administrator" error
            if "administrator" in (result_msg or "").lower() or "cannot be processed" in (result_msg or "").lower():
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"❌ <b>FLW Transfer Rejected</b> — Order <code>{oid}</code>\n\n"
                        f"Flutterwave returned an account restriction error:\n"
                        f"<code>{api_err}</code>\n\n"
                        f"⚠️ <b>Action required:</b> Log into your Flutterwave dashboard and check:\n"
                        f"  • Account limits or KYC requirements\n"
                        f"  • Transfer restrictions or compliance holds\n"
                        f"  • Contact Flutterwave support if this persists\n\n"
                        f"Order has NOT been marked paid. Mark manually when resolved."
                    ),
                    parse_mode="HTML")
            else:
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"❌ <b>FLW Transfer Failed</b> — Order <code>{oid}</code>\n\n"
                        f"<code>{api_err}</code>\n\n"
                        f"Order has NOT been marked paid. Mark manually."
                    ),
                    parse_mode="HTML")
            return

        # ── result_data is guaranteed non-None from here ──
        transfer_id  = str(result_data.get("id") or "")
        status       = str(result_data.get("status") or "NEW")
        tid_safe     = _esc(transfer_id)
        # tx_ref is the reference we sent — used by webhook to look up this job
        tx_ref       = str(result_data.get("reference") or ref)

        # ── Register transfer so webhook can reconnect to user + order ──
        if tx_ref:
            _flw_transfer_registry[tx_ref] = {
                "order_id":  order_id,
                "user_id":   chat_id,
                "slot":      user_slot,
                "amount":    amount,
                "currency":  currency,
                "pay_term":  pay_term,
            }
            logger.info(
                f"[FLW] Transfer registered | ref={tx_ref!r} transfer_id={transfer_id!r} "
                f"user={chat_id} order={order_id}"
            )

        logger.info(
            f"[FLW] Transfer initiated | user={chat_id} slot={user_slot} order={order_id} "
            f"transfer_id={transfer_id!r} initial_status={status!r}"
        )

        # ── Register transfer in global registry for webhook reconnection ──
        # This allows the /flw-webhook endpoint to find the right user + order
        # when Flutterwave sends a status update callback, even if polling timed out.
        _flw_transfer_registry[transfer_id] = {
            "transfer_ref": transfer_id,
            "order_id":     order_id,
            "user_id":      chat_id,       # Telegram chat_id of the bot user
            "slot":         user_slot,
            "amount":       amount,
            "currency":     currency,
            "pay_term":     pay_term,
            "verified_name": verified_name,
        }
        logger.info(f"[FLW] Registered transfer {transfer_id!r} in registry for order {order_id}")

        # ── Handle immediate FAILED status on creation ──
        if status == "FAILED":
            complete_msg = str(result_data.get("complete_message") or "Rejected by bank")
            logger.warning(f"[FLW] Transfer immediately FAILED | user={chat_id} order={order_id} | {complete_msg}")
            _s(chat_id).unpaid_log.append({
                "order_id": order_id, "account_no": account_no,
                "bank": bank_name or pay_type_name, "amount": amount,
                "reason": complete_msg,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            cmsg_safe = _esc(complete_msg)
            if "insufficient" in complete_msg.lower() or "funds" in complete_msg.lower():
                fail_text = (
                    f"❌ <b>FLW Failed — Insufficient Funds</b>\n\nOrder: <code>{oid}</code>\n"
                    f"Amount needed: <b>{amount:,.2f} {currency}</b>\n\n"
                    f"👉 Top up Flutterwave → Balances → Fund Wallet\n\n"
                    f"Order has NOT been marked paid."
                )
            else:
                fail_text = (
                    f"❌ <b>FLW Transfer Failed on Creation</b>\n\nOrder: <code>{oid}</code>\n"
                    f"Transfer ID: <code>{tid_safe}</code>\n"
                    f"Reason: <code>{cmsg_safe}</code>\n\n"
                    f"Order has NOT been marked paid. Mark manually."
                )
            await bot.send_message(chat_id=chat_id, text=fail_text, parse_mode="HTML")
            return

        # ── STEP 5: Poll transfer status up to 60 seconds ──
        final_status = status
        for attempt in range(12):
            await asyncio.sleep(5)
            if final_status in ("SUCCESSFUL", "FAILED"):
                break
            poll      = await asyncio.get_event_loop().run_in_executor(
                None, get_transfer_status, transfer_id, flw_secret_key
            )
            # Safely extract — data can be null even on polling
            poll_data    = poll.get("data") or {}
            final_status = str(poll_data.get("status") or final_status)
            logger.debug(
                f"[FLW] Poll attempt={attempt+1} | user={chat_id} order={order_id} "
                f"transfer_id={transfer_id!r} status={final_status!r}"
            )

        logger.info(
            f"[FLW] Final transfer status | user={chat_id} slot={user_slot} order={order_id} "
            f"transfer_id={transfer_id!r} final_status={final_status!r}"
        )

        if final_status == "SUCCESSFUL":
            # ── Guard: webhook may have already handled this order ──
            # The FLW webhook fires within seconds of transfer completion.
            # If it arrived while we were polling, it already marked Bybit paid
            # and sent a success notification. Don't duplicate.
            if _is_order_final(order_id):
                logger.info(
                    f"[FLW] Poll got SUCCESSFUL but order already finalized by webhook "
                    f"| user={chat_id} order={order_id} — skipping duplicate notification"
                )
                return
            # ── STEP 6: ONLY now mark Bybit order paid ──
            pay_type   = str(pay_term.get("paymentType", ""))
            payment_id = str(pay_term.get("id", ""))
            bybit_ok   = False
            if pay_type and payment_id:
                pr = await asyncio.get_event_loop().run_in_executor(
                    None, partial(mark_order_paid, order_id, pay_type, payment_id,
                                  creds=get_user_creds(chat_id))
                )
                bybit_ok = (pr or {}).get("retCode", -1) == 0
                logger.info(
                    f"[FLW] Bybit mark-paid | user={chat_id} order={order_id} "
                    f"bybit_ok={bybit_ok} retCode={(pr or {}).get('retCode','?')}"
                )
            _s(chat_id).paid_order_ids.add(order_id)
            _track_buy_volume(chat_id, order_id, order_detail)
            _set_order_final(order_id, "completed")
            # ── Update order message: remove action buttons, show ✅ Completed badge ──
            await _update_order_message_final(bot, chat_id, order_id, "Transfer Completed", "completed")
            # ── STEP 7: Send Telegram confirmation ──
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"✅ <b>FLW Transfer Successful</b>\n\n"
                    f"Amount: <b>₦{amount:,.2f}</b>\n"
                    f"Recipient: <b>{vname_safe}</b>\n"
                    f"Order: <code>{oid}</code>\n"
                    f"Transfer ID: <code>{tid_safe}</code>\n"
                    f"✅ Bybit order marked as paid."
                ),
                parse_mode="HTML")

        elif final_status == "FAILED":
            # Fetch final state for complete_message
            last_poll    = await asyncio.get_event_loop().run_in_executor(
                None, get_transfer_status, transfer_id, flw_secret_key
            )
            last_data    = last_poll.get("data") or {}
            complete_msg = str(last_data.get("complete_message") or "")
            logger.warning(
                f"[FLW] Transfer FAILED after polling | user={chat_id} order={order_id} "
                f"transfer_id={transfer_id!r} complete_message={complete_msg!r}"
            )
            _s(chat_id).unpaid_log.append({
                "order_id": order_id, "account_no": account_no,
                "bank": bank_name or pay_type_name, "amount": amount,
                "reason": complete_msg or "Transfer FAILED after polling",
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            # ── Update order message to show ❌ Transfer Failed badge ──
            await _update_order_message_final(bot, chat_id, order_id, "Transfer Failed", "failed")
            cmsg_safe = _esc(complete_msg)
            if "insufficient" in complete_msg.lower() or "funds" in complete_msg.lower():
                fail_text = (
                    f"❌ <b>FLW Failed — Insufficient Funds</b>\n\n"
                    f"Order: <code>{oid}</code>\n"
                    f"Amount: <b>{amount:,.2f} {currency}</b>\n\n"
                    f"👉 Top up Flutterwave → Balances → Fund Wallet\n\n"
                    f"Order has NOT been marked paid."
                )
            else:
                reason_line = f"Reason: <code>{cmsg_safe}</code>\n" if complete_msg else ""
                fail_text = (
                    f"❌ <b>FLW Transfer FAILED</b>\n\n"
                    f"Order: <code>{oid}</code>\n"
                    f"Transfer ID: <code>{tid_safe}</code>\n"
                    f"{reason_line}"
                    f"Order has NOT been marked paid. Mark manually."
                )
            await bot.send_message(chat_id=chat_id, text=fail_text, parse_mode="HTML")

        else:
            # Status still pending after 60s polling — check if webhook already handled it
            if _is_order_final(order_id):
                logger.info(
                    f"[FLW] Still-pending exit but order already finalized by webhook "
                    f"| user={chat_id} order={order_id} — skipping pending notification"
                )
                return
            fstatus_safe = _esc(final_status)
            logger.info(
                f"[FLW] Transfer still pending after polling | user={chat_id} order={order_id} "
                f"transfer_id={transfer_id!r} status={final_status!r}"
            )
            await _update_order_message_final(
                context.bot if hasattr(bot, "bot") else bot,
                chat_id, order_id, "Transfer Pending", "skipped"
            )
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"⏳ <b>FLW Transfer Pending</b>\n\n"
                    f"Order: <code>{oid}</code>\n"
                    f"Transfer ID: <code>{tid_safe}</code> | Status: <code>{fstatus_safe}</code>\n\n"
                    f"Order has NOT been marked paid yet.\n"
                    f"Flutterwave webhook will confirm and auto-mark when complete.\n"
                    f"Transfer ID is registered — webhook will reconnect automatically."
                ),
                parse_mode="HTML")

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logger.error(
            f"[FLW] _flw_autopay UNHANDLED ERROR | user={chat_id} order={order_id} | "
            f"error={e}\n{tb}"
        )
        oid      = _esc(order_id)
        err_safe = _esc(str(e)[:250])
        try:
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"❌ <b>FLW Auto-Pay Error</b> — Order <code>{oid}</code>\n\n"
                    f"<code>{err_safe}</code>\n\n"
                    f"Order has NOT been marked paid. Mark manually."
                ),
                parse_mode="HTML")
        except Exception as _notify_err:
            logger.error(f"[FLW] Could not notify user {chat_id} of error: {_notify_err}")


# ─────────────────────────────────────────
# 🟡 PAGA PAYMENT QUEUE WORKER
# Processes Paga payments strictly one at a time.
# Orders arriving while one is processing are queued and notified.
# ─────────────────────────────────────────
async def _paga_queue_worker():
    """
    Single background worker that drains the Paga payment queue.
    Each order is fully resolved (success / fail / pending timeout)
    before the next one starts — prevents Paga rate-limit rejections
    when multiple Bybit orders arrive simultaneously.
    """
    global _paga_queue_list
    logger.info("[Paga Queue] Worker started")
    while True:
        try:
            item = await _paga_queue.get()
            if item is None:
                logger.info("[Paga Queue] Worker received stop signal")
                break

            bot, chat_id, order_id, order_detail = item

            # Remove from display list
            _paga_queue_list = [x for x in _paga_queue_list if x[0] != order_id]

            remaining = _paga_queue.qsize()
            pos_msg   = f"\n\n📋 *{remaining} order(s) still in queue after this.*" if remaining > 0 else ""

            logger.info(f"[Paga Queue] Processing order {order_id} | queue remaining={remaining}")

            if remaining > 0:
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"🟡 <b>Paga Queue</b> — Processing order <code>{_esc(order_id)}</code>\n"
                        f"📋 <code>{remaining}</code> order(s) waiting after this one."
                    ),
                    parse_mode="HTML"
                )

            try:
                await _paga_autopay(bot, chat_id, order_id, order_detail)
            except Exception as e:
                logger.error(f"[Paga Queue] Error processing {order_id}: {e}")
                try:
                    oid      = _esc(order_id)
                    err_safe = _esc(str(e)[:200])
                    await bot.send_message(
                        chat_id=chat_id,
                        text=f"❌ <b>Paga Queue error</b> — Order <code>{oid}</code>\n<code>{err_safe}</code>",
                        parse_mode="HTML"
                    )
                except Exception:
                    pass

            _paga_queue.task_done()

        except Exception as e:
            logger.error(f"[Paga Queue] Worker loop error: {e}")
            await asyncio.sleep(2)

    logger.info("[Paga Queue] Worker stopped")


def _enqueue_paga_order(bot, chat_id, order_id, order_detail):
    """
    Add a Paga payment job to the queue.
    Also updates the display list with order summary for status reporting.
    """
    global _paga_queue_list
    amount   = order_detail.get("amount", "?")
    pay_term = order_detail.get("confirmedPayTerm", {}) or {}
    if not pay_term:
        terms    = order_detail.get("paymentTermList", [])
        pay_term = terms[0] if terms else {}
    bank = pay_term.get("bankName", "") or pay_term.get("paymentType", "?")
    _paga_queue_list.append((order_id, amount, bank))
    _paga_queue.put_nowait((bot, chat_id, order_id, order_detail))
    pos = _paga_queue.qsize()
    logger.info(f"[Paga Queue] Enqueued {order_id} | queue size={pos}")
    return pos


def _is_order_final(order_id: str) -> bool:
    """Return True if this order has already reached a final state (prevents duplicate actions)."""
    return order_id in _order_final_states


def _set_order_final(order_id: str, state: str):
    """Mark an order as final. state ∈ {'completed','rejected','warned','failed','expired'}."""
    _order_final_states[order_id] = state


async def _update_order_message(bot, chat_id: int, order_id: str,
                                 status_text: str, *, keep_buttons: bool = False):
    """
    Edit the original BUY order Telegram message to:
      1. Append a status line at the end
      2. Remove all inline keyboard buttons (unless keep_buttons=True)

    Falls back silently if the message is no longer editable (e.g. too old).
    """
    msg_id = _s(chat_id).order_msg_ids.get(order_id)
    if not msg_id:
        return
    try:
        # Fetch the current message text if possible, then append status
        try:
            current = await bot.get_message_text(chat_id=chat_id, message_id=msg_id)
        except Exception:
            current = None

        new_markup = InlineKeyboardMarkup([]) if not keep_buttons else None

        if current:
            new_text = current + f"\n\n{status_text}"
            try:
                await bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=msg_id,
                    text=new_text,
                    reply_markup=new_markup,
                    parse_mode="HTML"
                )
                return
            except Exception:
                pass
        # If we can't edit the text, at minimum remove the buttons
        await bot.edit_message_reply_markup(
            chat_id=chat_id,
            message_id=msg_id,
            reply_markup=InlineKeyboardMarkup([])
        )
    except Exception as e:
        logger.debug(f"[AutoPay] Could not update message for order {order_id}: {e}")


# ─────────────────────────────────────────
# 🔒 ORDER LOCK + STATE HELPERS
# ─────────────────────────────────────────

def _get_order_lock(chat_id: int, order_id: str) -> asyncio.Lock:
    """Return (and create if needed) the asyncio.Lock for (chat_id, order_id).
    Prevents concurrent auto-pay and manual button taps on the same order."""
    key = (chat_id, order_id)
    if key not in _order_action_locks:
        _order_action_locks[key] = asyncio.Lock()
    return _order_action_locks[key]


def _is_order_finalized(chat_id: int, order_id: str) -> bool:
    """Return True if this order has already reached a terminal state.
    All further callbacks for this order are silently ignored."""
    return (chat_id, order_id) in _order_final_states


def _set_order_final_state(chat_id: int, order_id: str, state: str):
    """Mark an order as having reached a terminal state.
    Valid states: 'completed', 'rejected', 'warned', 'failed', 'expired', 'skipped'"""
    _order_final_states[(chat_id, order_id)] = state
    logger.info(f"[OrderState] ({chat_id}, {order_id}) → {state}")


async def _update_order_message_final(
    bot, chat_id: int, order_id: str,
    status_text: str, state: str
):
    """Edit the original BUY order Telegram message to show a final status badge
    and remove all action buttons so users cannot re-press them.

    state: one of 'completed', 'rejected', 'warned', 'failed', 'skipped', 'expired'
    """
    _set_order_final_state(chat_id, order_id, state)

    badge_map = {
        "completed": "✅ Completed",
        "rejected":  "❌ Rejected",
        "warned":    "⚠️ Warning Sent",
        "failed":    "❌ Transfer Failed",
        "skipped":   "⏭ Skipped",
        "expired":   "⏰ Expired",
    }
    badge_label = badge_map.get(state, f"ℹ️ {state.title()}")
    status_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton(badge_label, callback_data=f"order_status_noop_{order_id}")]
    ])

    msg_id = _s(chat_id).order_msg_ids.get(order_id)
    if not msg_id:
        return
    try:
        await bot.edit_message_reply_markup(
            chat_id=chat_id,
            message_id=msg_id,
            reply_markup=status_keyboard
        )
        logger.info(f"[OrderMsg] Updated order message for {order_id} → state={state}")
    except Exception as e:
        logger.debug(f"[OrderMsg] Could not update order message for {order_id}: {e}")


async def _remove_order_buttons(bot, chat_id: int, order_id: str):
    """Remove pay buttons from the BUY order notification message after auto-pay success.
    Delegates to _update_order_message_final with 'completed' state."""
    await _update_order_message_final(bot, chat_id, order_id, "Completed", "completed")


# ─────────────────────────────────────────
# 🔔 FLW WEBHOOK PROCESSOR
# Called by the web server (server.py / main.py) when Flutterwave POSTs a webhook.
# STEP 1: Verify signature using FLW_SECRET_HASH per user
# STEP 2: Check transfer status == SUCCESSFUL
# STEP 3: Look up order via _flw_transfer_registry
# STEP 4: Mark Bybit order paid
# STEP 5: Notify Telegram user
# STEP 6: Remove/update buttons
# ─────────────────────────────────────────
async def handle_flw_webhook(bot, payload: dict, signature_header: str | None):
    """
    Process an incoming Flutterwave webhook event.

    Args:
        bot: Telegram Bot instance
        payload: Parsed JSON body from Flutterwave
        signature_header: Value of the 'verif-hash' (or 'X-Flw-Signature') HTTP header

    Returns:
        (ok: bool, reason: str)
    """
    import hmac, hashlib

    logger.info(f"[FLW Webhook] Received | event={payload.get('event','?')} "
                f"has_signature={'yes' if signature_header else 'no'}")

    # ── STEP 1: Verify signature ──
    # The FLW_SECRET_HASH is stored per user. We must find which user owns this transfer
    # first, then verify the signature against their secret hash.
    # However, we can do a fast pre-check: look up the transfer ref in the registry first.

    data       = payload.get("data", {}) or {}
    event_type = payload.get("event", "")

    # Flutterwave sends transfer events as "transfer.completed"
    if "transfer" not in event_type.lower() and "transfer" not in str(payload.get("event_type", "")).lower():
        logger.info(f"[FLW Webhook] Non-transfer event: {event_type!r} — ignoring")
        return True, "not_transfer"

    ref    = str(data.get("reference") or data.get("narration", "")).strip()
    status = str(data.get("status", "")).upper()

    logger.info(f"[FLW Webhook] Transfer event | ref={ref!r} status={status!r}")

    if not ref:
        logger.warning("[FLW Webhook] No reference in payload — cannot identify transfer")
        return False, "no_reference"

    # ── Look up registry ──
    entry = _flw_transfer_registry.get(ref)
    if not entry:
        logger.warning(f"[FLW Webhook] ref={ref!r} not in registry — may be from a different session or manual transfer")
        return False, "unknown_ref"

    chat_id  = entry["user_id"]
    order_id = entry["order_id"]
    amount   = entry["amount"]
    currency = entry.get("currency", "NGN")
    pay_term = entry.get("pay_term", {})

    # ── STEP 1b: Verify signature against this user's FLW_SECRET_HASH ──
    secret_hash = db.get_api(chat_id, "flw_secret_hash")
    if not signature_header:
        logger.warning(f"[FLW Webhook] ⚠️ No signature header — rejecting for security | ref={ref!r} user={chat_id}")
        return False, "no_signature"

    if secret_hash:
        if signature_header != secret_hash:
            logger.warning(
                f"[FLW Webhook] 🔒 Signature MISMATCH — rejecting | "
                f"ref={ref!r} user={chat_id} expected={secret_hash[:6]}... got={signature_header[:6]}..."
            )
            return False, "invalid_signature"
        logger.info(f"[FLW Webhook] ✅ Signature verified | ref={ref!r} user={chat_id}")
    else:
        logger.warning(f"[FLW Webhook] No FLW_SECRET_HASH configured for user={chat_id} — skipping verification")

    # ── STEP 2: Check status ──
    if status != "SUCCESSFUL":
        reason_map = {
            "FAILED":    "Transfer failed",
            "REVERSED":  "Transfer reversed",
            "CANCELLED": "Transfer cancelled",
            "PENDING":   "Transfer still pending",
        }
        reason_msg = reason_map.get(status, f"Transfer status: {status}")
        logger.warning(f"[FLW Webhook] Non-success status={status!r} | ref={ref!r} user={chat_id} order={order_id}")

        if status in ("FAILED", "REVERSED", "CANCELLED"):
            # Notify user of failure
            oid = _esc(order_id)
            try:
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"❌ <b>FLW Transfer {status}</b>\n\n"
                        f"Order: <code>{oid}</code>\n"
                        f"Amount: <b>{amount:,.2f} {currency}</b>\n"
                        f"Reason: {_esc(reason_msg)}\n\n"
                        f"Order has <b>NOT</b> been marked paid. Mark manually."
                    ),
                    parse_mode="HTML"
                )
            except Exception as _notify_err:
                logger.error(f"[FLW Webhook] Could not notify user {chat_id}: {_notify_err}")
            # Remove buttons since transfer is definitively done (failed)
            _set_order_final(order_id, "failed")
            await _remove_order_buttons(bot, chat_id, order_id)
        return False, f"status_{status.lower()}"

    # ── STEP 3: Guard against duplicate webhook processing ──
    if order_id in _s(chat_id).paid_order_ids:
        logger.info(f"[FLW Webhook] Order {order_id} already marked paid — ignoring duplicate")
        return True, "already_paid"

    # ── STEP 4: Mark Bybit order paid ──
    pay_type   = str(pay_term.get("paymentType", ""))
    payment_id = str(pay_term.get("id", ""))
    bybit_ok   = False
    if pay_type and payment_id:
        try:
            pr = await asyncio.get_event_loop().run_in_executor(
                None, partial(mark_order_paid, order_id, pay_type, payment_id,
                              creds=get_user_creds(chat_id))
            )
            bybit_ok = (pr or {}).get("retCode", -1) == 0
            logger.info(
                f"[FLW Webhook] Bybit mark-paid | user={chat_id} order={order_id} "
                f"bybit_ok={bybit_ok} retCode={(pr or {}).get('retCode','?')}"
            )
        except Exception as _bp_err:
            logger.error(f"[FLW Webhook] Bybit mark-paid error | user={chat_id} order={order_id}: {_bp_err}")
    else:
        logger.warning(f"[FLW Webhook] Missing pay_type or payment_id — cannot mark Bybit paid | order={order_id}")

    _s(chat_id).paid_order_ids.add(order_id)
    try:
        _od_bv = await asyncio.get_event_loop().run_in_executor(
            None, partial(get_order_detail, order_id, creds=get_user_creds(chat_id))
        )
        if _od_bv.get("retCode", -1) == 0:
            _track_buy_volume(chat_id, order_id, _od_bv.get("result", {}))
    except Exception as _bv_err:
        logger.debug(f"[BuyVolume] FLW webhook fetch failed for {order_id}: {_bv_err}")
    _set_order_final(order_id, "completed")

    # ── STEP 5: Remove buttons + update message ──
    await _remove_order_buttons(bot, chat_id, order_id)

    # ── STEP 6: Notify user ──
    recipient_name = str(data.get("beneficiary_name") or data.get("full_name") or "Recipient")
    transfer_id    = str(data.get("id") or "")
    oid        = _esc(order_id)
    tid_safe   = _esc(transfer_id)
    rname_safe = _esc(recipient_name)
    bybit_line = "✅ Bybit order marked as paid." if bybit_ok else "⚠️ Could not auto-mark on Bybit — please mark manually."

    try:
        await bot.send_message(
            chat_id=chat_id,
            text=(
                f"✅ <b>Flutterwave Transfer Successful</b>\n\n"
                f"Amount: <b>₦{amount:,.0f}</b>\n"
                f"Recipient: <b>{rname_safe}</b>\n"
                f"Order: <code>{oid}</code>\n"
                f"Transfer ID: <code>{tid_safe}</code>\n\n"
                f"{bybit_line}"
            ),
            parse_mode="HTML"
        )
        logger.info(f"[FLW Webhook] ✅ Webhook processed successfully | ref={ref!r} user={chat_id} order={order_id} bybit_ok={bybit_ok}")
    except Exception as _notify_err:
        logger.error(f"[FLW Webhook] Could not send success notification to user {chat_id}: {_notify_err}")

    # Clean up registry to free memory
    _flw_transfer_registry.pop(ref, None)

    return True, "success"


# ─────────────────────────────────────────
# 🟡 PAGA SUCCESS / FAILURE HELPERS
# ─────────────────────────────────────────
async def _paga_handle_success(bot, chat_id, order_id, pay_term, amount, holder_name, txn_id, ref):
    """Mark Bybit order paid and notify admin on Paga success."""
    pay_type   = str(pay_term.get("paymentType", ""))
    payment_id = str(pay_term.get("id", ""))
    bybit_ok   = False
    if pay_type and payment_id:
        pr       = await asyncio.get_event_loop().run_in_executor(
            None, partial(mark_order_paid, order_id, pay_type, payment_id, creds=get_user_creds(chat_id))
        )
        bybit_ok = pr.get("retCode", -1) == 0
    _s(chat_id).paid_order_ids.add(order_id)
    try:
        _od_bv = await asyncio.get_event_loop().run_in_executor(
            None, partial(get_order_detail, order_id, creds=get_user_creds(chat_id))
        )
        if _od_bv.get("retCode", -1) == 0:
            _track_buy_volume(chat_id, order_id, _od_bv.get("result", {}))
    except Exception as _bv_err:
        logger.debug(f"[BuyVolume] Paga success fetch failed for {order_id}: {_bv_err}")
    logger.info(f"[Paga] ✅ SUCCESS: txnId={txn_id} | Bybit={bybit_ok}")
    await _remove_order_buttons(bot, chat_id, order_id)
    await bot.send_message(chat_id=chat_id,
        text=(
            f"✅ <b>Paga Payment SUCCESS</b>\n\n"
            f"Order: <code>{order_id}</code>\n"
            f"Amount: <b>{amount:,.2f} NGN</b> → <code>{holder_name}</code>\n"
            f"Transaction ID: <code>{txn_id or 'N/A'}</code>\n"
            f"Reference: <code>{ref}</code>\n"
            f"Bybit marked paid: {'✅' if bybit_ok else '⚠️ Mark manually'}"
        ),
        parse_mode="HTML")


async def _paga_handle_failure(bot, chat_id, order_id, account_no, bank, amount, code, message_txt):
    """Log unpaid order and notify admin on Paga failure."""
    err_lower = (message_txt or "").lower()
    _s(chat_id).unpaid_log.append({
        "order_id":   order_id,
        "account_no": account_no,
        "bank":       bank,
        "amount":     amount,
        "reason":     message_txt or f"Paga responseCode={code}",
        "timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    logger.error(f"[Paga] ❌ FAILED: order={order_id} code={code} msg={message_txt}")
    if "insufficient" in err_lower or "balance" in err_lower or "funds" in err_lower:
        fail_text = (
            f"❌ <b>Paga Failed — Insufficient Funds</b>\n\n"
            f"Order: <code>{order_id}</code>\nAmount needed: <b>{amount:,.2f} NGN</b>\n\n"
            f"👉 Top up your Paga business account balance.\n"
            f"Mark this order manually."
        )
    else:
        fail_text = (
            f"❌ <b>Paga Transfer Failed</b>\n\n"
            f"Order: <code>{order_id}</code>\n"
            f"Code: <code>{code}</code> | Message: <code>{(message_txt or 'Unknown')[:200]}</code>\n\n"
            f"Mark order manually."
        )
    await bot.send_message(chat_id=chat_id, text=fail_text, parse_mode="HTML")


# ─────────────────────────────────────────
# 🟡 PAGA AUTO-PAY
# Flow: Name Match → Buyer Protection → validate account → depositToBank → poll → mark paid
# ─────────────────────────────────────────
async def _paga_autopay(bot, chat_id, order_id, order_detail):
    from paga import match_bank_uuid, validate_account, deposit_to_bank, check_status
    import os

    # Load this user's Paga credentials from DB
    paga_api_key    = db.get_api(chat_id, "paga_api_key")
    paga_credential = db.get_api(chat_id, "paga_credential")
    paga_principal  = db.get_api(chat_id, "paga_principal")

    if not (paga_api_key and paga_credential and paga_principal):
        oid = _esc(order_id)
        await bot.send_message(chat_id=chat_id,
            text=(
                f"❌ <b>Paga Auto-Pay</b> — Order <code>{oid}</code>\n\n"
                "No Paga API configured.\n"
                "Go to 🔑 <b>Set APIs</b> → Set Paga API first."
            ),
            parse_mode="HTML")
        return

    try:
        # ── Name Match check ──
        if _s(chat_id).name_match_enabled:
            has_info, _, _ = _has_account_info(order_detail)
            if not has_info:
                logger.info(f"[Paga NameMatch] Missing info on order {order_id} — marking paid + warn")
                pay_term_nm = order_detail.get("confirmedPayTerm", {}) or {}
                if not pay_term_nm:
                    terms_nm    = order_detail.get("paymentTermList", [])
                    pay_term_nm = terms_nm[0] if terms_nm else {}
                pt  = str(pay_term_nm.get("paymentType", ""))
                pid = str(pay_term_nm.get("id", ""))
                if pt and pid:
                    await asyncio.get_event_loop().run_in_executor(
                        None, partial(mark_order_paid, order_id, pt, pid, creds=get_user_creds(chat_id))
                    )
                    _s(chat_id).paid_order_ids.add(order_id)
                    _track_buy_volume(chat_id, order_id, order_detail)
                await asyncio.get_event_loop().run_in_executor(
                    None, partial(send_chat_message, order_id, NO_ACCOUNT_WARN_MSG,
                                  creds=get_user_creds(chat_id))
                )
                _s(chat_id).expecting_cancel_ids.add(order_id)
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"🔍 <b>Name Match — Missing Info</b>\n\n"
                        f"Order: <code>{order_id}</code>\n"
                        f"Account details incomplete — Paga transfer skipped.\n"
                        f"Marked paid on Bybit + seller asked to cancel."
                    ),
                    parse_mode="HTML")
                return

        pay_term = order_detail.get("confirmedPayTerm", {}) or {}
        if not pay_term:
            terms    = order_detail.get("paymentTermList", [])
            pay_term = terms[0] if terms else {}

        account_no    = pay_term.get("accountNo", "").strip()
        bank_name     = pay_term.get("bankName",  "").strip()
        pay_cfg       = pay_term.get("paymentConfigVo", {}) or pay_term.get("paymentConfig", {}) or {}
        pay_type_name = pay_cfg.get("paymentName", "").strip()
        amount_str    = order_detail.get("amount", "0")
        seller_name   = pay_term.get("realName", order_detail.get("sellerRealName", "Seller"))

        if not account_no:
            oid = _esc(order_id)
            await bot.send_message(chat_id=chat_id,
                text=f"❌ <b>Paga Auto-Pay</b> — Order <code>{oid}</code>\nNo account number found.",
                parse_mode="HTML")
            return

        bank_uuid = match_bank_uuid(bank_name, pay_type_name,
                                    paga_principal, paga_credential, paga_api_key)
        if not bank_uuid:
            oid  = _esc(order_id)
            bank = _esc(bank_name or pay_type_name)
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"❌ <b>Paga Auto-Pay</b> — Order <code>{oid}</code>\n"
                    f"Unknown bank: <code>{bank}</code>\nMark this order manually."
                ),
                parse_mode="HTML")
            return

        amount = float(amount_str)

        # ── Buyer Protection ──
        if _s(chat_id).buyer_protection_on:
            release_mins    = float(order_detail.get("_seller_release_mins", 0))
            release_unknown = bool(order_detail.get("_seller_release_unknown", False))
            if release_unknown or release_mins >= _s(chat_id).buyer_protection_mins:
                reason = (
                    "Seller release time unknown (fetch failed) — flagged as high-risk"
                    if release_unknown else
                    f"Seller avg release time ({release_mins:.0f} min) ≥ threshold ({_s(chat_id).buyer_protection_mins} min)"
                )
                logger.info(f"[Paga BuyerProtection] Skipping — {reason}")
                _s(chat_id).unpaid_log.append({
                    "order_id":   order_id,
                    "account_no": account_no,
                    "bank":       bank_name or pay_type_name,
                    "amount":     amount,
                    "reason":     reason,
                    "timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })
                pay_type   = str(pay_term.get("paymentType", ""))
                payment_id = str(pay_term.get("id", ""))
                if pay_type and payment_id:
                    await asyncio.get_event_loop().run_in_executor(
                        None, partial(mark_order_paid, order_id, pay_type, payment_id, creds=get_user_creds(chat_id))
                    )
                    _s(chat_id).paid_order_ids.add(order_id)
                    _track_buy_volume(chat_id, order_id, order_detail)
                await asyncio.get_event_loop().run_in_executor(
                    None, partial(send_chat_message, order_id, SELLER_WARN_MSG,
                                  creds=get_user_creds(chat_id))
                )
                _s(chat_id).expecting_cancel_ids.add(order_id)
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"🛡 <b>Buyer Protection Triggered</b> — Order <code>{order_id}</code>\n\n"
                        f"Seller release time: <code>{release_mins:.0f} min</code> ≥ <code>{_s(chat_id).buyer_protection_mins} min</code>\n"
                        f"✅ Marked paid on Bybit + warning sent.\n"
                        f"Paga transfer was skipped."
                    ),
                    parse_mode="HTML")
                return

        # ── Step 1: Validate account ──
        await bot.send_message(chat_id=chat_id,
            text=f"⏳ <b>Paga</b> Validating account <code>{_esc(account_no)}</code> ({_esc(bank_name or pay_type_name)})...",
            parse_mode="HTML")

        validate = await asyncio.get_event_loop().run_in_executor(
            None, validate_account, account_no, bank_uuid, amount,
            paga_principal, paga_credential, paga_api_key
        )

        if validate.get("responseCode") != 0 or "error" in validate:
            err = validate.get("message", validate.get("error", "Unknown error"))
            _s(chat_id).unpaid_log.append({
                "order_id":   order_id,
                "account_no": account_no,
                "bank":       bank_name or pay_type_name,
                "amount":     amount,
                "reason":     f"Paga account validation failed: {err}",
                "timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"❌ <b>Paga Account Invalid</b> — Order <code>{order_id}</code>\n\n"
                    f"Account <code>{_esc(account_no)}</code> @ {_esc(bank_name or pay_type_name)} failed validation.\n"
                    f"Reason: <code>{_esc(str(err))}</code>\n\nTransfer aborted. Mark order manually."
                ),
                parse_mode="HTML")
            return

        # Use helper functions that try all known field names (visible in Render logs)
        from paga import _extract_account_name, _extract_fee
        verified_name = _extract_account_name(validate, fallback=seller_name)
        fee           = _extract_fee(validate)
        logger.info(f"[Paga] Validated: {verified_name} | fee={fee}")

        await bot.send_message(chat_id=chat_id,
            text=(
                f"✅ <b>Account Verified</b>: <b>{verified_name}</b>\n"
                f"Account: <code>{account_no}</code> ({bank_name or pay_type_name})\n"
                f"Fee: <b>₦{fee:,.2f}</b>\n\n"
                f"⏳ Sending <b>{amount:,.2f} NGN</b>..."
            ),
            parse_mode="HTML")
        # ── Step 2: Send transfer ──
        render_url   = os.environ.get("RENDER_EXTERNAL_URL", "").rstrip("/")
        callback_url = f"{render_url}/paga-webhook" if render_url else ""
        sender_name  = _s(chat_id).settings.get("sender_name", "Akinrinade Akinniyi")
        ref          = f"p2p{order_id[-16:]}"
        narration    = f"{sender_name[:14]} P2P"   # Paga remarks: 30 char limit

        result = await asyncio.get_event_loop().run_in_executor(
            None, deposit_to_bank,
            account_no, bank_uuid, amount,
            verified_name, "",          # recipient_name, recipient_phone
            narration, callback_url, ref,
            paga_principal, paga_credential, paga_api_key
        )

        if "error" in result:
            err_msg = result["error"]
            ip = await _get_current_ip()
            if "401" in err_msg or "403" in err_msg or "IP" in err_msg:
                oid      = _esc(order_id)
                err_safe = _esc(err_msg[:200])
                ip_safe  = _esc(ip)
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"❌ <b>Paga blocked</b> — Order <code>{oid}</code>\n\n"
                        f"<code>{err_safe}</code>\n\n"
                        f"👉 Whitelist IP <code>{ip_safe}</code> on Paga dashboard → Settings → IP Whitelist"
                    ),
                    parse_mode="HTML")
            else:
                oid      = _esc(order_id)
                err_safe = _esc(err_msg[:300])
                await bot.send_message(chat_id=chat_id,
                    text=f"❌ <b>Paga error</b> — Order <code>{oid}</code>\n<code>{err_safe}</code>",
                    parse_mode="HTML")
            return

        response_code = result.get("responseCode", -1)
        txn_id        = result.get("transactionId", "") or ""
        message_txt   = result.get("message", "") or ""
        from paga import _extract_account_name, check_status
        holder_name   = _extract_account_name(result, fallback=verified_name)

        # ── responseCode meanings from Paga docs ──
        # 0  → SUCCESS (immediate)
        # 3  → PENDING (processing, must poll check_status)
        # anything else → FAILED

        if response_code == 0:
            # Immediate success — mark Bybit paid
            await _paga_handle_success(
                bot, chat_id, order_id, pay_term,
                amount, holder_name, txn_id, ref
            )

        elif response_code == 3 or message_txt.upper() == "PENDING":
            # ── PENDING: poll check_status up to 12×10s = 120 seconds ──
            logger.info(f"[Paga] PENDING — polling check_status for ref={ref}")
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"⏳ <b>Paga Transfer Pending</b>\n\n"
                    f"Order: <code>{order_id}</code>\n"
                    f"Amount: <b>{amount:,.2f} NGN</b> → <code>{holder_name}</code>\n"
                    f"Reference: <code>{ref}</code>\n\n"
                    f"Polling for status update (up to 2 minutes)..."
                ),
                parse_mode="HTML")

            final_code = response_code
            final_msg  = message_txt
            final_txn  = txn_id

            for attempt in range(12):
                await asyncio.sleep(10)
                poll = await asyncio.get_event_loop().run_in_executor(
                    None, check_status, ref,
                    paga_principal, paga_credential, paga_api_key
                )
                final_code = poll.get("responseCode", -1)
                final_msg  = poll.get("message", "") or ""
                final_txn  = poll.get("transactionId", "") or final_txn
                logger.info(
                    f"[Paga] Poll {attempt+1}/12 → code={final_code} "
                    f"msg={final_msg} txnId={final_txn}"
                )
                if final_code == 0:
                    break
                if final_code not in (3, -1) and final_msg.upper() != "PENDING":
                    break  # definitive failure

            if final_code == 0:
                await _paga_handle_success(
                    bot, chat_id, order_id, pay_term,
                    amount, holder_name, final_txn, ref
                )
            elif final_code == 3 or final_msg.upper() == "PENDING":
                # Still pending after 2 min — notify but don't mark failed
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"⏳ <b>Paga Still Pending After 2 Min</b>\n\n"
                        f"Order: <code>{order_id}</code>\n"
                        f"Reference: <code>{ref}</code>\n\n"
                        f"Paga webhook will notify you when complete.\n"
                        f"Check your Paga dashboard if no update arrives.\n"
                        f"Do NOT mark Bybit order paid yet."
                    ),
                    parse_mode="HTML")
            else:
                await _paga_handle_failure(
                    bot, chat_id, order_id,
                    account_no, bank_name or pay_type_name,
                    amount, final_code, final_msg
                )
        else:
            # Immediate failure
            await _paga_handle_failure(
                bot, chat_id, order_id,
                account_no, bank_name or pay_type_name,
                amount, response_code, message_txt
            )

    except Exception as e:
        logger.error(f"[Paga] _paga_autopay error: {e}")
        oid      = _esc(order_id)
        err_safe = _esc(str(e)[:200])
        await bot.send_message(chat_id=chat_id,
            text=f"❌ <b>Paga error</b> — Order <code>{oid}</code>\n<code>{err_safe}</code>",
            parse_mode="HTML")


# ─────────────────────────────────────────
# 💬 CHAT MONITOR — Poll Bybit order chats
# Fetches new messages every 12 seconds for all active orders.
# Forwards new messages to Telegram with a Reply button.
# ─────────────────────────────────────────

def _get_active_order_ids(chat_id: int) -> set:
    """Return all order IDs currently being tracked (buy + sell, not yet released)."""
    sess   = _s(chat_id)
    active = set()
    active.update(sess.seen_order_ids - sess.paid_order_ids)
    for oid in sess.seen_sell_ids:
        if not oid.startswith("paid_") and oid not in sess.released_ids:
            active.add(oid)
    active.update(sess.paid_order_ids)
    return active


async def _poll_order_chat(bot, chat_id: int, order_id: str):
    """
    Fetch latest messages for one order.
    Forward only NEW messages from the counterparty to Telegram.

    Own-message detection uses bybit_uid (set in AD PRICE BOT → Set UID) as the
    primary identifier, checked against both userId and accountId fields in the message.
    Auto-learns accountId and nick from the first matching message for faster future matching.
    """
    try:
        result = await asyncio.get_event_loop().run_in_executor(
            None, partial(get_chat_messages, order_id, "1", "30", creds=get_user_creds(chat_id))
        )
        rc = result.get("retCode", result.get("ret_code", -1))
        if rc != 0:
            return

        # Bybit wraps messages in result.result (list)
        inner    = result.get("result", {})
        messages = inner.get("result", inner) if isinstance(inner, dict) else inner
        if not isinstance(messages, list):
            return

        my_uid     = str(_s(chat_id).settings.get("bybit_uid", "")).strip()
        _chat_msgs = _s(chat_id).seen_chat_msgs   # per-user dict: {order_id: set(msg_ids)}

        if order_id not in _chat_msgs:
            # First poll — learn my accountId and nick by matching bybit_uid
            for m in messages:
                uid  = str(m.get("userId",    ""))
                acct = str(m.get("accountId", ""))
                nck  = str(m.get("nickName",  ""))
                # Match on userId OR accountId
                if my_uid and (uid == my_uid or acct == my_uid):
                    if acct and not _s(chat_id).my_account_id:
                        _s(chat_id).my_account_id = acct
                        logger.info(f"[ChatMonitor] Learned my accountId={acct} nick='{nck}'")
                    if nck and not _s(chat_id).my_nick:
                        _s(chat_id).my_nick = nck
                    break
            # Seed seen IDs — do not forward existing messages on startup
            _chat_msgs[order_id] = {str(m.get("id", "")) for m in messages}
            return

        already_seen = _chat_msgs[order_id]

        # Reverse: messages are newest-first — forward in chronological order
        for msg in reversed(messages):
            msg_id       = str(msg.get("id",              ""))
            msg_type     = int(msg.get("msgType",         0))
            content      = str(msg.get("message",        "")).strip()
            nick         = str(msg.get("nickName",   "Unknown"))
            user_id      = str(msg.get("userId",          ""))
            account_id   = str(msg.get("accountId",       ""))
            role         = str(msg.get("roleType",        ""))
            only_cust    = int(msg.get("onlyForCustomer", 0))

            if msg_id in already_seen:
                continue
            already_seen.add(msg_id)

            # ── Skip system/admin types ──
            if msg_type in (0, 5, 6):
                continue
            if role == "sys":
                continue
            if only_cust == 1:
                continue
            if not content:
                continue

            # ── Primary filter: bybit_uid matches userId OR accountId ──
            # This is the most reliable check — uses the UID you explicitly set
            if my_uid and (user_id == my_uid or account_id == my_uid):
                # Also learn accountId for future faster matching
                if account_id and not _s(chat_id).my_account_id:
                    _s(chat_id).my_account_id = account_id
                if nick and not _s(chat_id).my_nick:
                    _s(chat_id).my_nick = nick
                logger.debug(f"[ChatMonitor] ⏭ Own msg {msg_id} (uid match)")
                continue

            # ── Secondary filter: learned accountId ──
            if _s(chat_id).my_account_id and account_id == _s(chat_id).my_account_id:
                logger.debug(f"[ChatMonitor] ⏭ Own msg {msg_id} (accountId match)")
                continue

            # ── Tertiary filter: learned nick ──
            if _s(chat_id).my_nick and nick == _s(chat_id).my_nick:
                logger.debug(f"[ChatMonitor] ⏭ Own msg {msg_id} (nick match)")
                continue

            # ── This is a counterparty message — forward it ──
            logger.debug(f"[ChatMonitor] ✅ Forwarding msg {msg_id} from '{nick}' (userId={user_id} acctId={account_id})")
            type_label = {1: "💬", 2: "🖼 Image", 7: "📄 PDF", 8: "🎥 Video"}.get(msg_type, "💬")
            display_content = content if len(content) <= 300 else content[:297] + "..."

            text = (
                f"💬 <b>New Bybit Message</b>\n\n"
                f"🆔 Order: <code>{order_id}</code>\n"
                f"👤 From: <b>{nick}</b>\n"
                f"{type_label} _{display_content}_"
            )

            reply_kb = InlineKeyboardMarkup([[
                InlineKeyboardButton(
                    "↩️ Reply",
                    callback_data=f"chatreply_{order_id}_{nick[:20]}"
                )
            ]])

            await bot.send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=reply_kb,
                parse_mode="HTML"
            )
            logger.info(
                f"[ChatMonitor] ✅ Forwarded msg {msg_id} from '{nick}' "
                f"(acctId={account_id}) on order {order_id}"
            )

    except Exception as e:
        logger.error(f"[ChatMonitor] _poll_order_chat {order_id} error: {e}")


async def chat_monitor_loop(bot, chat_id: int):
    """Background loop — polls all active order chats every 8 seconds.

    Rate-limit fix: orders are polled SEQUENTIALLY with a 1.5 s gap between each,
    instead of all at once via asyncio.gather.  With up to ~5 active orders this
    keeps us comfortably below Bybit's 10 read-req/s per-UID limit even when the
    order monitor is also running in parallel.
    """
    # Note: chat_monitor_enabled is set to True by the toggle handler BEFORE
    # this task is created, so the UI reflects the change immediately.
    logger.info("💬 CHAT MONITOR STARTED")

    while _s(chat_id).chat_monitor_enabled:
        try:
            active_ids = _get_active_order_ids(chat_id)
            for oid in active_ids:
                # Stop mid-loop immediately if monitor was disabled
                if not _s(chat_id).chat_monitor_enabled:
                    break
                try:
                    await _poll_order_chat(bot, chat_id, oid)
                except Exception as e:
                    logger.error(f"[ChatMonitor] Poll error for order {oid}: {e}")
                # 1.5 s gap between each order poll — prevents burst spikes
                await asyncio.sleep(1.5)
        except Exception as e:
            logger.error(f"[ChatMonitor] Loop error: {e}")

        await asyncio.sleep(8)

    logger.info("💬 CHAT MONITOR STOPPED")


async def order_monitor_loop(bot, chat_id):
    """
    Per-user order monitor loop. chat_id == user_id in private Telegram chats.

    ISOLATION: Each user gets their own task, their own creds (via get_user_creds),
    their own sess object. No shared state with other users.
    """
    sess = _s(chat_id)
    sess.order_monitor_running = True
    logger.info(f"🔔 ORDER MONITOR STARTED for user {chat_id} (slot {_get_user_slot_str(chat_id)})")

    _ip_error_notified = False   # Track if we already warned this user about IP issue

    while sess.order_monitor_running:
        try:
            # ── Load THIS user's credentials using THEIR slot (not global) ──
            creds = get_user_creds(chat_id)

            # ── Guard: no API key saved for this user's slot ──
            if not is_admin(chat_id) and not creds.get("key"):
                sess.order_monitor_running = False
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        "❌ <b>Order Monitor stopped.</b>\n\n"
                        f"No Bybit API key found for Account {_get_user_slot_str(chat_id)}.\n"
                        "Go to 🔑 <b>Set APIs</b> and add your API key first."
                    ),
                    parse_mode="HTML"
                )
                break

            buy_res, sell_incoming_res, sell_paid_res = await asyncio.gather(
                asyncio.get_event_loop().run_in_executor(None, partial(get_pending_orders, creds=creds)),
                asyncio.get_event_loop().run_in_executor(None, partial(get_incoming_sell_orders, creds=creds)),
                asyncio.get_event_loop().run_in_executor(None, partial(get_sell_orders, creds=creds)),
            )

            # ── IP Whitelist error (10010) — stop polling, notify once ──
            for res, label in [
                (buy_res,          "pending orders"),
                (sell_incoming_res, "incoming sell orders"),
                (sell_paid_res,    "paid sell orders"),
            ]:
                rc  = res.get("retCode", res.get("ret_code", -1))
                msg = res.get("retMsg",  res.get("ret_msg", ""))
                if rc == 10010 or (rc != 0 and "IP" in str(msg).upper()):
                    if not _ip_error_notified:
                        _ip_error_notified = True
                        ip = await _get_current_ip()
                        await bot.send_message(
                            chat_id=chat_id,
                            text=(
                                "🚫 <b>Bybit IP Whitelist Error (10010)</b>\n\n"
                                f"Your API key for Account {_get_user_slot_str(chat_id)} "
                                "is not whitelisted for this server IP.\n\n"
                                f"👉 Add <code>{_esc(ip)}</code> to your Bybit API key's IP whitelist:\n"
                                "Bybit → Account → API Management → Edit Key → Bind IP\n\n"
                                "⚠️ Order monitor has been <b>paused</b> to prevent error spam.\n"
                                "Restart monitoring after whitelisting the IP."
                            ),
                            parse_mode="HTML"
                        )
                        sess.order_monitor_running = False
                    break
            if not sess.order_monitor_running:
                break

            _ip_error_notified = False   # Reset on successful poll

            def _items(res):
                rc = res.get("retCode", res.get("ret_code", -1))
                return res.get("result", {}).get("items", []) if rc == 0 else []

            buy_items       = _items(buy_res)
            sell_incoming   = _items(sell_incoming_res)
            sell_paid_items = _items(sell_paid_res)

            # Build list of (coroutine) for new orders — stored as coroutines NOT tasks
            # so we can run them sequentially and avoid bursting the Bybit rate limit
            # when multiple orders arrive in the same poll cycle.
            tasks = []
            for item in buy_items:
                oid = item.get("id")
                if oid and oid not in sess.seen_order_ids:
                    sess.seen_order_ids.add(oid)
                    tasks.append(_handle_buy_order(bot, chat_id, oid))

            for item in sell_incoming:
                oid = item.get("id")
                if oid and oid not in sess.seen_sell_ids:
                    sess.seen_sell_ids.add(oid)
                    tasks.append(_handle_sell_incoming(bot, chat_id, oid))

            for item in sell_paid_items:
                oid         = item.get("id")
                release_key = f"paid_{oid}"
                if oid and release_key not in sess.seen_sell_ids:
                    sess.seen_sell_ids.add(release_key)
                    tasks.append(_handle_sell_paid(bot, chat_id, oid))

            # Rate-limit fix: run each new-order handler one at a time with a short gap.
            # Each handler makes 2 API calls (get_order_detail + get_counterparty_info).
            # Running them all in parallel when 5 orders arrive = 10 simultaneous calls
            # which reliably triggers Bybit's 10 req/s limit (retCode 10006).
            for coro in tasks:
                if not sess.order_monitor_running:
                    break
                try:
                    await coro
                except Exception as e:
                    logger.error(f"[Orders] Task error for user {chat_id}: {e}")
                # 1.2 s stagger — allows up to ~8 new orders per 10 s cycle
                # before the next poll, well within the 10 req/s read limit.
                if tasks.index(coro) < len(tasks) - 1:
                    await asyncio.sleep(1.2)

        except asyncio.CancelledError:
            logger.info(f"[Orders] Monitor task cancelled for user {chat_id}")
            break
        except Exception as e:
            logger.error(f"[Orders] Loop error for user {chat_id}: {e}")

        # ── Check flagged orders for a seller cancel request ──
        # Per Bybit's official /v5/p2p/order/info docs, status 100/110 mean
        # "objectioning" / "waiting for the user to raise an objection" —
        # that's the APPEAL/DISPUTE flow, not a seller cancel request. That
        # was the actual bug this whole time: we were watching the wrong
        # field. Bybit provides a dedicated boolean for exactly this:
        #
        #   needBuyerExamineCancel: true → a seller cancel application is
        #   pending and the buyer (us) should show the Accept/Reject UI.
        #
        # We check this DIRECTLY on each order we're expecting a cancel
        # from — the same targeted get_order_detail() lookup used
        # everywhere else in the bot — rather than scanning a broad
        # status-filtered list.
        #
        # This only runs for orders the bot itself flagged (buyer-protection /
        # slow-release / name-match) and marked paid + warned the seller about.
        # If the seller never requests a cancel on a flagged order, nothing
        # happens — we just keep checking it each cycle until it resolves.
        expecting = list(_s(chat_id).expecting_cancel_ids)
        for oid in expecting:
            if oid in _s(chat_id).pending_cancel_reviews:
                continue   # already notified, waiting on the user's Accept/Reject
            try:
                od = await asyncio.get_event_loop().run_in_executor(
                    None, partial(get_order_detail, oid, creds=creds)
                )
                if od.get("retCode", -1) != 0:
                    logger.debug(f"[CancelPoll] Could not fetch order {oid} for {chat_id}: {od.get('retMsg')}")
                    continue
                result = od.get("result", {})
                status = str(result.get("status", ""))
                if result.get("needBuyerExamineCancel"):
                    asyncio.create_task(
                        _handle_seller_cancel_request(bot, chat_id, oid)
                    )
                elif status not in ("10", "20") and not result.get("needBuyerExamineCancel"):
                    # Order left the normal in-flight window (10=waiting buy pay,
                    # 20=waiting seller release) some other way — appeal, or it
                    # simply completed/cancelled without a cancel review ever
                    # being needed. Stop tracking it so we don't poll forever.
                    _s(chat_id).expecting_cancel_ids.discard(oid)
            except Exception as _ce:
                logger.debug(f"[CancelPoll] Error checking order {oid} for {chat_id}: {_ce}")

        await asyncio.sleep(10)

    sess.order_monitor_running = False
    logger.info(f"🔕 ORDER MONITOR STOPPED for user {chat_id}")


# ─────────────────────────────────────────
# 🚫 SELLER CANCEL REQUEST HANDLER
# ─────────────────────────────────────────

def _cancel_review_buttons(order_id: str) -> InlineKeyboardMarkup:
    """Inline buttons for seller cancel review: Accept or Reject."""
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Accept Cancellation", callback_data=f"sc_accept_{order_id}"),
            InlineKeyboardButton("❌ Reject",              callback_data=f"sc_reject_{order_id}"),
        ]
    ])


def _cancel_reject_reason_buttons(order_id: str) -> InlineKeyboardMarkup:
    """Inline buttons for rejection reason selection."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(
            "💸 I have already made payment",
            callback_data=f"sc_reason_{order_id}_haveMadePayment"
        )],
        [InlineKeyboardButton(
            "💰 I have not received a full refund",
            callback_data=f"sc_reason_{order_id}_haveNotReceivedFullRefund"
        )],
        [InlineKeyboardButton(
            "📝 Other reason",
            callback_data=f"sc_reason_{order_id}_others"
        )],
    ])


async def _handle_seller_cancel_request(bot, chat_id: int, order_id: str):
    """
    Called when order_monitor_loop sees needBuyerExamineCancel=true on a
    flagged buy order (i.e. one the bot itself marked paid + warned the
    seller about).
    Fetches full order details, builds the notification with original flag reason
    (buyer protection slow-seller), then sends Accept / Reject buttons to the user.
    Per-user: uses chat_id for creds and session state.
    """
    sess = _s(chat_id)

    # Guard: only process once per order
    if order_id in sess.pending_cancel_reviews:
        return

    try:
        # Fetch full order detail for payment info
        det = await asyncio.get_event_loop().run_in_executor(
            None, partial(get_order_detail, order_id, creds=get_user_creds(chat_id))
        )
        if det.get("retCode", -1) != 0:
            logger.warning(f"[CancelReview] Could not fetch order {order_id}: {det.get('retMsg')}")
            return
        order_detail = det.get("result", {})

        # Fetch seller info for release time display
        seller_uid  = order_detail.get("targetUserId", "")
        seller_info = {}
        if seller_uid:
            si = await asyncio.get_event_loop().run_in_executor(
                None, partial(get_counterparty_info, str(seller_uid), order_id,
                              creds=get_user_creds(chat_id))
            )
            if si.get("retCode", -1) == 0:
                seller_info = si.get("result", {})

        # Build flag reason — reuse buyer protection data if available
        try:
            release_mins = float(seller_info.get("averageReleaseTime", "0") or 0)
        except (ValueError, TypeError):
            release_mins = 0.0

        thresh = sess.buyer_protection_mins if sess.buyer_protection_on else 0
        if sess.buyer_protection_on and release_mins >= thresh:
            flag_reason = (
                f"Seller avg release time: {release_mins:.0f} min "
                f"≥ your threshold: {thresh} min"
            )
        else:
            flag_reason = "Seller requested order cancellation"

        # Store in session so the button handler can retrieve it
        sess.pending_cancel_reviews[order_id] = {
            "order_detail": order_detail,
            "seller_info":  seller_info,
            "flag_reason":  flag_reason,
        }

        # Format the order details the same way as the original buy order message
        order_detail["_seller_release_mins"] = release_mins
        msg = format_order_message(order_detail, seller_info, uid=chat_id)

        quantity = order_detail.get("quantity", "—")
        amount   = order_detail.get("amount",   "—")
        currency = order_detail.get("currencyId", "—")

        await bot.send_message(
            chat_id=chat_id,
            text=(
                f"🚨 <b>Seller Cancel Request</b>\n\n"
                f"The seller has requested to cancel this buy order.\n\n"
                f"⚠️ <b>Reason flagged:</b> {_esc(flag_reason)}\n\n"
                f"{msg}\n\n"
                f"<b>What would you like to do?</b>\n"
                f"• <b>Accept</b> — allow the seller to cancel the order\n"
                f"• <b>Reject</b> — refuse the cancellation and choose a reason"
            ),
            reply_markup=_cancel_review_buttons(order_id),
            parse_mode="HTML"
        )
        logger.info(
            f"[CancelReview] Notified user {chat_id} about cancel request "
            f"for order {order_id} | reason: {flag_reason}"
        )

    except Exception as e:
        logger.error(f"[CancelReview] Error handling cancel request {order_id} for user {chat_id}: {e}")


async def _handle_cancel_review(bot, chat_id: int, order_id: str,
                                 examine_result: str, reject_reason_key: str = ""):
    """
    Execute the buyer's decision (PASS or REJECT) via Bybit API.
    Called from button_handler after the user taps Accept or a Reject reason.
    """
    sess = _s(chat_id)
    creds = get_user_creds(chat_id)

    # Map short key → full Bybit reason string
    reason_map = {
        "haveMadePayment":          "buyerRefuseOrderCancelReason_haveMadePayment",
        "haveNotReceivedFullRefund": "buyerRefuseOrderCancelReason_haveNotReceivedFullRefund",
        "others":                   "buyerRefuseOrderCancelReason_others",
    }
    full_reason = reason_map.get(reject_reason_key, "") if reject_reason_key else ""

    try:
        result = await asyncio.get_event_loop().run_in_executor(
            None, partial(
                review_seller_cancel,
                order_id,
                examine_result,
                full_reason,
                "",   # rejectProofs — not required for text reasons
                "",   # rejectRemark
                creds=creds,
            )
        )
        ret_code = result.get("retCode", -1)
        ret_msg  = result.get("retMsg", "")

        if ret_code == 0:
            # Clean up session
            sess.pending_cancel_reviews.pop(order_id, None)
            sess.expecting_cancel_ids.discard(order_id)
            if examine_result == "PASS":
                # Order will be cancelled — mark as finalized
                _set_order_final(order_id, "cancelled")
                sess.paid_order_ids.discard(order_id)
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"✅ <b>Cancellation Accepted</b>\n\n"
                        f"Order <code>{_esc(order_id)}</code> has been cancelled.\n"
                        f"The seller's cancellation request was approved."
                    ),
                    parse_mode="HTML"
                )
            else:
                # Readable label for the reason
                reason_labels = {
                    "haveMadePayment":           "I have already made payment",
                    "haveNotReceivedFullRefund":  "I have not received a full refund",
                    "others":                    "Other reason",
                }
                reason_label = reason_labels.get(reject_reason_key, reject_reason_key)
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"❌ <b>Cancellation Rejected</b>\n\n"
                        f"Order <code>{_esc(order_id)}</code> — the seller's cancellation "
                        f"request has been rejected.\n\n"
                        f"<b>Reason given:</b> {_esc(reason_label)}\n\n"
                        f"The order remains active. Bybit will handle the dispute."
                    ),
                    parse_mode="HTML"
                )
            logger.info(
                f"[CancelReview] {examine_result} for order {order_id} "
                f"user={chat_id} reason={full_reason!r} retCode=0"
            )
        else:
            await bot.send_message(
                chat_id=chat_id,
                text=(
                    f"❌ <b>Cancel Review Failed</b>\n\n"
                    f"Order: <code>{_esc(order_id)}</code>\n"
                    f"Error: <code>{_esc(ret_msg)}</code>\n\n"
                    f"Please try again or handle manually on Bybit."
                ),
                parse_mode="HTML"
            )
            logger.warning(
                f"[CancelReview] API error | order={order_id} user={chat_id} "
                f"examine={examine_result} retCode={ret_code} msg={ret_msg!r}"
            )
    except Exception as e:
        logger.error(f"[CancelReview] _handle_cancel_review error: {e}")
        await bot.send_message(
            chat_id=chat_id,
            text=(
                f"❌ <b>Error processing cancel review</b>\n"
                f"<code>{_esc(str(e)[:200])}</code>"
            ),
            parse_mode="HTML"
        )


def _track_buy_volume(uid: int, order_id: str, order_detail: dict):
    """Add this buy order's coin quantity (quantity/tokenId — NOT the fiat
    amount/currencyId) to the user's rolling 24h buy-volume analytics.
    Safe to call from every pay path — idempotent per order_id, so it will
    never double-count even if called more than once for the same order."""
    try:
        token = str(order_detail.get("tokenId", "")).upper().strip()
        qty   = order_detail.get("quantity", "0")
        if token and qty:
            _s(uid).record_buy_volume(order_id, token, qty)
    except Exception as _e:
        logger.debug(f"[BuyVolume] Failed to record order {order_id} for user {uid}: {_e}")


async def _handle_buy_order(bot, chat_id, order_id):
    try:
        det = await asyncio.get_event_loop().run_in_executor(None, partial(get_order_detail, order_id, creds=get_user_creds(chat_id)))
        if det.get("retCode", -1) != 0:
            return
        order_detail = det.get("result", {})
        seller_uid   = order_detail.get("targetUserId", "")

        # ── Fetch seller counterparty info (used for buyer-protection release-time check) ──
        # IMPORTANT: this call can fail transiently (rate limit, timeout). Previously a
        # failure here silently left seller_info={} for the rest of the function, which
        # made release_mins default to 0 and silently bypassed buyer protection entirely
        # — the order would get auto-paid with no warning even if the seller's real
        # release time was high. We now retry once and, if still unavailable, flag it
        # as unknown rather than silently treating it as "safe" (0 min).
        seller_info = {}
        seller_info_unknown = False
        if seller_uid:
            si = await asyncio.get_event_loop().run_in_executor(
                None, partial(get_counterparty_info, str(seller_uid), order_id, creds=get_user_creds(chat_id))
            )
            if si.get("retCode", -1) == 0:
                seller_info = si.get("result", {})
            else:
                logger.warning(
                    f"[BuyOrder] get_counterparty_info failed for order {order_id} "
                    f"user={chat_id} seller={seller_uid}: retCode={si.get('retCode')} "
                    f"msg={si.get('retMsg','')!r} — retrying once"
                )
                await asyncio.sleep(2)
                si_retry = await asyncio.get_event_loop().run_in_executor(
                    None, partial(get_counterparty_info, str(seller_uid), order_id, creds=get_user_creds(chat_id))
                )
                if si_retry.get("retCode", -1) == 0:
                    seller_info = si_retry.get("result", {})
                else:
                    seller_info_unknown = True
                    logger.warning(
                        f"[BuyOrder] get_counterparty_info retry also failed for order "
                        f"{order_id} user={chat_id}: retCode={si_retry.get('retCode')} "
                        f"msg={si_retry.get('retMsg','')!r} — release time unknown, "
                        f"treating as high-risk for buyer protection"
                    )

        # Populate payment name cache for this user so Type XXX is resolved to real names
        creds_for_map = get_user_creds(chat_id)
        if chat_id not in _payment_name_cache or not _payment_name_cache[chat_id]:
            fetched_map = await asyncio.get_event_loop().run_in_executor(
                None, partial(_fetch_user_payment_map, creds_for_map)
            )
            if fetched_map:
                _payment_name_cache[chat_id] = fetched_map

        msg = format_order_message(order_detail, seller_info, uid=chat_id)
        sent_msg = await bot.send_message(
            chat_id=chat_id,
            text=f"🛒 <b>BUY Order — Pay Seller</b>\n{msg}",
            reply_markup=order_buttons(order_id),
            parse_mode="HTML"
        )
        # Store message_id so auto-pay can remove buttons without a query object
        _s(chat_id).order_msg_ids[order_id] = sent_msg.message_id

        # ── Persist cumulative buy order count to DB ──
        try:
            user_rec = db.get_user(chat_id)
            if user_rec is not None:
                new_buy_count = (user_rec.get("total_buy_orders") or 0) + 1
                db.update_user_stats(chat_id, total_buy_orders=new_buy_count)
        except Exception as _stat_err:
            logger.debug(f"[Stats] Could not update buy count for {chat_id}: {_stat_err}")

        # ── Name Match check (Bybit auto-pay path) ──
        if _s(chat_id).name_match_enabled and (_s(chat_id).auto_pay_enabled or _s(chat_id).flw_pay_enabled or _s(chat_id).paga_pay_enabled):
            has_info, _, _ = _has_account_info(order_detail)
            if not has_info and order_id not in _s(chat_id).paid_order_ids:
                pay_term_nm = order_detail.get("confirmedPayTerm", {}) or {}
                if not pay_term_nm:
                    terms_nm    = order_detail.get("paymentTermList", [])
                    pay_term_nm = terms_nm[0] if terms_nm else {}
                pt  = str(pay_term_nm.get("paymentType", ""))
                pid = str(pay_term_nm.get("id", ""))
                if pt and pid:
                    await asyncio.get_event_loop().run_in_executor(
                        None, partial(mark_order_paid, order_id, pt, pid,
                                      creds=get_user_creds(chat_id))
                    )
                    _s(chat_id).paid_order_ids.add(order_id)
                    _track_buy_volume(chat_id, order_id, order_detail)
                    await _update_order_message_final(bot, chat_id, order_id, "Warning Sent", "warned")
                await asyncio.get_event_loop().run_in_executor(
                    None, partial(send_chat_message, order_id, NO_ACCOUNT_WARN_MSG,
                                  creds=get_user_creds(chat_id))
                )
                _s(chat_id).expecting_cancel_ids.add(order_id)
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"🔍 <b>Name Match — Missing Info</b>\n\n"
                        f"Order <code>{order_id}</code>\nNo account details found.\n"
                        f"Marked paid + seller asked to cancel."
                    ),
                    parse_mode="HTML")
                return

        # ── compute seller release time once (shared by all pay paths) ──
        try:
            seller_release = float(seller_info.get("averageReleaseTime", "0") or 0)
        except (ValueError, TypeError):
            seller_release = 0
        order_detail["_seller_release_mins"]    = seller_release
        order_detail["_seller_release_unknown"] = seller_info_unknown

        if _s(chat_id).paga_pay_enabled and order_id not in _s(chat_id).paid_order_ids:
            await asyncio.sleep(5)
            # ── Enqueue instead of calling directly ──
            # This ensures orders are paid one at a time, preventing
            # Paga rate-limit failures when multiple orders arrive at once.
            pos = _enqueue_paga_order(bot, chat_id, order_id, order_detail)
            if pos > 1:
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"🟡 <b>Paga Queue</b> — Order <code>{order_id}</code> added\n"
                        f"📋 Position: <code>{_esc(str(pos))}</code> in queue\n"
                        f"Will be processed after the current order completes."
                    ),
                    parse_mode="HTML"
                )

        elif _s(chat_id).flw_pay_enabled and order_id not in _s(chat_id).paid_order_ids:
            await asyncio.sleep(5)
            await _flw_autopay(bot, chat_id, order_id, order_detail)

        elif _s(chat_id).auto_pay_enabled and order_id not in _s(chat_id).paid_order_ids:
            try:
                release_mins = float(seller_info.get("averageReleaseTime", "0") or 0)
            except (ValueError, TypeError):
                release_mins = 0

            # ── Buyer Protection check BEFORE marking paid ──
            # If we couldn't confirm the seller's release time (API fetch failed
            # twice), fail safe: treat it as if it's above your threshold so the
            # order gets the warning + review flow instead of silently auto-paying
            # with zero visibility into the actual release time.
            bp_triggered = _s(chat_id).buyer_protection_on and (
                seller_info_unknown or release_mins >= _s(chat_id).buyer_protection_mins
            )
            if bp_triggered:
                pay_term_bp = order_detail.get("confirmedPayTerm", {}) or {}
                if not pay_term_bp:
                    terms_bp    = order_detail.get("paymentTermList", [])
                    pay_term_bp = terms_bp[0] if terms_bp else {}
                pt_bp  = str(pay_term_bp.get("paymentType", ""))
                pid_bp = str(pay_term_bp.get("id", ""))
                if pt_bp and pid_bp and order_id not in _s(chat_id).paid_order_ids:
                    pr_bp = await asyncio.get_event_loop().run_in_executor(
                        None, partial(mark_order_paid, order_id, pt_bp, pid_bp,
                                      creds=get_user_creds(chat_id))
                    )
                    if pr_bp.get("retCode", -1) == 0:
                        _s(chat_id).paid_order_ids.add(order_id)
                        _track_buy_volume(chat_id, order_id, order_detail)
                        # Use the "warned" badge, not "completed" — this order was
                        # flagged by buyer protection, it wasn't a plain successful
                        # auto-pay. The badge should make that visible at a glance.
                        await _update_order_message_final(bot, chat_id, order_id, "Warning Sent", "warned")
                await asyncio.get_event_loop().run_in_executor(
                    None, partial(send_chat_message, order_id, SELLER_WARN_MSG,
                                  creds=get_user_creds(chat_id))
                )
                _s(chat_id).expecting_cancel_ids.add(order_id)
                bp_reason = (
                    "Buyer Protection: seller release time unknown (fetch failed) — flagged as high-risk"
                    if seller_info_unknown else
                    f"Buyer Protection: seller release {release_mins:.0f} min ≥ {_s(chat_id).buyer_protection_mins} min"
                )
                _s(chat_id).unpaid_log.append({
                    "order_id":   order_id,
                    "account_no": str(pay_term_bp.get("accountNo","—")),
                    "bank":       get_payment_name(str(pay_term_bp.get("paymentType",""))),
                    "amount":     float(order_detail.get("amount","0")),
                    "reason":     bp_reason,
                    "timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"🛡 <b>Buyer Protection</b> — Order <code>{_esc(order_id)}</code>\n\n"
                        + (
                            "Seller release time: <code>unknown</code> (info fetch failed twice — flagged as high-risk)\n"
                            if seller_info_unknown else
                            f"Seller release: <code>{release_mins:.0f} min</code> ≥ <code>{_s(chat_id).buyer_protection_mins} min</code>\n"
                        )
                        + "✅ Marked paid on Bybit + warning sent to seller."
                    ),
                    parse_mode="HTML"
                )
            else:
                # ── Normal auto-pay path ──
                await asyncio.sleep(5)   # brief delay before marking

                # Re-fetch order to confirm it is still unpaid
                recheck = await asyncio.get_event_loop().run_in_executor(
                    None, partial(get_order_detail, order_id, creds=get_user_creds(chat_id))
                )
                if recheck.get("retCode", -1) != 0:
                    logger.warning(f"[AutoPay] Could not re-fetch order {order_id} — skipping")
                    return
                recheck_detail = recheck.get("result", {})
                # Bybit order status: 10=pending, 20=paid, 30=done, 40=cancelled
                if str(recheck_detail.get("status","")) not in ("10",):
                    logger.info(f"[AutoPay] Order {order_id} already processed (status={recheck_detail.get('status')}) — skipping")
                    return
                if order_id in _s(chat_id).paid_order_ids:
                    return   # already handled by a parallel path

                pay_term = recheck_detail.get("confirmedPayTerm", {}) or {}
                if not pay_term:
                    terms    = recheck_detail.get("paymentTermList", [])
                    pay_term = terms[0] if terms else {}

                payment_type = str(pay_term.get("paymentType", ""))
                payment_id   = str(pay_term.get("id", ""))

                if payment_type and payment_id:
                    pr = await asyncio.get_event_loop().run_in_executor(
                        None, partial(mark_order_paid, order_id, payment_type, payment_id,
                                      creds=get_user_creds(chat_id))
                    )
                    if pr.get("retCode", -1) == 0:
                        _s(chat_id).paid_order_ids.add(order_id)
                        _track_buy_volume(chat_id, order_id, order_detail)
                        await _remove_order_buttons(bot, chat_id, order_id)
                        await bot.send_message(
                            chat_id=chat_id,
                            text=f"💳 <b>Auto-Pay ✅</b> Order <code>{_esc(order_id)}</code> marked paid.",
                            parse_mode="HTML"
                        )
                    else:
                        await bot.send_message(
                            chat_id=chat_id,
                            text=(
                                f"❌ <b>Auto-Pay failed</b> — Order <code>{_esc(order_id)}</code>\n"
                                f"<code>{_esc(pr.get('retMsg',''))}</code>"
                            ),
                            parse_mode="HTML"
                        )
    except Exception as e:
        logger.error(f"[BUY] _handle_buy_order {order_id} error: {e}")


async def _handle_sell_incoming(bot, chat_id, order_id):
    try:
        det = await asyncio.get_event_loop().run_in_executor(None, partial(get_order_detail, order_id, creds=get_user_creds(chat_id)))
        if det.get("retCode", -1) != 0:
            return
        order_detail = det.get("result", {})
        buyer_uid    = order_detail.get("targetUserId", "")

        buyer_info = {}
        if buyer_uid:
            bi = await asyncio.get_event_loop().run_in_executor(
                None, partial(get_counterparty_info, str(buyer_uid), order_id, creds=get_user_creds(chat_id))
            )
            if bi.get("retCode", -1) == 0:
                buyer_info = bi.get("result", {})

        msg = format_sell_order_message(order_detail, buyer_info, uid=chat_id)
        await bot.send_message(
            chat_id=chat_id,
            text=f"💰 <b>SELL Order — Awaiting Buyer Payment</b>\n{msg}",
            parse_mode="HTML"
        )

        # ── 🚨 Fraud Check (SELL orders only) ──
        # Try every possible field Bybit may use for buyer name
        buyer_name = (
            order_detail.get("buyerRealName", "").strip()
            or order_detail.get("targetRealName", "").strip()
            or buyer_info.get("realName", "").strip()
            or buyer_info.get("nickName", "").strip()
            or ""
        )

        # Always show verification status so you know it ran
        scammer_count = get_scammer_count()
        if not buyer_name:
            await bot.send_message(
                chat_id=chat_id,
                text=(
                    f"🔍 <b>Fraud Check — Order <code>{order_id}</code></b>\n\n"
                    f"⚠️ Buyer name not available yet at this stage.\n"
                    f"Name will be checked again when buyer pays (status 20).\n"
                    f"_(Database: {scammer_count} names loaded)_"
                ),
                parse_mode="HTML"
            )
        else:
            await bot.send_message(
                chat_id=chat_id,
                text=f"🔍 <b>Verifying buyer name...</b>\n👤 <code>{buyer_name}</code>",
                parse_mode="HTML"
            )
            fraud = await asyncio.get_event_loop().run_in_executor(
                None, check_buyer_name, buyer_name
            )
            if fraud["flagged"]:
                match_label = {
                    "exact":   "🔴 Exact match",
                    "partial": "🟠 Partial match",
                    "fuzzy":   "🟡 Similar name",
                }.get(fraud["match_type"], "⚠️ Match")
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"🚨 <b>FRAUD WARNING — Order <code>{order_id}</code></b>\n\n"
                        f"👤 Buyer: <b>{buyer_name}</b>\n"
                        f"{match_label}: <code>{fraud['matched_name']}</code>\n"
                        f"Similarity: <code>{fraud['similarity']:.0%}</code>\n\n"
                        f"⛔ <b>Do NOT accept payment from this buyer.</b>\n"
                        f"Fraudulent / chargeback records found.\n\n"
                        f"👉 Request order cancellation immediately."
                    ),
                    parse_mode="HTML"
                )
                logger.warning(
                    f"[FraudCheck] 🚨 FLAGGED {order_id} | buyer='{buyer_name}' "
                    f"matched='{fraud['matched_name']}' type={fraud['match_type']}"
                )
            else:
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"✅ <b>Buyer Verified — Not in fraud list</b>\n\n"
                        f"👤 <code>{buyer_name}</code>\n"
                        f"_(Checked against {scammer_count} names)_"
                    ),
                    parse_mode="HTML"
                )
                logger.info(f"[FraudCheck] ✅ Clean: '{buyer_name}' on order {order_id}")

        # ── Custom sell message ──
        if _s(chat_id).sell_msg_enabled and _s(chat_id).sell_custom_msg:
            for i in range(_s(chat_id).sell_msg_count):
                await asyncio.get_event_loop().run_in_executor(
                    None, partial(send_chat_message, order_id, _s(chat_id).sell_custom_msg,
                                  creds=get_user_creds(chat_id))
                )
                if i < _s(chat_id).sell_msg_count - 1:
                    await asyncio.sleep(1)
    except Exception as e:
        logger.error(f"[SELL incoming] {order_id} error: {e}")


async def _handle_sell_paid(bot, chat_id, order_id):
    try:
        det = await asyncio.get_event_loop().run_in_executor(None, partial(get_order_detail, order_id, creds=get_user_creds(chat_id)))
        if det.get("retCode", -1) != 0:
            return
        order_detail = det.get("result", {})
        buyer_uid    = order_detail.get("targetUserId", "")

        buyer_info = {}
        if buyer_uid:
            bi = await asyncio.get_event_loop().run_in_executor(
                None, partial(get_counterparty_info, str(buyer_uid), order_id, creds=get_user_creds(chat_id))
            )
            if bi.get("retCode", -1) == 0:
                buyer_info = bi.get("result", {})

        msg = format_sell_order_message(order_detail, buyer_info, uid=chat_id)
        await bot.send_message(
            chat_id=chat_id,
            text=f"✅ <b>SELL Order — Buyer Has Paid! Release Coin Now</b>\n{msg}",
            reply_markup=sell_order_buttons(order_id),
            parse_mode="HTML"
        )

        # ── Persist cumulative sell order count to DB ──
        try:
            user_rec = db.get_user(chat_id)
            if user_rec is not None:
                new_sell_count = (user_rec.get("total_sell_orders") or 0) + 1
                db.update_user_stats(chat_id, total_sell_orders=new_sell_count)
        except Exception as _stat_err:
            logger.debug(f"[Stats] Could not update sell count for {chat_id}: {_stat_err}")

        # ── 🚨 Fraud Check at paid stage (buyer name most reliable here) ──
        buyer_name = (
            order_detail.get("buyerRealName", "").strip()
            or order_detail.get("targetRealName", "").strip()
            or buyer_info.get("realName", "").strip()
            or ""
        )
        if buyer_name:
            scammer_count = get_scammer_count()
            await bot.send_message(
                chat_id=chat_id,
                text=f"🔍 <b>Verifying buyer name before release...</b>\n👤 <code>{buyer_name}</code>",
                parse_mode="HTML"
            )
            fraud = await asyncio.get_event_loop().run_in_executor(
                None, check_buyer_name, buyer_name
            )
            if fraud["flagged"]:
                match_label = {
                    "exact":   "🔴 Exact match",
                    "partial": "🟠 Partial match",
                    "fuzzy":   "🟡 Similar name",
                }.get(fraud["match_type"], "⚠️ Match")
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"🚨 <b>FRAUD WARNING — DO NOT RELEASE</b>\n\n"
                        f"Order: <code>{order_id}</code>\n"
                        f"👤 Buyer: <b>{buyer_name}</b>\n"
                        f"{match_label}: <code>{fraud['matched_name']}</code>\n"
                        f"Similarity: <code>{fraud['similarity']:.0%}</code>\n\n"
                        f"⛔ <b>Do NOT release coins to this buyer.</b>\n"
                        f"Fraudulent / chargeback records found.\n\n"
                        f"👉 Open a dispute or request cancellation."
                    ),
                    parse_mode="HTML"
                )
                logger.warning(
                    f"[FraudCheck] 🚨 PAID-STAGE FLAGGED {order_id} | "
                    f"buyer='{buyer_name}' matched='{fraud['matched_name']}'"
                )
            else:
                await bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"✅ <b>Buyer Verified — Not in fraud list</b>\n\n"
                        f"👤 <code>{buyer_name}</code>\n"
                        f"_(Checked against {scammer_count} names)_\n\n"
                        f"Safe to release coins."
                    ),
                    parse_mode="HTML"
                )

    except Exception as e:
        logger.error(f"[SELL paid] {order_id} error: {e}")


# ─────────────────────────────────────────
# 💲 Float price calc
# ─────────────────────────────────────────
def _extract_bybit_max(error_msg: str) -> str | None:
    import re
    match = re.search(r'higher than ([\d.]+)', error_msg)
    if match:
        return match.group(1).rstrip(".")
    return None


def _extract_bybit_bounds(error_msg: str):
    """
    Parse BOTH bounds from Bybit's out-of-range message, e.g.
    'The fixed price set is lower than X or higher than Y.'
    Returns (min_str, max_str) — either may be None if not present in
    this particular message.

    Matches optional thousands-separator commas too (e.g. "79,126,068.63")
    and strips them before returning — NGN bounds are large enough that if
    Bybit ever formats them with commas, a comma-blind regex would match
    only the digits up to the first comma (e.g. "79" instead of
    "79126068.63") and silently post a price a thousand times too small.
    """
    import re
    min_match = re.search(r'lower than ([\d,]*\d(?:\.\d+)?)',  error_msg or "")
    max_match = re.search(r'higher than ([\d,]*\d(?:\.\d+)?)', error_msg or "")
    min_val = min_match.group(1).replace(",", "").rstrip(".") if min_match else None
    max_val = max_match.group(1).replace(",", "").rstrip(".") if max_match else None
    return min_val, max_val


def _safety_margin(price: Decimal) -> Decimal:
    """
    Small buffer used when retrying at a Bybit-stated boundary, so we land
    safely INSIDE the valid range instead of riding exactly on the edge.
    Posting the exact boundary value sometimes gets rejected again if
    Bybit's live min/max shifted a hair between it telling us the number
    and us submitting it (normal market movement over the round-trip).
    0.02% of price, with a small currency-agnostic floor so it's never
    effectively zero on cheap-priced pairs.
    """
    margin = price * Decimal("0.0002")
    return margin if margin > Decimal("0.01") else Decimal("0.01")


def _wants_live_ceiling(currency_id: str, token_id: str, float_pct) -> bool:
    """
    True when the floating % is set to the very top of what this bot
    allows for this pair, or one point below it (e.g. 110% or 111% for
    NGN/BTC, 130% or 131% for USD/BTC — get_max_float_pct(NGN,BTC) is 111,
    get_max_float_pct(USD,BTC) is 131).

    At that setting, the user's real intent is "give me the highest price
    Bybit will currently let me list at" — not literally "market price
    times this exact multiplier". A fixed multiplier can drift from what
    Bybit will actually accept as the live order book moves, so instead of
    trusting the formula number, the caller (auto_update_loop) deliberately
    forces an out-of-range rejection every cycle so Bybit hands back its
    OWN live ceiling in the error message, then posts that instead. See
    the 912120022 handling in auto_update_loop for where that gets used.
    """
    try:
        pct = float(float_pct)
    except (TypeError, ValueError):
        return False
    hi = get_max_float_pct(currency_id, token_id)
    return hi - 1 <= pct <= hi


def calc_floating_price(ad_data, float_pct, local_usdt_ref):
    """
    Calculate floating price for any supported currency/token pair.

    Formula:
      NGN/USD:  token_usdt_price × local_usdt_ref × float_pct / 100
      GHS/GBP/EUR/RUB/KES:  token_usdt_price × local_usdt_ref × float_pct / 100
      (same formula — local_usdt_ref is the local currency per 1 USDT rate)

    For USDT/USDC pairs the ref is not needed (token IS the dollar).
    """
    currency = ad_data.get("currencyId", "").upper()
    token    = ad_data.get("tokenId",   "").upper()

    token_price = get_token_usdt_price(token)
    if token_price <= 0:
        return None, f"Failed to fetch {token}/USDT price from Bybit"

    # Currencies that need a local/USDT reference rate
    needs_ref = currency_needs_ref(currency) or currency == "NGN"

    if needs_ref:
        if local_usdt_ref <= 0:
            return None, f"{currency}/USDT reference price not set — tap 💱 Set {currency}/USDT Ref"
        raw = token_price * local_usdt_ref * float_pct / 100
    elif currency == "USD":
        # USD: token_price already in USD
        raw = token_price * float_pct / 100
    else:
        # Unknown currency — treat as direct
        raw = token_price * float_pct / 100

    return str(Decimal(str(raw)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)), None


async def _try_fast_chase(bot, chat_id, sess, slot_idx, ad_data, s, float_pct, creds, _quant):
    """
    Runs every 10 seconds while Ad 1 waits out the rest of its scheduled
    interval (floating mode, single-ad users only — see auto_update_loop,
    which gates this to slot_idx == -1 and sess.total_ad_slots() == 1
    before ever calling it, so a user with 2+ ads never triggers this and
    can't out-update everyone else on the same pair).

    If the live reference price has risen by at least this pair's minimum
    gap since the last post, post it right away instead of waiting for the
    next scheduled cycle. Pulls from the SAME rolling modify budget as the
    scheduled cycle (_can_modify_ad1/_record_modify_ad1) so the two
    together can never exceed Bybit's 10-modifies-per-5-minutes limit on a
    single ad.

    Deliberately minimal error handling: this is a bonus opportunistic
    update, not a scheduled cycle, so any failure here is silently
    skipped rather than counted towards the auto-stop failure counter —
    the next tick (10s) or the regular scheduled cycle picks it up either
    way. Only moves that increase the price are chased (a seller wants a
    higher price sooner, not a lower one).
    """
    if not _can_modify_ad1(sess):
        return
    local_usdt_ref = float(sess.shared_local_usdt_ref or s.get("local_usdt_ref") or 0)
    new_p_str, err = calc_floating_price(ad_data, float_pct, local_usdt_ref)
    if err:
        return
    new_p = Decimal(new_p_str)
    cur_p = _ad_current_price(sess, slot_idx)
    if cur_p <= 0:
        return
    gap = get_min_price_gap(ad_data.get("currencyId",""), ad_data.get("tokenId",""), new_p)
    if new_p - cur_p < gap:
        return   # hasn't risen enough to be worth a fresh post yet

    new_p_str = str(new_p.quantize(_quant, rounding=ROUND_HALF_UP))
    _record_modify_ad1(sess)
    result = await asyncio.get_event_loop().run_in_executor(
        _ad_executor, modify_ad, s["ad_id"], new_p_str, ad_data, creds
    )
    if result.get("retCode", result.get("ret_code", -1)) == 0:
        _set_ad_current_price(sess, slot_idx, new_p)
        now = datetime.now().strftime("%H:%M:%S")
        await bot.send_message(chat_id=chat_id,
            text=(
                f"⚡ <b>Fast update</b> <code>{now}</code>\n"
                f"💲 <code>{new_p_str}</code> — price rose before the next scheduled cycle"
            ),
            parse_mode="HTML")
    # Non-zero result: silently skipped, see docstring above.


# ─────────────────────────────────────────
# 🔄 PRICE UPDATE LOOP
# ─────────────────────────────────────────
async def auto_update_loop(bot, chat_id, slot_idx: int = -1):
    """
    Runs one ad's price-update cycle. slot_idx=-1 is Ad 1 (the original
    single-ad behavior, completely unchanged); 0/1 are Ad 2/Ad 3.

    Auto-stop-on-failure: when a user has MORE THAN ONE ad active, two
    consecutive failed updates on this specific ad auto-stops just this
    slot and notifies the user to fix it on Bybit directly. This is now
    only a fallback for genuinely unexpected errors — the two known,
    recoverable cases are handled directly instead of counting as
    failures: 912120022 (out-of-range — retried with Bybit's own stated
    max) and 90043 (new price rounds to the same value the ad already has
    live — retried with a small nudge). Single-ad users keep the original
    behavior on any other error: it keeps retrying forever, since there's
    no other ad's price at risk for them.
    """
    sess = _s(chat_id)
    _set_ad_running(sess, slot_idx, True)
    label     = _ad_slot_label(slot_idx)
    s         = _ad_settings(sess, slot_idx)
    ad_data   = _ad_data_of(sess, slot_idx)
    interval  = s.get("interval", 2)
    increment = Decimal(str(s.get("increment","0.05")))
    if s.get("mode") == "fixed":
        _set_ad_current_price(sess, slot_idx, Decimal(str(ad_data.get("price","0"))))

    # ── Load this user's credentials ONCE at loop start ──
    # Re-read from DB so any key updates take effect on next loop restart.
    creds = get_user_creds(chat_id)
    if not creds or not creds.get("key"):
        await bot.send_message(chat_id=chat_id,
            text=(
                f"❌ <b>{label} Auto-Update stopped</b>\n\n"
                "No Bybit API key found for your account.\n"
                "Go to 🔑 <b>Set APIs</b> → <b>Set Bybit API</b> first."
            ),
            parse_mode="HTML")
        _set_ad_running(sess, slot_idx, False)
        return

    cycle = 0
    while _ad_running(sess, slot_idx):
        cycle += 1
        now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        mode = s.get("mode","fixed")
        prefix = f"[{label}] " if sess.total_ad_slots() > 1 else ""

        if mode == "fixed":
            new_p    = _ad_current_price(sess, slot_idx) + increment
            _quant   = Decimal("0.00000001")   # unchanged — fixed mode's original precision
            chase_ceiling = False   # live-ceiling chase only applies to floating mode
        else:
            try:
                float_pct = float(s.get("float_pct") or 0)
            except (TypeError, ValueError):
                float_pct = 0
            if float_pct <= 0:
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"⚠️ {prefix}<b>Cycle {cycle}</b> — Float % isn't set yet.\n"
                        f"Set it from the AD Price Bot menu, then restart {label}."
                    ), parse_mode="HTML")
                _set_ad_running(sess, slot_idx, False)
                _set_ad_task(sess, slot_idx, None)
                return
            local_usdt_ref = float(sess.shared_local_usdt_ref or s.get("local_usdt_ref") or 0)
            new_p_str, err = calc_floating_price(ad_data, float_pct, local_usdt_ref)
            if err:
                await bot.send_message(chat_id=chat_id,
                    text=f"⚠️ {prefix}<b>Cycle {cycle} float error</b>\n<code>{_esc(str(err))}</code>", parse_mode="HTML")
                for _ in range(interval * 60):
                    if not _ad_running(sess, slot_idx): break
                    await asyncio.sleep(1)
                continue
            new_p  = Decimal(new_p_str)   # calc_floating_price's own 0.01 precision
            _quant = Decimal("0.01")
            chase_ceiling = _wants_live_ceiling(
                ad_data.get("currencyId",""), ad_data.get("tokenId",""), float_pct
            )

        # ── Same-float multi-ad support: same % is allowed across ads on
        # the same pair now — this nudges the actual PRICE apart instead,
        # only when another active ad on the identical pair would
        # otherwise land within the minimum gap. ──
        if sess.total_ad_slots() > 1:
            new_p = _resolve_price_collision(
                sess, slot_idx,
                ad_data.get("currencyId",""), ad_data.get("tokenId",""),
                new_p
            )
        new_p_str = str(new_p.quantize(_quant, rounding=ROUND_HALF_UP))

        # ── Live-ceiling chase (max floating %) ──
        # At the top float % for this pair, submit a price deliberately
        # far beyond anything Bybit would ever accept, guaranteeing an
        # out-of-range (912120022) rejection. Bybit's error message always
        # includes its OWN current real max — the 912120022 handler below
        # then posts exactly that. This makes the ad always track Bybit's
        # true live ceiling every cycle instead of a fixed formula number
        # that can drift from it as the order book moves.
        if chase_ceiling:
            probe_price = (new_p * Decimal("5")).quantize(_quant, rounding=ROUND_HALF_UP)
            submit_price, submit_str = probe_price, str(probe_price)
        else:
            submit_price, submit_str = new_p, new_p_str

        if slot_idx == -1:
            _record_modify_ad1(sess)   # shared budget with the fast-chase check
        result   = await asyncio.get_event_loop().run_in_executor(
            _ad_executor, modify_ad, s["ad_id"], submit_str, ad_data, creds
        )
        ret_code = result.get("retCode", result.get("ret_code",-1))
        ret_msg  = result.get("retMsg",  result.get("ret_msg","Unknown"))

        if ret_code == 912120022:
            # Out-of-range — Bybit tells us its own max/min. The FIRST
            # retry posts that exact number — it's what Bybit itself just
            # told us is valid, and with live BTC/ETH prices constantly
            # moving there's essentially no risk of it colliding with
            # anything. Only if that exact-boundary attempt ALSO fails
            # do later attempts back off with a small safety margin, each
            # time reading a FRESH boundary from the latest response.
            #
            # One specific failure mode handled here: if this ad already
            # sits EXACTLY at Bybit's stated boundary (e.g. a previous
            # cycle already parked it at the live max), resubmitting that
            # same exact number gets rejected with 90043 ("price differs
            # by less than 0%") instead of 912120022 — Bybit is telling us
            # the number is valid, just unchanged. That used to bubble up
            # as a totally different, unhandled error and get counted as
            # a cycle failure after only 2 occurrences, auto-stopping the
            # ad. Now it's treated the same way the top-level 90043 branch
            # handles it: nudge off it by this pair's minimum price gap,
            # in whichever direction we were already heading, and keep
            # trying — rather than stopping the ad over a price that's
            # perfectly fine, just identical to what's already live.
            last_code, last_msg = ret_code, ret_msg
            posted_price   = None
            was_too_high   = None
            candidate      = None
            for _attempt in range(4):
                if last_code == 912120022:
                    min_str, max_str = _extract_bybit_bounds(last_msg)
                    was_too_high = max_str is not None
                    bound_str = max_str if was_too_high else min_str
                    if not bound_str:
                        break   # couldn't parse a boundary at all — nothing more we can do
                    bound_dec = Decimal(bound_str)
                    if _attempt < 2:
                        # Use Bybit's own string EXACTLY — no re-quantizing. Bybit
                        # doesn't always use the same decimal precision we do
                        # (e.g. 3 decimals on some USD pairs vs 2 on NGN), so
                        # rounding this to our own precision can round IT UP
                        # past Bybit's actual limit and fail again for a reason
                        # that has nothing to do with the price being wrong.
                        candidate_str = bound_str
                        candidate     = bound_dec
                    else:
                        # Last resort: nudge off the boundary. Round in the SAFE
                        # direction (down if we were too high, up if too low) so
                        # quantizing can never push us back out of range.
                        margin    = _safety_margin(bound_dec)
                        candidate = (bound_dec - margin) if was_too_high else (bound_dec + margin)
                        safe_rounding = ROUND_FLOOR if was_too_high else ROUND_CEILING
                        candidate     = candidate.quantize(_quant, rounding=safe_rounding)
                        candidate_str = str(candidate)
                elif last_code == 90043 and candidate is not None:
                    gap = get_min_price_gap(ad_data.get("currencyId",""), ad_data.get("tokenId",""), candidate)
                    candidate = (candidate - gap) if was_too_high else (candidate + gap)
                    candidate = candidate.quantize(_quant, rounding=ROUND_HALF_UP)
                    candidate_str = str(candidate)
                else:
                    break   # a genuinely different error — stop retrying, fall through to failure handling

                if slot_idx == -1:
                    _record_modify_ad1(sess)
                retry_result = await asyncio.get_event_loop().run_in_executor(
                    _ad_executor, modify_ad, s["ad_id"], candidate_str, ad_data, creds
                )
                last_code = retry_result.get("retCode", retry_result.get("ret_code",-1))
                last_msg  = retry_result.get("retMsg",  retry_result.get("ret_msg","Unknown"))
                if last_code == 0:
                    posted_price = candidate
                    break
                if last_code not in (912120022, 90043):
                    break   # a different error now — stop retrying, fall through to failure handling

            if posted_price is not None:
                _reset_ad_failures(sess, slot_idx)
                _set_ad_current_price(sess, slot_idx, posted_price)
                if chase_ceiling:
                    await bot.send_message(chat_id=chat_id,
                        text=(
                            f"✅ {prefix}<b>Cycle {cycle}</b> <code>{now}</code>\n"
                            f"🏔 Chasing Bybit's live ceiling (max float {float_pct}%)\n"
                            f"💲 Posted at Bybit's real-time max: <code>{posted_price}</code> ({mode.upper()})"
                        ),
                        parse_mode="HTML")
                else:
                    await bot.send_message(chat_id=chat_id,
                        text=(
                            f"✅ {prefix}<b>Cycle {cycle}</b> <code>{now}</code>\n"
                            f"⚠️ Original <code>{new_p_str}</code> was out of range\n"
                            f"💲 Posted within Bybit's limit: <code>{posted_price}</code> ({mode.upper()})"
                        ),
                        parse_mode="HTML")
            else:
                if await _handle_ad_cycle_failure(bot, chat_id, sess, slot_idx, label, cycle, last_code, last_msg, ad_data):
                    return


        elif ret_code == 90043:
            # "The price of this P2P ad differs from your existing ad by
            # less than 0%." — this fires when the computed price rounds
            # to the SAME value the ad already has live on Bybit (happens
            # when the underlying market barely moves between cycles, or
            # when this ad hasn't changed since its last successful post).
            # This is fully recoverable — nudge the price by this pair's
            # minimum gap and retry, same pattern as the out-of-range case
            # above, instead of counting it as a failure. If one nudge
            # isn't enough (e.g. the pair's flat gap is tiny relative to
            # its price scale), each further attempt doubles the nudge; if
            # Bybit instead reports we've now gone out of range, switch to
            # posting its stated boundary directly.
            last_code, last_msg = ret_code, ret_msg
            candidate = new_p
            posted_price = None
            for _attempt in range(3):
                if last_code == 90043:
                    gap = get_min_price_gap(ad_data.get("currencyId",""), ad_data.get("tokenId",""), candidate) * (2 ** _attempt)
                    candidate = candidate - gap
                    candidate = candidate.quantize(_quant, rounding=ROUND_HALF_UP)
                elif last_code == 912120022:
                    min_str, max_str = _extract_bybit_bounds(last_msg)
                    bound_str = max_str if max_str else min_str
                    if not bound_str:
                        break
                    candidate = Decimal(bound_str)
                else:
                    break   # a genuinely different error — stop retrying, fall through to failure handling

                candidate_str = str(candidate)
                if slot_idx == -1:
                    _record_modify_ad1(sess)
                retry_result = await asyncio.get_event_loop().run_in_executor(
                    _ad_executor, modify_ad, s["ad_id"], candidate_str, ad_data, creds
                )
                last_code = retry_result.get("retCode", retry_result.get("ret_code",-1))
                last_msg  = retry_result.get("retMsg",  retry_result.get("ret_msg","Unknown"))
                if last_code == 0:
                    posted_price = candidate
                    break
                if last_code not in (90043, 912120022):
                    break

            if posted_price is not None:
                _reset_ad_failures(sess, slot_idx)
                _set_ad_current_price(sess, slot_idx, posted_price)
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"✅ {prefix}<b>Cycle {cycle}</b> <code>{now}</code>\n"
                        f"⚠️ Price unchanged from last post — nudged\n"
                        f"💲 <code>{posted_price}</code> ({mode.upper()})"
                    ),
                    parse_mode="HTML")
            else:
                if await _handle_ad_cycle_failure(bot, chat_id, sess, slot_idx, label, cycle, last_code, last_msg, ad_data):
                    return

        elif ret_code == 0:
            _reset_ad_failures(sess, slot_idx)
            _set_ad_current_price(sess, slot_idx, submit_price)
            if chase_ceiling:
                # Extremely unlikely — the probe price (5x the formula
                # number) was accepted outright instead of triggering an
                # out-of-range rejection. Flag it clearly rather than
                # silently leaving the ad listed at that price, since this
                # means Bybit's real ceiling for this ad is currently even
                # higher than expected.
                await bot.send_message(chat_id=chat_id,
                    text=(
                        f"⚠️ {prefix}<b>Cycle {cycle}</b> <code>{now}</code>\n"
                        f"🏔 Live-ceiling probe was accepted directly at <code>{submit_str}</code> — "
                        f"Bybit's real max right now is at or above this. Double-check this ad on Bybit."
                    ),
                    parse_mode="HTML")
            else:
                await bot.send_message(chat_id=chat_id,
                    text=f"✅ {prefix}<b>Cycle {cycle}</b> <code>{now}</code>\n💲 <code>{submit_str}</code> ({mode.upper()})",
                    parse_mode="HTML")
        else:
            if await _handle_ad_cycle_failure(bot, chat_id, sess, slot_idx, label, cycle, ret_code, ret_msg, ad_data):
                return

        _fast_chase = (slot_idx == -1 and mode == "floating"
                       and sess.total_ad_slots() == 1)
        for _tick in range(interval * 60):
            if not _ad_running(sess, slot_idx): break
            if _fast_chase and _tick > 0 and _tick % 10 == 0:
                await _try_fast_chase(bot, chat_id, sess, slot_idx, ad_data, s, float_pct, creds, _quant)
            await asyncio.sleep(1)

    logger.info(f"🛑 PRICE LOOP STOPPED ({label}) for user {chat_id}")


async def _handle_ad_cycle_failure(bot, chat_id, sess, slot_idx, label, cycle, ret_code, ret_msg, ad_data=None) -> bool:
    """
    Shared failure handler for auto_update_loop. Returns True if the loop
    should stop immediately (this slot was auto-stopped), False if it
    should keep going to its normal inter-cycle sleep.
    """
    prefix = f"[{label}] " if sess.total_ad_slots() > 1 else ""
    extra = ""
    if ad_data:
        _ecur = ad_data.get("currencyId","").upper()
        extra = f"\n💱 Update {_ecur}/USDT ref if rate changed" if (currency_needs_ref(_ecur) or _ecur == "NGN") else ""

    if sess.total_ad_slots() > 1:
        fail_count = _increment_ad_failures(sess, slot_idx)
        if fail_count >= 2:
            _set_ad_running(sess, slot_idx, False)
            _set_ad_task(sess, slot_idx, None)
            await bot.send_message(chat_id=chat_id,
                text=(
                    f"🛑 <b>{label} auto-stopped</b>\n\n"
                    f"2 failed updates in a row — likely too close to another ad's price, "
                    f"or a Bybit-side rejection.\n"
                    f"Last error: <code>{ret_code}</code> — <code>{_esc(str(ret_msg))}</code>{extra}\n\n"
                    f"Check/edit this ad directly on Bybit, then restart it from the bot."
                ),
                parse_mode="HTML")
            return True

    await bot.send_message(chat_id=chat_id,
        text=f"❌ {prefix}<b>Cycle {cycle} failed</b>\n<code>{ret_code}</code> — <code>{_esc(str(ret_msg))}</code>{extra}",
        parse_mode="HTML")
    return False



# ─────────────────────────────────────────
# 📤 Send / edit menu with banner image
# ─────────────────────────────────────────
async def send_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send the main menu with the banner image attached."""
    uid     = update.effective_user.id
    chat_id = update.effective_chat.id
    text    = main_menu_text(uid)
    kb      = main_menu_keyboard(uid)
    try:
        await context.bot.send_photo(
            chat_id=chat_id,
            photo=BANNER_URL,
            caption=text,
            reply_markup=kb,
            parse_mode="HTML"
        )
    except Exception as e:
        logger.warning(f"[Menu] Failed to send photo, falling back to text: {e}")
        await context.bot.send_message(
            chat_id=chat_id, text=text, reply_markup=kb, parse_mode="HTML"
        )


async def edit_menu(query, text: str, keyboard: InlineKeyboardMarkup):
    """Edit the existing menu message (photo caption or plain text).
    Tries caption first (photo messages), falls back to text, then sends new message."""
    # Try caption edit (for photo/banner messages)
    try:
        await query.edit_message_caption(caption=text, reply_markup=keyboard, parse_mode="HTML")
        return
    except Exception:
        pass
    # Try text edit (for plain text messages)
    try:
        await query.edit_message_text(text=text, reply_markup=keyboard, parse_mode="HTML")
        return
    except Exception as e:
        logger.warning(f"[edit_menu] edit failed: {e}")
    # Last resort — send as new message
    try:
        await query.message.reply_text(text=text, reply_markup=keyboard, parse_mode="HTML")
    except Exception as e:
        logger.error(f"[edit_menu] send fallback also failed: {e}")


def _esc(value: str) -> str:
    """HTML-escape a string so it is safe inside parse_mode='HTML' messages.
    Escapes &, <, > which are the only three Telegram HTML mode cares about.
    API keys often contain underscores, dashes, dots — none of those need escaping.
    """
    return (value or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


async def _typing(context: ContextTypes.DEFAULT_TYPE, chat_id: int, min_delay: float = 0.5, max_delay: float = 1.4):
    """
    Show the 'typing…' indicator and hold for a short human-feeling delay
    before the caller sends its reply. Called at the top of every
    user-facing handler (buttons, free text, commands) so nothing lands
    instantly. Failures here are swallowed — this is cosmetic only and
    must never block or crash a real response.
    """
    try:
        await context.bot.send_chat_action(chat_id=chat_id, action="typing")
    except Exception:
        pass
    try:
        await asyncio.sleep(random.uniform(min_delay, max_delay))
    except Exception:
        pass


_TG_USERNAME_RE = re.compile(r'^@[A-Za-z][A-Za-z0-9_]{4,31}$')
_PHONE_RE       = re.compile(r'^\+[1-9]\d{6,14}$')


def validate_contact(text: str):
    """
    Validates a contact string as either a Telegram username (@name) or a
    WhatsApp/phone number in international format (+countrycode...).
    Returns (ok, normalized_contact, error_message) — error_message is ""
    when ok is True.
    """
    t = (text or "").strip()
    if not t:
        return False, "", "❌ Please send your Telegram username or WhatsApp number."
    if t.startswith("@"):
        if _TG_USERNAME_RE.match(t):
            return True, t, ""
        return False, "", (
            "❌ That doesn't look like a valid Telegram username.\n"
            "It should start with @ and be 5-32 characters (letters, numbers, underscores).\n"
            "Example: <code>@johnsmith</code>"
        )
    if t.startswith("+"):
        if _PHONE_RE.match(t):
            return True, t, ""
        return False, "", (
            "❌ That doesn't look like a valid WhatsApp number.\n"
            "Include your country code starting with +, digits only, no spaces or dashes.\n"
            "Example: <code>+2348012345678</code>"
        )
    return False, "", (
        "❌ Please start with <code>@</code> for a Telegram username (e.g. <code>@johnsmith</code>) "
        "or <code>+</code> for a WhatsApp number with country code (e.g. <code>+2348012345678</code>)."
    )


async def _submit_upgrade_request(bot, uid: int, uname: str, dname: str, contact: str):
    """
    Shared upgrade-request submission — saves to DB and notifies every
    admin, now including the contact detail the user provided. Used from
    the free-text message handler (after contact info passes validation);
    confirming to the user is handled separately by the caller since that
    differs by call site.
    """
    logger.info(f"[Upgrade] Request from uid={uid} uname=@{uname} — saving to DB")
    try:
        db.request_upgrade(uid, uname, dname, contact)
        logger.info(f"[Upgrade] DB write OK for uid={uid}")
    except Exception as _db_err:
        logger.error(f"[Upgrade] DB write FAILED for uid={uid}: {_db_err}")

    _ref_line = ""
    _referrer_id = db.get_referrer(uid)
    if _referrer_id:
        _referrer = db.get_user(_referrer_id)
        _rname = _esc(_referrer.get("username") or _referrer.get("display_name") or str(_referrer_id)) if _referrer else str(_referrer_id)
        _ref_line = f"🎁 Referred by: @{_rname} (<code>{_referrer_id}</code>)\n\n"
    _admin_msg = (
        f"🔔 <b>New Upgrade Request!</b>\n\n"
        f"👤 User ID: <code>{uid}</code>\n"
        f"Username: @{uname if uname else 'None'}\n"
        f"Name: {_esc(dname)}\n"
        f"📞 Contact: {_esc(contact)}\n\n"
        f"{_ref_line}"
        f"Approve: <code>/upgrade {uid} 30</code>"
    )
    for _admin_id in list(_admin_chat_ids):
        try:
            await bot.send_message(chat_id=_admin_id, text=_admin_msg, parse_mode="HTML")
            logger.info(f"[Upgrade] Admin {_admin_id} notified for uid={uid}")
        except Exception as _notify_err:
            logger.error(f"[Upgrade] Could not notify admin {_admin_id}: {_notify_err}")


async def edit_menu_html(query, text: str, keyboard: InlineKeyboardMarkup):
    """Like edit_menu but uses HTML parse mode — safe for raw API keys / UUIDs."""
    try:
        await query.edit_message_caption(caption=text, reply_markup=keyboard, parse_mode="HTML")
        return
    except Exception:
        pass
    try:
        await query.edit_message_text(text=text, reply_markup=keyboard, parse_mode="HTML")
        return
    except Exception as e:
        logger.warning(f"[edit_menu_html] edit failed: {e}")
    try:
        await query.message.reply_text(text=text, reply_markup=keyboard, parse_mode="HTML")
    except Exception as e:
        logger.error(f"[edit_menu_html] send fallback also failed: {e}")


# ─────────────────────────────────────────
# /start   /menu
# ─────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tuser = update.effective_user
    user, is_new = _get_or_register_user(tuser)
    if is_admin(tuser.id):
        _admin_chat_ids.add(update.message.chat_id)

    # ── Referral capture ──
    # Deep link looks like https://t.me/<bot>?start=ref_R<code> — only ever
    # takes effect for genuinely new accounts, and never for self-referral
    # (db.record_referral_join() enforces both, plus "referrer must exist").
    if is_new and context.args:
        payload = context.args[0]
        if payload.startswith("ref_"):
            referrer_id = db.resolve_referral_code(payload[len("ref_"):])
            if referrer_id:
                linked = db.record_referral_join(
                    tuser.id, referrer_id, tuser.username or "", tuser.full_name or ""
                )
                if linked:
                    logger.info(f"[Referral] {tuser.id} linked to referrer {referrer_id}")
                    try:
                        await context.bot.send_message(
                            chat_id=referrer_id,
                            text=(
                                f"🎁 <b>New referral!</b>\n\n"
                                f"@{tuser.username or tuser.full_name or tuser.id} joined using your link.\n"
                                f"You'll earn ₦{REFERRAL_REWARD_NGN:,} once they upgrade to Pro and the admin approves it."
                            ),
                            parse_mode="HTML"
                        )
                    except Exception:
                        pass

    # Auto-downgrade expired pro users
    db.check_and_auto_downgrade(tuser.id)

    # ── Always refresh plan badge from DB so Pro shows instantly after upgrade ──
    global _current_user_id, _current_plan_badge
    _current_user_id    = tuser.id
    _current_plan_badge = sub.plan_badge(tuser.id)

    # ── Load persisted settings from disk into session ──
    # This ensures settings (Ad ID, UID, mode, interval, etc.) survive restarts.
    _load_settings_from_disk(tuser.id)

    # Load scammer list if empty
    if get_scammer_count() == 0:
        asyncio.get_event_loop().run_in_executor(None, load_scammers)

    # Load help-agent knowledge base if empty
    if help_agent.get_entry_count() == 0:
        asyncio.get_event_loop().run_in_executor(None, help_agent.load_knowledge)

    await send_menu(update, context)


async def menu_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start(update, context)


# ─────────────────────────────────────────
# 🏓 Ping commands
# ─────────────────────────────────────────
async def ping_bybit_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test Bybit API — works for all Pro users and admin, using their own saved keys."""
    uid   = update.effective_user.id
    creds = get_user_creds(uid)
    if not is_admin(uid) and not creds.get("key"):
        await update.message.reply_text(
            "❌ *No Bybit API set.*\n\nGo to 🔑 *Set APIs* → Set Bybit Account 1 API first.",
            parse_mode="HTML"
        )
        return
    uid  = update.effective_user.id
    slot = _get_user_slot_str(uid)   # per-user slot
    await update.message.reply_text(f"⏳ Testing Bybit Account {slot} API...")
    from bybit import ping_api
    result   = await asyncio.get_event_loop().run_in_executor(None, partial(ping_api, creds=creds))
    ret_code = result.get("retCode", -1)
    if ret_code == 0:
        info      = result.get("result", {})
        perms     = info.get("permissions", {})
        ips       = info.get("ips", [])
        fiat_p2p  = perms.get("FiatP2P", [])
        has_ads   = "Advertising" in fiat_p2p
        read_only = info.get("readOnly", 1)
        plines    = [f"  {'✅' if v else '➖'} {k}: {', '.join(v) if v else 'none'}" for k,v in perms.items()]
        ad_stat   = "✅ Can edit ads" if has_ads and not read_only else \
                    "⚠️ Read only"   if has_ads else "❌ No P2P permission"
        await update.message.reply_text(
            f"✅ <b>Bybit Account {slot} API connected!</b>\n\n"
            f"🔑 <code>...{info.get('apiKey','')[-6:]}</code>\n"
            f"🔒 Read only: <code>{'Yes' if read_only else 'No'}</code>\n"
            f"🌍 IPs: <code>{', '.join(ips) if ips else 'None'}</code>\n\n"
            f"🔓 <b>Permissions:</b>\n" + "\n".join(plines) + f"\n\n🛒 <b>P2P: {ad_stat}</b>",
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(
            f"❌ <b>API failed</b>\n<code>{_esc(result.get('retMsg',''))}</code>", parse_mode="HTML"
        )


async def ping_flutterwave_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test Flutterwave API — uses the user's own saved FLW secret key from DB."""
    uid        = update.effective_user.id
    secret_key = db.get_api(uid, "flw_secret_key")

    if not secret_key:
        await update.message.reply_text(
            "❌ <b>No Flutterwave API set.</b>\n\n"
            "Go to 🔑 <b>Set APIs</b> → Set Flutterwave API first.\n\n"
            "You need to provide 3 credentials:\n"
            "  FLW_PUBLIC_KEY\n"
            "  FLW_SECRET_HASH\n"
            "  FLW_SECRET_KEY",
            parse_mode="HTML"
        )
        return

    await update.message.reply_text("⏳ Testing Flutterwave v3 API...")
    from flutterwave import ping_flutterwave
    result = await asyncio.get_event_loop().run_in_executor(None, ping_flutterwave, secret_key)
    if "error" in result:
        ip = await _get_current_ip()
        err_text = _esc(result["error"][:300])
        ip_safe  = _esc(ip)
        await update.message.reply_text(
            f"❌ <b>Flutterwave connection failed</b>\n\n"
            f"<code>{err_text}</code>\n\n"
            f"• Ensure FLW_SECRET_KEY starts with <code>FLWSECK_</code>\n"
            f"• Whitelist IP <code>{ip_safe}</code> on Flutterwave → Settings → API → IP Whitelist",
            parse_mode="HTML"
        )
    else:
        banks = result.get("banks", [])
        if banks:
            lines = [f"✅ <b>Flutterwave Connected!</b> {len(banks)} Nigerian banks:\n"]
            for bank in banks[:60]:
                code = _esc(bank.get("code", ""))
                name = _esc(bank.get("name", ""))
                lines.append(f"<code>{code}</code> — {name}")
            msg = "\n".join(lines)
            if len(msg) > 4000:
                msg = msg[:4000] + "\n...(truncated)"
            await update.message.reply_text(msg, parse_mode="HTML")
        else:
            await update.message.reply_text(
                "✅ <b>Flutterwave v3 Connected!</b>\nSecret key valid ✅\nDynamic bank matching active ✅",
                parse_mode="HTML"
            )


async def ping_paga_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test Paga API — uses the user's own saved Paga keys from DB."""
    uid        = update.effective_user.id
    principal  = db.get_api(uid, "paga_principal")
    credential = db.get_api(uid, "paga_credential")
    api_key    = db.get_api(uid, "paga_api_key")

    if not (principal and credential and api_key):
        await update.message.reply_text(
            "❌ <b>No Paga API set.</b>\n\n"
            "Go to 🔑 <b>Set APIs</b> → Set Paga API first.\n\n"
            "You need to provide:\n"
            "  PAGA_API_KEY\n"
            "  PAGA_CREDENTIAL\n"
            "  PAGA_PRINCIPAL",
            parse_mode="HTML"
        )
        return

    await update.message.reply_text("⏳ Testing Paga Business API...")
    from paga import ping_paga
    result = await asyncio.get_event_loop().run_in_executor(
        None, ping_paga, principal, credential, api_key
    )
    if "error" in result:
        ip      = await _get_current_ip()
        err_s   = _esc(result["error"][:300])
        ip_safe = _esc(ip)
        await update.message.reply_text(
            f"❌ <b>Paga connection failed</b>\n\n"
            f"<code>{err_s}</code>\n\n"
            f"• <b>PAGA_PRINCIPAL</b> = Public Key / Principal\n"
            f"• <b>PAGA_CREDENTIAL</b> = Live Primary Secret Key\n"
            f"• <b>PAGA_API_KEY</b> = HMAC Hash Key\n"
            f"• Whitelist IP <code>{ip_safe}</code> on Paga dashboard → Settings → IP Whitelist",
            parse_mode="HTML"
        )
    else:
        banks = result.get("banks", [])
        if banks:
            lines = [f"✅ <b>Paga Connected!</b> {len(banks)} banks available:\n"]
            for bank in banks[:50]:
                uuid_safe = _esc(bank.get("uuid", "?")[:8])
                name_safe = _esc(bank.get("name", ""))
                lines.append(f"<code>{uuid_safe}...</code> — {name_safe}")
            msg = "\n".join(lines)
            if len(msg) > 4000:
                msg = msg[:4000] + "\n...(truncated)"
            await update.message.reply_text(msg, parse_mode="HTML")
        else:
            await update.message.reply_text(
                "✅ <b>Paga Connected!</b>\nCredentials valid ✅\nDynamic bank UUID matching active ✅",
                parse_mode="HTML"
            )


# ─────────────────────────────────────────
# 🎛️ BUTTON HANDLER
# ─────────────────────────────────────────
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Main callback button handler. Wrapped with full exception logging so
    any crash is visible in Render logs with exact line + traceback.
    """
    try:
        await _button_handler_inner(update, context)
    except Exception as _bh_err:
        import traceback
        logger.error(
            f"[ButtonHandler] UNHANDLED EXCEPTION\n"
            f"  data={getattr(getattr(update, 'callback_query', None), 'data', '?')!r}\n"
            f"  user={getattr(getattr(update, 'callback_query', None), 'from_user', None)}\n"
            f"  error={_bh_err}\n"
            f"{traceback.format_exc()}"
        )
        try:
            q = update.callback_query
            if q:
                await q.answer("⚠️ An error occurred. Please try again.", show_alert=True)
        except Exception:
            pass


async def _button_handler_inner(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # All per-user state accessed via _s(tuser.id).field — no globals needed

    query   = update.callback_query
    data    = query.data if update.callback_query else ""
    chat_id = query.message.chat_id if query and query.message else 0
    logger.debug(f"[ButtonHandler] Received callback: data={data!r} chat_id={chat_id}")
    try:
        await query.answer()
    except Exception as _ans_err:
        logger.warning(f"[ButtonHandler] query.answer() failed: {_ans_err}")

    # ── Register/update user on every interaction ──
    global _current_user_id, _current_plan_badge
    tuser = query.from_user
    user_rec, _ = _get_or_register_user(tuser)
    db.check_and_auto_downgrade(tuser.id)
    _current_user_id    = tuser.id
    _current_plan_badge = sub.plan_badge(tuser.id)

    # ── Per-user isolated state for non-admin users ──
    # Admin uses the global user_state; non-admins get their own isolated dict
    if is_admin(tuser.id):
        _btn_state = user_state
    else:
        if "state" not in context.user_data:
            context.user_data["state"] = {}
        _btn_state = context.user_data["state"]
    # NOTE: Bybit credentials are loaded per-call via get_user_creds(tuser.id)
    # — no globals are mutated here. See get_user_creds() for details.


    # ── Pro feature guard ──
    # Block non-admin free users from ALL functional sections.
    # They can only access: main_menu, upgrade_plan, upgrade_request_yes, bot_status,
    # get_my_ip, section_apis, set_api_*, delete_apis, delete_apis_confirm, reset_*
    _FREE_ALLOWED = {
        "main_menu", "upgrade_plan", "upgrade_request_yes",
        "referrals", "contact_support",
        "bot_status", "reset_confirm", "reset_do",
        "section_apis", "set_api_bybit", "set_api_flw", "set_api_paga",
        "set_api_bybit_1", "set_api_bybit_2",
        "delete_apis", "delete_apis_confirm",
        "delete_bybit1_apis", "delete_bybit1_confirm",
        "delete_bybit2_apis", "delete_bybit2_confirm",
        "delete_flw_apis",   "delete_flw_confirm",
        "delete_paga_apis",  "delete_paga_confirm",
    }
    _is_free_allowed = (
        data in _FREE_ALLOWED
        or data.startswith("switch_account_")
        or data.startswith("sc_accept_")
        or data.startswith("sc_reject_")
        or data.startswith("sc_reason_")
    )
    if not is_admin(tuser.id) and not sub.is_pro(tuser.id) and not _is_free_allowed:
        await query.answer(
            "🔒 Upgrade to Pro to access this feature.",
            show_alert=True
        )
        await edit_menu(query,
            "🔒 *Pro Plan Required*\n\nYou need a Pro plan to use this bot.\n\nTap *⬆️ Upgrade Plan* to request access from the admin.",
            main_menu_keyboard(tuser.id)
        )
        return

    # Legacy per-feature guard (still applies for admin-visible toggles)
    if sub.requires_pro(data) and not sub.is_pro(tuser.id) and not is_admin(tuser.id):
        await query.answer(
            "🔒 Pro plan required. Tap Upgrade Plan in the menu.",
            show_alert=True
        )
        return

    # ── 🏠 Main menu ──
    if data == "main_menu":
        # Always refresh plan badge when returning to main menu so upgrades
        # are reflected immediately without needing a redeploy.
        db.check_and_auto_downgrade(tuser.id)
        _current_plan_badge = sub.plan_badge(tuser.id)
        await edit_menu(query, main_menu_text(tuser.id), main_menu_keyboard(tuser.id))

    # ── 🌍 Get My IP ──
    elif data == "get_my_ip":
        await query.edit_message_caption(caption="⏳ Fetching public IP...", parse_mode="HTML") \
            if query.message.photo else await query.edit_message_text("⏳ Fetching public IP...")
        import requests as _req
        ip = None
        for svc in ["https://api.ipify.org", "https://ifconfig.me/ip", "https://icanhazip.com"]:
            try:
                ip = _req.get(svc, timeout=5).text.strip()
                if ip: break
            except Exception:
                continue
        txt = (
            f"🌍 <b>Public IP Address</b>\n\n<code>{ip}</code>\n\n"
            "👉 Add this to your Bybit API whitelist if it changed."
        ) if ip else "❌ Could not fetch IP. Try again."
        try:
            await query.edit_message_caption(caption=txt, reply_markup=InlineKeyboardMarkup(back_main()), parse_mode="HTML")
        except Exception:
            await query.edit_message_text(txt, reply_markup=InlineKeyboardMarkup(back_main()), parse_mode="HTML")

    # ── 🔑 Switch Account ──
    elif data.startswith("switch_account_"):
        idx      = int(data.split("_")[-1])
        accounts = get_all_accounts()
        # Allow up to 2 slots regardless of env keys — credentials come from DB
        if idx >= max(len(accounts), 2):
            await query.answer("Invalid account", show_alert=True)
            return
        if _s(tuser.id).refresh_running or _s(tuser.id).order_monitor_running:
            await query.answer("⚠️ Stop all running tasks before switching accounts.", show_alert=True)
            return

        # ── PER-USER slot switch — does NOT affect any other user ──
        # NEVER call set_active_account() here — that modifies the global
        # bybit._active_index which is shared across ALL users.

        # ── Save current slot's AD BOT settings before switching ──
        old_slot_str = _get_user_slot_str(tuser.id)
        _s(tuser.id).settings[f"mode_{old_slot_str}"]          = _s(tuser.id).settings.get("mode", "fixed")
        _s(tuser.id).settings[f"increment_{old_slot_str}"]     = _s(tuser.id).settings.get("increment", "0.05")
        _s(tuser.id).settings[f"float_pct_{old_slot_str}"]     = _s(tuser.id).settings.get("float_pct", "")
        _s(tuser.id).settings[f"local_usdt_ref_{old_slot_str}"]= _s(tuser.id).settings.get("local_usdt_ref", "")
        _s(tuser.id).settings[f"interval_{old_slot_str}"]      = _s(tuser.id).settings.get("interval", 2)
        _save_settings(tuser.id)   # persist before slot change

        _s(tuser.id).selected_slot = idx   # only this user changes
        new_slot_str = _get_user_slot_str(tuser.id)

        # Clear volatile order/ad data (other users are untouched)
        _s(tuser.id).ad_data.clear()
        _s(tuser.id).seen_order_ids.clear(); _s(tuser.id).paid_order_ids.clear()
        _s(tuser.id).seen_sell_ids.clear();  _s(tuser.id).released_ids.clear()

        # ── Restore new slot's saved AD BOT settings (do NOT overwrite with defaults) ──
        _s(tuser.id).settings["ad_id"]          = _s(tuser.id).settings.get(f"ad_id_{new_slot_str}", "")
        _s(tuser.id).settings["bybit_uid"]      = _s(tuser.id).settings.get(f"bybit_uid_{new_slot_str}", "")
        _s(tuser.id).settings["mode"]           = _s(tuser.id).settings.get(f"mode_{new_slot_str}", "fixed")
        _s(tuser.id).settings["increment"]      = _s(tuser.id).settings.get(f"increment_{new_slot_str}", "0.05")
        _s(tuser.id).settings["float_pct"]      = _s(tuser.id).settings.get(f"float_pct_{new_slot_str}", "")
        _s(tuser.id).settings["local_usdt_ref"] = _s(tuser.id).settings.get(f"local_usdt_ref_{new_slot_str}", "")
        _s(tuser.id).settings["interval"]       = _s(tuser.id).settings.get(f"interval_{new_slot_str}", 2)

        acct_label = accounts[idx]["label"] if idx < len(accounts) else f"Account {idx + 1}"
        logger.info(f"[Slot] User {tuser.id} switched to slot {idx+1} ({acct_label}) — other users unaffected")
        await edit_menu(query,
            f"✅ <b>Switched to {acct_label}</b>\n\nYour session cleared.\n\n" + main_menu_text(tuser.id),
            main_menu_keyboard(tuser.id)
        )

    # ── Section navigations ──
    elif data == "section_ads":
        _sess_ads = _s(tuser.id)
        if _sess_ads.editing_slot >= len(_sess_ads.extra_ad_slots):
            _sess_ads.editing_slot = -1   # guard against a stale index if a slot was removed elsewhere
        await edit_menu(query, ads_section_text(tuser.id), ads_section_keyboard(tuser.id))

    elif data == "section_orders":
        await edit_menu(query, orders_section_text(tuser.id), orders_section_keyboard(tuser.id))

    elif data == "section_autopay":
        await edit_menu(query, autopay_section_text(tuser.id), autopay_section_keyboard(tuser.id))

    # ── 📡 Bot Status ──
    elif data == "bot_status":
        done, total, bar = setup_progress(tuser.id)
        r_status = f"🟢 Running | `{str(_s(tuser.id).current_price) if _s(tuser.id).current_price else _s(tuser.id).ad_data.get('price','—')}`" \
                   if _s(tuser.id).refresh_running else "🔴 Stopped"
        o_status = "🔔 Active — every 10s" if _s(tuser.id).order_monitor_running else "🔕 Stopped"
        bp_s = f"🛡 ON ({_s(tuser.id).buyer_protection_mins}min)" if _s(tuser.id).buyer_protection_on else "🛡 OFF"
        nm_s = "🔍 ON" if _s(tuser.id).name_match_enabled else "🔍 OFF"
        txt = (
            f"📡 <b>Bot Status</b>\n\n"
            f"🔑 Active: <b>{(get_all_accounts()[_s(tuser.id).selected_slot] if get_all_accounts() and _s(tuser.id).selected_slot < len(get_all_accounts()) else (get_all_accounts()[0] if get_all_accounts() else {'label': f'Account {_s(tuser.id).selected_slot + 1}'}))['label']}</b>\n"
            f"Setup: {bar} <code>{done}/{total}</code>\n\n"
            f"📊 Price Bot: {r_status}\n"
            f"📦 Order Monitor: {o_status}\n"
            f"💳 Auto-Pay: {'ON' if _s(tuser.id).auto_pay_enabled else 'OFF'}\n"
            f"💸 FLW Pay: {'ON' if _s(tuser.id).flw_pay_enabled else 'OFF'}\n"
            f"{bp_s} | {nm_s}\n\n"
            f"🆔 Ad: <code>{_s(tuser.id).settings.get('ad_id') or 'Not set'}</code>\n"
            f"🔀 Mode: <code>{_s(tuser.id).settings.get('mode','fixed').upper()}</code>\n"
            f"⏱ Interval: <code>{_s(tuser.id).settings.get('interval',2)} min</code>\n\n"
            f"BUY seen: <code>{len(_s(tuser.id).seen_order_ids)}</code> | Paid: <code>{len(_s(tuser.id).paid_order_ids)}</code>\n"
            f"SELL seen: <code>{len(_s(tuser.id).seen_sell_ids)}</code> | Released: <code>{len(_s(tuser.id).released_ids)}</code>"
        )
        await edit_menu(query, txt, InlineKeyboardMarkup(back_main()))

    # ── 🔁 Reset confirm ──
    elif data == "reset_confirm":
        await edit_menu(query,
            "⚠️ *Reset Session?*\n\nThis clears all settings and stops all running tasks.",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Yes, Reset", callback_data="reset_do")],
                [InlineKeyboardButton("❌ Cancel",     callback_data="main_menu")],
            ])
        )

    elif data == "reset_do":
        _s(tuser.id).refresh_running = False; _s(tuser.id).order_monitor_running = False
        _s(tuser.id).auto_pay_enabled = False; _s(tuser.id).flw_pay_enabled = False; _s(tuser.id).paga_pay_enabled = False
        _s(tuser.id).buyer_protection_on = False; _s(tuser.id).name_match_enabled = False
        _s(tuser.id).chat_monitor_enabled = False
        if _s(tuser.id).chat_monitor_task:
            _s(tuser.id).chat_monitor_task.cancel()
            _s(tuser.id).chat_monitor_task = None
        _s(tuser.id).seen_chat_msgs.clear()
        _s(tuser.id).reply_state.clear()
        _s(tuser.id).order_msg_ids.clear()
        _s(tuser.id).my_account_id = ""
        _s(tuser.id).my_nick       = ""
        if _s(tuser.id).refresh_task:      _s(tuser.id).refresh_task.cancel();      _s(tuser.id).refresh_task = None
        if _s(tuser.id).order_monitor_task: _s(tuser.id).order_monitor_task.cancel(); _s(tuser.id).order_monitor_task = None
        _s(tuser.id).current_price = Decimal("0"); _s(tuser.id).ad_data.clear()
        _s(tuser.id).seen_order_ids = set(); _s(tuser.id).paid_order_ids = set()
        _s(tuser.id).seen_sell_ids = set(); _s(tuser.id).released_ids = set()
        _s(tuser.id).sell_msg_enabled = False; _s(tuser.id).sell_msg_count = 1
        # Reset ONLY this user's slot — NOT global
        _s(tuser.id).selected_slot = 0
        for k, v in [("ad_id",""),("bybit_uid",""),("mode","fixed"),
                     ("increment","0.05"),("float_pct",""),("local_usdt_ref",""),("interval",2)]:
            _s(tuser.id).settings[k] = v
        _s(tuser.id).settings.pop("manage_ad_id",   None)
        _s(tuser.id).settings.pop("manage_ad_data", None)
        _s(tuser.id).settings.pop("post_ad_qty",    None)
        await edit_menu(query,
            "✅ *Session reset!* All settings cleared.\n\nTap /menu to start fresh.",
            InlineKeyboardMarkup(back_main())
        )

    # ── ℹ️ Auto-pay info ──
    elif data == "autopay_info":
        await edit_menu(query,
            "ℹ️ *How Auto-Pay Works*\n\n"
            "1. Order Monitor must be running\n"
            "2. New BUY order arrives → bot waits 5 seconds\n"
            "3. Reads full order and payment details\n"
            "4. Marks the order as paid on Bybit automatically\n"
            "5. 🛡 If Buyer Protection is ON and seller release time ≥ threshold,\n"
            "   bot also sends a warning message to the seller\n"
            "6. 🔍 If Name Match is ON and account info is missing,\n"
            "   bot marks paid + tells seller to cancel\n\n"
            "⚠️ Ensure you have funds to cover orders before enabling.",
            InlineKeyboardMarkup(back_section("section_autopay"))
        )

    # ── ℹ️ FLW info ──
    elif data == "flw_info":
        await edit_menu(query,
            "ℹ️ *How Flutterwave Auto-Pay Works*\n\n"
            "1. Order Monitor must be running\n"
            "2. New BUY order → bot waits 5 seconds\n"
            "3. 🔍 Name Match: if account info missing → mark paid + ask seller to cancel\n"
            "4. 🛡 Buyer Protection: if seller release time ≥ threshold → mark paid + warn seller (no FLW transfer)\n"
            "5. Verifies seller's bank account via Flutterwave\n"
            "6. Sends NGN transfer\n"
            "7. Polls status up to 60s — if SUCCESSFUL → marks Bybit order paid\n\n"
            "⚠️ Cannot run with Bybit Auto-Pay simultaneously.\n"
            "⚠️ Keep enough NGN balance on Flutterwave.",
            InlineKeyboardMarkup(back_section("section_autopay"))
        )

    # ── 🛡 Buyer Protection menu ──
    elif data == "buyer_protection_menu":
        await edit_menu(query, buyer_protection_menu_text(tuser.id), buyer_protection_menu_keyboard(tuser.id))

    elif data == "toggle_buyer_protection":
        _s(tuser.id).buyer_protection_on = not _s(tuser.id).buyer_protection_on
        status = "✅ ON" if _s(tuser.id).buyer_protection_on else "❌ OFF"
        await edit_menu(query,
            f"🛡 <b>Buyer Protection {status}</b>\n\nThreshold: <code>{_s(tuser.id).buyer_protection_mins} min</code>\n\n"
            + buyer_protection_menu_text(tuser.id),
            buyer_protection_menu_keyboard(tuser.id)
        )

    elif data.startswith("bp_set_") and data != "bp_set_custom":
        mins = int(data.split("_")[-1])
        _s(tuser.id).buyer_protection_mins = mins
        await edit_menu(query,
            f"✅ <b>Buyer Protection threshold set to <code>{mins} min</code></b>\n\n" + buyer_protection_menu_text(tuser.id),
            buyer_protection_menu_keyboard(tuser.id)
        )

    elif data == "bp_set_custom":
        _btn_state["action"]       = "bp_custom_threshold"
        _btn_state["prev_section"] = "buyer_protection_menu"
        await edit_menu(query,
            f"✏️ <b>Custom Buyer Protection Threshold</b>\n\n"
            f"Current: <code>{_s(tuser.id).buyer_protection_mins} min</code>\n\n"
            "Send the number of minutes you want to use as the threshold.\n"
            "Example: `25`",
            InlineKeyboardMarkup(back_section("section_autopay"))
        )

    # ── 🔍 Name Match toggle ──
    elif data == "toggle_name_match":
        _s(tuser.id).name_match_enabled = not _s(tuser.id).name_match_enabled
        status = "✅ ON" if _s(tuser.id).name_match_enabled else "❌ OFF"
        await edit_menu(query,
            f"🔍 <b>Name Match {status}</b>\n\n"
            + ("When enabled, if the bot detects no account name or account number "
               "on a BUY order, it will:\n\n"
               "  • Mark the order as paid on Bybit\n"
               "  • Tell the seller to request a cancel\n"
               "  • Skip Flutterwave transfer entirely\n\n"
               if _s(tuser.id).name_match_enabled else
               "Name Match is now disabled.\n\n")
            + autopay_section_text(tuser.id),
            autopay_section_keyboard(tuser.id)
        )

    # ── 💳 Toggle Auto-Pay ──
    elif data == "toggle_auto_pay":
        _s(tuser.id).auto_pay_enabled = not _s(tuser.id).auto_pay_enabled
        if _s(tuser.id).auto_pay_enabled and _s(tuser.id).flw_pay_enabled:
            _s(tuser.id).flw_pay_enabled = False
        if _s(tuser.id).auto_pay_enabled and _s(tuser.id).paga_pay_enabled:
            _s(tuser.id).paga_pay_enabled = False
        await edit_menu(query, autopay_section_text(tuser.id), autopay_section_keyboard(tuser.id))

    # ── 🟢 Toggle Flutterwave Pay ──
    elif data == "toggle_flw_pay":
        if not _s(tuser.id).flw_pay_enabled:
            # All users (including admin) must have all 3 FLW keys in DB
            _flw_ready = all(db.get_api(tuser.id, k) for k in (
                "flw_public_key", "flw_secret_hash", "flw_secret_key"
            ))
            if not _flw_ready:
                await query.answer(
                    "❌ Flutterwave API incomplete. Go to 🔑 Set APIs → Set Flutterwave API and enter all 3 credentials.",
                    show_alert=True
                )
                return
        _s(tuser.id).flw_pay_enabled = not _s(tuser.id).flw_pay_enabled
        if _s(tuser.id).flw_pay_enabled and _s(tuser.id).auto_pay_enabled:
            _s(tuser.id).auto_pay_enabled = False
        if _s(tuser.id).flw_pay_enabled and _s(tuser.id).paga_pay_enabled:
            _s(tuser.id).paga_pay_enabled = False
        await edit_menu(query, autopay_section_text(tuser.id), autopay_section_keyboard(tuser.id))

    # ── 🟡 Toggle Paga Pay ──
    elif data == "toggle_paga_pay":
        if not _s(tuser.id).paga_pay_enabled:
            _paga_key = db.get_api(tuser.id, "paga_principal")
            if not _paga_key:
                await query.answer(
                    "❌ No Paga API saved. Go to 🔑 Set APIs → Set Paga API first.",
                    show_alert=True
                )
                return
        _s(tuser.id).paga_pay_enabled = not _s(tuser.id).paga_pay_enabled
        if _s(tuser.id).paga_pay_enabled and _s(tuser.id).auto_pay_enabled:
            _s(tuser.id).auto_pay_enabled = False
        if _s(tuser.id).paga_pay_enabled and _s(tuser.id).flw_pay_enabled:
            _s(tuser.id).flw_pay_enabled = False
        await edit_menu(query, autopay_section_text(tuser.id), autopay_section_keyboard(tuser.id))

    # ── ℹ️ Paga info ──
    elif data == "paga_info":
        await edit_menu(query,
            "ℹ️ *How Paga Auto-Pay Works*\n\n"
            "1. Order Monitor must be running\n"
            "2. New BUY order → bot waits 5 seconds\n"
            "3. 🔍 Name Match: if account info missing → mark paid + ask seller to cancel\n"
            "4. 🛡 Buyer Protection: if seller release time ≥ threshold → mark paid + warn seller (no Paga transfer)\n"
            "5. Fetches bank UUID from Paga's bank list\n"
            "6. Validates seller's bank account via Paga\n"
            "7. Sends NGN transfer via Paga depositToBank\n"
            "8. If successful → marks Bybit order paid\n"
            "9. Paga webhook notifies you in Telegram of transfer status\n\n"
            "⚠️ Only ONE of Bybit, Flutterwave, or Paga can be active at a time.\n"
            "⚠️ Keep enough NGN balance on your Paga business account.\n"
            "⚠️ Whitelist your Render IP on Paga dashboard → Settings → IP Whitelist.",
            InlineKeyboardMarkup(back_section("section_autopay"))
        )

    # ── ✏️ Set Sender Name ──
    elif data == "set_sender_name":
        _btn_state["action"]       = "sender_name"
        _btn_state["prev_section"] = "section_autopay"
        cur = _s(tuser.id).settings.get("sender_name", "Not set")
        await edit_menu(query,
            f"✏️ <b>Set Your Sender Name</b>\n\nCurrent: <code>{cur}</code>\n\n"
            "This name appears in the Flutterwave transfer narration:\n"
            f"<code>[Your Name] payment to [Receiver Name]</code>\n\n"
            "Send your full name — e.g. `Akinrinade Akinniyi`",
            InlineKeyboardMarkup(back_section("section_autopay"))
        )

    # ── 📋 View Unpaid Orders ──
    elif data == "view_unpaid_orders":
        if not _s(tuser.id).unpaid_log:
            await edit_menu(query,
                "📋 *Unpaid Orders*\n\nNo unpaid orders recorded this session. ✅",
                InlineKeyboardMarkup([
                    [InlineKeyboardButton("🗑 Clear Log", callback_data="clear_unpaid_log")],
                    *back_section("section_autopay")
                ])
            )
            return
        lines = [f"📋 *Unpaid Orders ({len(_s(tuser.id).unpaid_log)}):*\n"]
        for i, entry in enumerate(_s(tuser.id).unpaid_log[-20:], 1):
            lines.append(
                f"<b>{i}.</b> <code>{entry['order_id']}</code>\n"
                f"  👤 <code>{entry.get('account_no','—')}</code> ({entry.get('bank','—')})\n"
                f"  💵 <code>{entry.get('amount',0):,.2f} NGN</code>\n"
                f"  ❌ {entry.get('reason','Unknown')}\n"
                f"  🕐 {entry.get('timestamp','')}\n"
            )
        msg = "\n".join(lines)
        if len(msg) > 4000: msg = msg[:4000] + "\n...(truncated)"
        await edit_menu(query, msg,
            InlineKeyboardMarkup([
                [InlineKeyboardButton("🗑 Clear Log", callback_data="clear_unpaid_log")],
                *back_section("section_autopay")
            ])
        )

    elif data == "clear_unpaid_log":
        _s(tuser.id).unpaid_log.clear()
        await edit_menu(query, "✅ Unpaid orders log cleared.", InlineKeyboardMarkup(back_section("section_autopay")))

    # ── 📊 View Buy Volume Analytics (24h) ──
    elif data == "view_buy_volume":
        sess_bv = _s(tuser.id)
        lines_bv = sess_bv.get_buy_volume_lines()
        secs_left = sess_bv.buy_volume_reset_in_seconds()
        hrs_left  = secs_left // 3600
        mins_left = (secs_left % 3600) // 60
        if not lines_bv:
            body = "No buy orders recorded in the current 24h window yet."
        else:
            body = "<code>" + "\n".join(lines_bv) + "</code>"
        await edit_menu(query,
            f"📊 <b>Buy Volume — 24h Window</b>\n\n"
            f"{body}\n\n"
            f"⏱ Window resets in <b>{hrs_left}h {mins_left}m</b>\n\n"
            f"<i>Coin quantities only (not fiat) — total volume you've bought "
            f"across all accepted/auto-paid buy orders in this window. "
            f"Not affected by the hourly session reset.</i>",
            InlineKeyboardMarkup(back_section("section_autopay"))
        )

    # ── 💬 Toggle Chat Monitor ──
    elif data == "toggle_chat_monitor":
        if _s(tuser.id).chat_monitor_enabled:
            _s(tuser.id).chat_monitor_enabled = False
            if _s(tuser.id).chat_monitor_task:
                _s(tuser.id).chat_monitor_task.cancel()
                _s(tuser.id).chat_monitor_task = None
            await edit_menu(query,
                "💬 *Chat Monitor stopped.*\n\n" + orders_section_text(tuser.id),
                orders_section_keyboard(tuser.id)
            )
        else:
            # Conflict guard: block chat monitor while auto-update is running
            if _s(tuser.id).refresh_running:
                await edit_menu(query,
                    "⚠️ <b>Cannot start Chat Monitor</b>\n\n"
                    "<b>Ad Auto-Update</b> is currently running.\n\n"
                    "Running both simultaneously overloads the bot and causes delays "
                    "for all users.\n\n"
                    "Please stop Auto-Update first, then start Chat Monitor.",
                    InlineKeyboardMarkup(back_section("section_orders"))
                )
                return
            # Set flag BEFORE creating task so UI reflects it immediately
            _s(tuser.id).chat_monitor_enabled = True
            _s(tuser.id).chat_monitor_task = asyncio.create_task(
                chat_monitor_loop(context.bot, chat_id)
            )
            await edit_menu(query,
                "💬 *Chat Monitor started!*\nPolling Bybit order chats every 8 seconds.\n\n"
                + orders_section_text(tuser.id),
                orders_section_keyboard(tuser.id)
            )

    # ── ↩️ Chat Reply — set reply state ──
    elif data.startswith("chatreply_"):
        # Format: chatreply_{order_id}_{nick}
        parts    = data.split("_", 2)
        order_id = parts[1] if len(parts) > 1 else ""
        nick     = parts[2] if len(parts) > 2 else "counterparty"
        _s(tuser.id).reply_state[chat_id] = {"order_id": order_id, "nick": nick}
        _btn_state["action"]       = "chat_reply"
        _btn_state["prev_section"] = "section_orders"
        try:
            await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([]))
        except Exception:
            pass
        await context.bot.send_message(
            chat_id=chat_id,
            text=(
                f"↩️ <b>Reply to {nick}</b>\n"
                f"Order: <code>{order_id}</code>\n\n"
                "Type your message and send it.\n"
                "_Tap ❌ Cancel to cancel._"
            ),
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("❌ Cancel Reply", callback_data="cancel_chat_reply")
            ]]),
            parse_mode="HTML"
        )

    # ── ❌ Cancel Chat Reply ──
    elif data == "cancel_chat_reply":
        _s(tuser.id).reply_state.pop(chat_id, None)
        _btn_state["action"] = None
        await context.bot.send_message(
            chat_id=chat_id,
            text="❌ Reply cancelled.",
        )

    # ── 🔔 Toggle Order Monitor ──
    elif data == "toggle_order_monitor":
        if _s(tuser.id).order_monitor_running:
            _s(tuser.id).order_monitor_running = False
            if _s(tuser.id).order_monitor_task:
                _s(tuser.id).order_monitor_task.cancel()
                _s(tuser.id).order_monitor_task = None
            await edit_menu(query,
                "🔕 *Order monitoring stopped.*\n\n" + orders_section_text(tuser.id),
                orders_section_keyboard(tuser.id)
            )
        else:
            # Conflict guard: block order monitor while auto-update is running
            if _s(tuser.id).refresh_running:
                await edit_menu(query,
                    "⚠️ <b>Cannot start Order Monitor</b>\n\n"
                    "<b>Ad Auto-Update</b> is currently running.\n\n"
                    "Running both simultaneously overloads the bot and causes delays "
                    "for all users.\n\n"
                    "Please stop Auto-Update first, then start Order Monitor.",
                    InlineKeyboardMarkup(back_section("section_orders"))
                )
                return
            _s(tuser.id).order_monitor_task = asyncio.create_task(
                order_monitor_loop(context.bot, chat_id)
            )
            # _s(tuser.id).order_monitor_running is set to True inside the loop itself,
            # but we set it here immediately so the UI reflects it instantly
            _s(tuser.id).order_monitor_running = True
            await edit_menu(query,
                "🔔 *Order monitoring started!*\nChecking every 10 seconds.\n\n"
                + orders_section_text(tuser.id),
                orders_section_keyboard(tuser.id)
            )

    # ── 📋 Check Orders Now ──
    elif data == "check_orders_now":
        await edit_menu(query, "⏳ Checking for orders...", orders_section_keyboard(tuser.id))
        result   = await asyncio.get_event_loop().run_in_executor(None, partial(get_pending_orders, creds=get_user_creds(tuser.id)))
        ret_code = result.get("retCode", result.get("ret_code",-1))
        if ret_code == 0:
            items = result.get("result",{}).get("items",[])
            txt   = f"📦 Found `{len(items)}` active order(s)." if items else "📦 No active orders at this time."
        else:
            txt = f"❌ `{result.get('retMsg','')}`"
        await edit_menu(query, txt + "\n\n" + orders_section_text(tuser.id), orders_section_keyboard(tuser.id))

    # ── 🗑 Clear Seen Orders ──
    elif data == "clear_seen_orders":
        _s(tuser.id).seen_order_ids.clear(); _s(tuser.id).seen_sell_ids.clear()
        await edit_menu(query,
            "✅ All seen orders cleared. Bot will re-notify on next check.\n\n" + orders_section_text(tuser.id),
            orders_section_keyboard(tuser.id)
        )

    # ── ✉️ Toggle Sell Msg ──
    elif data == "toggle_sell_msg":
        _s(tuser.id).sell_msg_enabled = not _s(tuser.id).sell_msg_enabled
        await edit_menu(query, orders_section_text(tuser.id), orders_section_keyboard(tuser.id))

    # ── ✏️ Set Sell Message ──
    elif data == "set_sell_msg":
        _btn_state["action"]       = "sell_custom_msg"
        _btn_state["prev_section"] = "section_orders"
        cur = _s(tuser.id).sell_custom_msg[:80] + "..." if len(_s(tuser.id).sell_custom_msg) > 80 else _s(tuser.id).sell_custom_msg
        await edit_menu(query,
            f"✏️ <b>Set Sell Order Message</b>\n\nCurrent:\n_{cur}_\n\n"
            "Send your new custom message to send to buyers on SELL orders.",
            InlineKeyboardMarkup(back_section("section_orders"))
        )

    # ── 🔢 Set Message Count ──
    elif data == "set_sell_msg_count":
        _btn_state["action"]       = "sell_msg_count"
        _btn_state["prev_section"] = "section_orders"
        await edit_menu(query,
            f"🔢 <b>Set Message Count</b>\n\nCurrent: <code>{_s(tuser.id).sell_msg_count}x</code>\n\n"
            "How many times to send to buyer? (1–5)",
            InlineKeyboardMarkup(back_section("section_orders"))
        )

    # ── 🆔 Set Ad ID ──
    elif data == "set_ad_id":
        _btn_state["action"]       = "ad_id"
        _btn_state["prev_section"] = "section_ads"
        sess = _s(tuser.id)
        slot_idx = sess.editing_slot
        if slot_idx == -1:
            slot_str = _get_user_slot_str(tuser.id)
            cur = (
                sess.settings.get(f"ad_id_{slot_str}", "")
                or sess.settings.get("ad_id", "")
                or "Not set"
            )
            label = f"Account {slot_str}"
        else:
            cur = _ad_settings(sess, slot_idx).get("ad_id", "") or "Not set"
            label = _ad_slot_label(slot_idx)
        await edit_menu(query,
            f"🆔 <b>Set Ad ID — {label}</b>\n\nCurrent: <code>{_esc(cur)}</code>\n\n"
            "Send your Bybit Ad ID.\n💡 Use 📃 My Ads List to find it.\n\n"
            "Example: `2040156088201854976`",
            InlineKeyboardMarkup(back_section("section_ads"))
        )

    # ── 👤 Set UID ──
    elif data == "set_uid":
        _btn_state["action"]       = "bybit_uid"
        _btn_state["prev_section"] = "section_ads"
        slot_str = _get_user_slot_str(tuser.id)
        # Read the slot-keyed value first (what ads_section_text displays),
        # fall back to the generic key for backwards compatibility
        cur = (
            _s(tuser.id).settings.get(f"bybit_uid_{slot_str}", "")
            or _s(tuser.id).settings.get("bybit_uid", "")
            or "Not set"
        )
        await edit_menu(query,
            f"👤 <b>Set Bybit UID — Account {slot_str}</b>\n\nCurrent: <code>{_esc(cur)}</code>\n\n"
            "Bybit App → Profile → copy UID under your username.\n\n"
            "Example: `520097760`",
            InlineKeyboardMarkup(back_section("section_ads"))
        )

    # ── 🗑 Delete UID ──
    elif data == "delete_uid":
        slot_str = _get_user_slot_str(tuser.id)
        cur = (
            _s(tuser.id).settings.get(f"bybit_uid_{slot_str}", "")
            or _s(tuser.id).settings.get("bybit_uid", "")
            or "Not set"
        )
        await edit_menu(query,
            f"🗑 <b>Delete UID — Account {slot_str}</b>\n\n"
            f"Current UID: <code>{_esc(cur)}</code>\n\n"
            f"This removes the UID for Account {slot_str} only.\n"
            f"Other accounts and users are not affected.\n\n"
            f"Tap confirm to delete.",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Yes, Delete UID", callback_data="delete_uid_confirm")],
                [InlineKeyboardButton("❌ Cancel",          callback_data="set_uid")],
            ])
        )

    elif data == "delete_uid_confirm":
        slot_str = _get_user_slot_str(tuser.id)
        _s(tuser.id).settings[f"bybit_uid_{slot_str}"] = ""
        # Clear generic key only if it was pointing at this slot's value
        if _s(tuser.id).settings.get("bybit_uid") == _s(tuser.id).settings.get(f"bybit_uid_{slot_str}", ""):
            _s(tuser.id).settings["bybit_uid"] = ""
        # In all cases, sync generic key from slot key (which is now "")
        _s(tuser.id).settings["bybit_uid"] = ""
        _save_settings(tuser.id)
        logger.info(f"[UID] Deleted bybit_uid for user={tuser.id} slot={slot_str}")
        await edit_menu(query,
            f"✅ <b>UID deleted for Account {slot_str}.</b>\n\n"
            f"Tap 👤 Set UID to enter a new one.",
            InlineKeyboardMarkup(back_section("section_ads"))
        )

    # ── 📃 My Ads ──
    elif data == "fetch_my_ads":
        uid   = tuser.id
        creds = get_user_creds(tuser.id)
        # Guard: non-admin user with no API key saved
        # Guard: non-admin user with no API key saved
        if not is_admin(tuser.id) and not creds.get("key"):
            await edit_menu(query,
                "\u274c *No Bybit API set.*\n\nGo to \U0001f511 *Set APIs* \u2192 Set Bybit Account 1 API to add your key first.",
                InlineKeyboardMarkup(back_section("section_ads"))
            )
            return
            return
        await edit_menu(query, "⏳ Fetching your ads...", ads_section_keyboard(tuser.id))
        result   = await asyncio.get_event_loop().run_in_executor(None, partial(get_my_ads, creds=creds))
        ret_code = result.get("retCode", result.get("ret_code",-1))
        if ret_code == 0:
            items = result.get("result",{}).get("items",[])
            if not items:
                await edit_menu(query, "📃 No ads found.", InlineKeyboardMarkup(back_section("section_ads")))
                return
            bybit_uid = _s(tuser.id).settings.get("bybit_uid","")
            lines = ["📃 *Your P2P Ads:*\n"]
            for item in items:
                if bybit_uid and str(item.get("userId","")) != str(bybit_uid):
                    continue
                side  = "BUY" if str(item.get("side","")) == "0" else "SELL"
                stat  = {10:"🟢",20:"🔴",30:"✅"}.get(item.get("status",0),"❓")
                lines.append(
                    f"{stat} <b>{side}</b> <code>{item.get('tokenId','')}/{item.get('currencyId','')}</code>"
                    f" | 💲<code>{item.get('price','')}</code>\n🆔 <code>{item.get('id','')}</code>\n"
                )
            if len(lines) == 1: lines.append("No ads match your UID.")
            lines.append("\n_Tap any ID to copy → use 🆔 Set Ad ID_")
            msg = "\n".join(lines)
            if len(msg) > 4000: msg = msg[:4000] + "...(truncated)"
            await edit_menu(query, msg, InlineKeyboardMarkup(back_section("section_ads")))
        else:
            await edit_menu(query,
                f"❌ <code>{result.get('retMsg',result.get('ret_msg',''))}</code>",
                InlineKeyboardMarkup(back_section("section_ads"))
            )

    # ── 📋 Fetch Ad Details ──
    elif data == "fetch_ad":
        sess = _s(tuser.id)
        slot_idx = sess.editing_slot
        s = _ad_settings(sess, slot_idx)
        if not s.get("ad_id"):
            await edit_menu(query, "❌ Set your Ad ID first.", InlineKeyboardMarkup(back_section("section_ads")))
            return
        _creds = get_user_creds(tuser.id)
        if not is_admin(tuser.id) and not _creds.get("key"):
            await edit_menu(query,
                "\u274c *No Bybit API set.*\n\nGo to \U0001f511 *Set APIs* \u2192 Set Bybit Account 1 API first.",
                InlineKeyboardMarkup(back_section("section_ads"))
            )
            return
        await edit_menu(query, "⏳ Loading ad from Bybit...", ads_section_keyboard(tuser.id))
        result   = await asyncio.get_event_loop().run_in_executor(
            None, partial(get_ad_details, s["ad_id"], creds=_creds)
        )
        ret_code = result.get("retCode", result.get("ret_code",-1))
        if ret_code == 0:
            ad_data = _ad_data_of(sess, slot_idx)
            ad_data.clear()
            ad_data.update(result.get("result",{}))
            token    = ad_data.get("tokenId","—")
            currency = ad_data.get("currencyId","—")
            max_pct  = get_max_float_pct(currency, token)
            ad_stat  = {10:"🟢 Online",20:"🔴 Offline",30:"✅ Done"}.get(ad_data.get("status"),"?")
            next_hint = next_setup_hint(tuser.id) if slot_idx == -1 else "Now set mode + interval for this ad."
            await edit_menu(query,
                f"✅ <b>Ad Loaded! ({_ad_slot_label(slot_idx)})</b>\n\n"
                f"🆔 <code>{s['ad_id']}</code>\n"
                f"💱 <code>{token}/{currency}</code> | 💲 <code>{ad_data.get('price','')}</code>\n"
                f"Min: <code>{ad_data.get('minAmount','')}</code> | Max: <code>{ad_data.get('maxAmount','')}</code> | Qty: <code>{ad_data.get('lastQuantity','')}</code>\n"
                f"Status: {ad_stat} | Max float: <code>{max_pct}%</code>\n\n"
                f"_{next_hint}_",
                InlineKeyboardMarkup(back_section("section_ads"))
            )
        else:
            await edit_menu(query,
                f"❌ <code>{result.get('retMsg',result.get('ret_msg',''))}</code>",
                InlineKeyboardMarkup(back_section("section_ads"))
            )

    # ── 🔀 Switch Mode ──
    elif data == "switch_mode":
        sess = _s(tuser.id)
        slot_idx = sess.editing_slot
        s = _ad_settings(sess, slot_idx)
        new_mode = "floating" if s.get("mode") == "fixed" else "fixed"
        s["mode"] = new_mode
        next_hint = ""
        if slot_idx == -1:
            slot_str = _get_user_slot_str(tuser.id)
            sess.settings[f"mode_{slot_str}"] = new_mode
            _save_settings(tuser.id)
            next_hint = f"\n\n_{next_setup_hint(tuser.id)}_"
        note = " (takes effect next cycle)" if _ad_running(sess, slot_idx) else ""
        await edit_menu(query,
            f"🔀 <b>{_ad_slot_label(slot_idx)} switched to {new_mode.upper()}{note}</b>{next_hint}",
            InlineKeyboardMarkup(back_section("section_ads"))
        )

    # ── ➕ Set Increment ──
    elif data == "set_increment":
        _btn_state["action"]       = "increment"
        _btn_state["prev_section"] = "section_ads"
        sess = _s(tuser.id)
        s = _ad_settings(sess, sess.editing_slot)
        await edit_menu(query,
            f"➕ <b>Set Increment — {_ad_slot_label(sess.editing_slot)}</b>\n\nCurrent: <code>+{s.get('increment','0.05')}</code> per cycle\n\n"
            "Send the amount to add each cycle.\nExamples: `0.05` | `1` | `0.5`",
            InlineKeyboardMarkup(back_section("section_ads"))
        )

    # ── 📊 Set Float % ──
    elif data == "set_float_pct":
        sess = _s(tuser.id)
        slot_idx = sess.editing_slot
        ad_data = _ad_data_of(sess, slot_idx)
        if not ad_data:
            await edit_menu(query, "❌ Fetch Ad Details first.", InlineKeyboardMarkup(back_section("section_ads")))
            return
        token    = ad_data.get("tokenId","USDT").upper()
        currency = ad_data.get("currencyId","NGN").upper()
        max_pct  = get_max_float_pct(currency, token)
        min_pct  = get_min_float_pct(currency, token)
        needs_ref = currency_needs_ref(currency) or currency == "NGN"
        _btn_state["action"]       = "float_pct"
        _btn_state["prev_section"] = "section_ads"
        s   = _ad_settings(sess, slot_idx)
        cur = s.get("float_pct","") or "Not set"
        formula = (
            f"<code>{token}/USDT × {currency}/USDT ref × your% ÷ 100</code>"
            if needs_ref else
            f"<code>{token}/USDT × your% ÷ 100</code>"
        )
        other_pcts = sess.get_active_float_pcts(exclude_index=slot_idx, currency_id=currency, token_id=token)
        gap_note = (
            f"\n\nℹ️ Your other active {token}/{currency} ad(s) are also using: "
            f"{', '.join(f'{p}%' for p in other_pcts)} — that's fine, using the same % is allowed. "
            f"The bot automatically keeps the posted PRICES far enough apart."
            if other_pcts else ""
        )
        await edit_menu(query,
            f"📊 <b>Set Float % — {_ad_slot_label(slot_idx)}</b>\n\nPair: <code>{token}/{currency}</code> | Range: <code>{min_pct}%–{max_pct}%</code>\nCurrent: <code>{cur}</code>\n\n"
            f"Formula: {formula}{gap_note}\n\n"
            f"Send a value between <code>{min_pct}</code> and <code>{max_pct}</code>. Example: <code>105</code>",
            InlineKeyboardMarkup(back_section("section_ads"))
        )

    # ── 💱 Set NGN Ref ──
    elif data == "set_ngn_ref":
        _btn_state["action"]       = "ngn_usdt_ref"
        _btn_state["prev_section"] = "section_ads"
        sess = _s(tuser.id)
        ad_data = _ad_data_of(sess, sess.editing_slot)
        _rcur = ad_data.get("currencyId","NGN").upper() if ad_data else "NGN"
        cur   = sess.shared_local_usdt_ref or "Not set"
        shared_note = " (shared by all your ads)" if sess.total_ad_slots() > 1 else ""
        await edit_menu(query,
            f"💱 <b>{_rcur}/USDT Reference Price{shared_note}</b>\n\nCurrent: <code>{cur}</code>\n\n"
            f"Check Bybit P2P market for current {_rcur}/USDT rate.\n"
            f"Example: <code>{'1580' if _rcur == 'NGN' else '1.25' if _rcur == 'EUR' else '100'}</code> ({_rcur} per 1 USDT)",
            InlineKeyboardMarkup(back_section("section_ads"))
        )

    # ── ⏱ Set Interval ──
    elif data == "set_interval":
        _btn_state["action"]       = "interval"
        _btn_state["prev_section"] = "section_ads"
        sess = _s(tuser.id)
        s = _ad_settings(sess, sess.editing_slot)
        await edit_menu(query,
            f"⏱ <b>Set Interval — {_ad_slot_label(sess.editing_slot)}</b>\n\nCurrent: every <code>{s.get('interval',2)}</code> min\n\n"
            f"Send minutes between each price update (minimum {bybit.MIN_AD_INTERVAL_MINUTES}).\nExamples: `2` | `5` | `10`",
            InlineKeyboardMarkup(back_section("section_ads"))
        )

    # ── 🔄 Update Once Now ──
    elif data == "update_now":
        if _s(tuser.id).total_ad_slots() > 1:
            await edit_menu(query,
                "❌ <b>Update Once Now</b> is only available with a single active ad.\n\n"
                "Stop or remove your extra ads first if you need a one-off manual update.",
                InlineKeyboardMarkup(back_section("section_ads"))
            )
            return
        if not _s(tuser.id).ad_data or not _s(tuser.id).settings.get("ad_id"):
            await edit_menu(query, "❌ Load ad details first.", InlineKeyboardMarkup(back_section("section_ads")))
            return
        # ── Conflict guard: block manual update while order/chat monitor is running ──
        _sess_chk2 = _s(tuser.id)
        if _sess_chk2.order_monitor_running or _sess_chk2.chat_monitor_enabled:
            _active2 = []
            if _sess_chk2.order_monitor_running: _active2.append("Order Monitor")
            if _sess_chk2.chat_monitor_enabled:  _active2.append("Chat Monitor")
            await edit_menu(query,
                "⚠️ <b>Cannot update while monitoring is active</b>\n\n"
                f"<b>{' and '.join(_active2)}</b> is currently running.\n\n"
                "Stop your active monitors first before updating ad price.",
                InlineKeyboardMarkup(back_section("section_ads"))
            )
            return
        # Load per-user creds — MUST be done before modify_ad
        _update_creds = get_user_creds(tuser.id)
        if not _update_creds or not _update_creds.get("key"):
            await edit_menu(query,
                "❌ <b>No Bybit API key found.</b>\n\nGo to 🔑 Set APIs → Set Bybit API first.",
                InlineKeyboardMarkup(back_section("section_ads")))
            return
        mode = _s(tuser.id).settings.get("mode","fixed")
        await edit_menu(query, f"⏳ Updating ({mode} mode)...", ads_section_keyboard(tuser.id))
        if mode == "fixed":
            # ── FIX: always compute the NEXT price (base + increment), not the last applied price.
            # current_price is 0 if the auto-loop has never run, so we start from ad_data["price"].
            # This mirrors exactly what auto_update_loop does on each cycle.
            _increment = Decimal(str(_s(tuser.id).settings.get("increment", "0.05")))
            _base = _s(tuser.id).current_price if _s(tuser.id).current_price else Decimal(str(_s(tuser.id).ad_data.get("price", "0")))
            _next_price = _base + _increment
            price = str(_next_price.quantize(Decimal("0.00000001"), rounding=ROUND_HALF_UP))
        else:
            # ── FIX: call calc_floating_price directly (it is sync) — do NOT wrap in
            # run_in_executor which causes tuple-unpacking to silently fail.
            try:
                float_pct = float(_s(tuser.id).settings.get("float_pct") or 0)
            except (TypeError, ValueError):
                float_pct = 0
            if float_pct <= 0:
                await edit_menu(query,
                    "❌ Float % isn't set yet. Tap 📊 Set Float % first, then Update Once Now.",
                    InlineKeyboardMarkup(back_section("section_ads")))
                return
            local_usdt_ref = float(_s(tuser.id).settings.get("local_usdt_ref") or 0)
            price, err     = calc_floating_price(_s(tuser.id).ad_data, float_pct, local_usdt_ref)
            if err:
                await edit_menu(query, f"❌ <code>{_esc(str(err))}</code>", InlineKeyboardMarkup(back_section("section_ads")))
                return
        _record_modify_ad1(_s(tuser.id))
        result = await asyncio.get_event_loop().run_in_executor(
            _ad_executor, modify_ad, _s(tuser.id).settings["ad_id"], price, _s(tuser.id).ad_data, _update_creds
        )
        rc = result.get("retCode", result.get("ret_code",-1))
        rm = result.get("retMsg",  result.get("ret_msg",""))
        if rc == 912120022:
            # Same bounded retry as auto_update_loop — see there for why a
            # single exact-boundary attempt wasn't reliable enough.
            for _attempt in range(3):
                min_str, max_str = _extract_bybit_bounds(rm)
                was_too_high = max_str is not None
                bound_str = max_str if was_too_high else min_str
                if not bound_str:
                    break
                bound_dec = Decimal(bound_str)
                if _attempt < 2:
                    candidate_str = bound_str   # Bybit's own string, exactly — see auto_update_loop for why
                else:
                    margin        = _safety_margin(bound_dec)
                    candidate     = (bound_dec - margin) if was_too_high else (bound_dec + margin)
                    safe_rounding = ROUND_FLOOR if was_too_high else ROUND_CEILING
                    candidate_str = str(candidate.quantize(Decimal("0.01"), rounding=safe_rounding))
                _record_modify_ad1(_s(tuser.id))
                result = await asyncio.get_event_loop().run_in_executor(
                    _ad_executor, modify_ad, _s(tuser.id).settings["ad_id"], candidate_str, _s(tuser.id).ad_data, _update_creds
                )
                rc    = result.get("retCode", result.get("ret_code",-1))
                rm    = result.get("retMsg",  result.get("ret_msg",""))
                price = candidate_str
                if rc == 0 or rc != 912120022:
                    break
        elif rc == 90043:
            # Price rounds to the same value the ad already has live —
            # nudge by this pair's minimum gap and retry once (same
            # recoverable case as in auto_update_loop, see there for why).
            _ad_data_now = _s(tuser.id).ad_data
            _nudge = get_min_price_gap(_ad_data_now.get("currencyId",""), _ad_data_now.get("tokenId",""), Decimal(str(price)))
            _nudged_price = str((Decimal(str(price)) - _nudge).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            _record_modify_ad1(_s(tuser.id))
            result = await asyncio.get_event_loop().run_in_executor(
                _ad_executor, modify_ad, _s(tuser.id).settings["ad_id"], _nudged_price, _ad_data_now, _update_creds
            )
            rc    = result.get("retCode", result.get("ret_code",-1))
            rm    = result.get("retMsg",  result.get("ret_msg",""))
            price = _nudged_price
        if rc == 0:
            # ── Advance current_price so the next cycle (auto or manual) continues from here
            _s(tuser.id).current_price = Decimal(str(price))
            await edit_menu(query,
                f"✅ <b>Updated!</b> Price: <code>{price}</code> ({mode.upper()})\n\n_{next_setup_hint(tuser.id)}_",
                InlineKeyboardMarkup(back_section("section_ads"))
            )
        else:
            await edit_menu(query, f"❌ <code>{rc}</code> — <code>{_esc(rm)}</code>", InlineKeyboardMarkup(back_section("section_ads")))

    # ── 📢 Post/Remove Ad Manager — independent from auto-update ──
    # ── 📢 Post / Remove Ad Manager ──
    elif data == "post_ad_prompt":
        manage_id = _s(tuser.id).settings.get("manage_ad_id", "")
        mdata     = _s(tuser.id).settings.get("manage_ad_data", {})
        cur_id_line = f"Manage Ad ID: `{manage_id}`" if manage_id else "⚠️ No Manage Ad ID set yet."
        if mdata:
            stat   = {10:"🟢 Online", 20:"🔴 Offline", 30:"✅ Done"}.get(mdata.get("status"), "?")
            loaded = f"\nStatus: {stat} | 💲`{mdata.get('price','—')}`"
        else:
            loaded = "\n_No ad fetched yet._"
        await edit_menu(query,
            f"📢 <b>Post / Remove Ad Manager</b>\n\n"
            f"⚠️ Completely separate from Auto-Update.\n"
            f"Setting IDs here will NOT affect your auto-price bot.\n\n"
            f"{cur_id_line}{loaded}\n\n"
            f"• <b>Post Ad</b> — brings a paused/offline ad back online (same ID)\n"
            f"• <b>Remove Ad</b> — pauses/takes an online ad offline (same ID)",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("🆔 Set Manage Ad ID",       callback_data="set_manage_ad_id")],
                [InlineKeyboardButton("📋 Fetch Manage Ad",        callback_data="fetch_manage_ad")],
                [InlineKeyboardButton("🟢 Post Ad (go online)",    callback_data="post_ad_do")],
                [InlineKeyboardButton("🔴 Remove Ad (go offline)", callback_data="remove_ad_confirm")],
                *back_section("section_ads"),
            ])
        )

    elif data == "set_manage_ad_id":
        _btn_state["action"]       = "manage_ad_id"
        _btn_state["prev_section"] = "post_ad_prompt"
        cur     = _s(tuser.id).settings.get("manage_ad_id", "") or "Not set"
        auto_id = _s(tuser.id).settings.get("ad_id", "not set")
        await edit_menu(query,
            f"🆔 <b>Set Manage Ad ID</b>\n\n"
            f"Current Manage Ad ID: <code>{cur}</code>\n"
            f"Auto-Update Ad ID: <code>{auto_id}</code> (unchanged)\n\n"
            f"Send the Bybit Ad ID you want to post or remove.\n"
            f"Example: <code>2040156088201854976</code>",
            InlineKeyboardMarkup(back_manager())
        )

    elif data == "fetch_manage_ad":
        manage_id = _s(tuser.id).settings.get("manage_ad_id", "")
        if not manage_id:
            await edit_menu(query, "❌ Set a Manage Ad ID first.", InlineKeyboardMarkup(back_manager()))
            return
        await edit_menu(query, f"⏳ Fetching ad `{manage_id}`...", InlineKeyboardMarkup(back_manager()))
        result = await asyncio.get_event_loop().run_in_executor(None, partial(get_ad_details, manage_id, creds=get_user_creds(tuser.id)))
        rc = result.get("retCode", result.get("ret_code", -1))
        if rc == 0:
            mdata = result.get("result", {})
            _s(tuser.id).settings["manage_ad_data"] = mdata
            token    = mdata.get("tokenId", "—")
            currency = mdata.get("currencyId", "—")
            side_val = "BUY" if str(mdata.get("side", "1")) == "0" else "SELL"
            stat     = {10:"🟢 Online", 20:"🔴 Offline", 30:"✅ Done"}.get(mdata.get("status"), "?")
            await edit_menu(query,
                f"✅ <b>Manage Ad Loaded!</b>\n\n"
                f"🆔 <code>{manage_id}</code>\n"
                f"💱 <code>{token}/{currency}</code> | Side: <code>{side_val}</code>\n"
                f"💲 Price: <code>{mdata.get('price','—')}</code> | Qty: <code>{mdata.get('lastQuantity', mdata.get('quantity','—'))}</code>\n"
                f"Min: <code>{mdata.get('minAmount','—')}</code> | Max: <code>{mdata.get('maxAmount','—')}</code>\n"
                f"Status: {stat}\n\n"
                f"<i>Tap Post Ad if offline, or Remove Ad if online.</i>",
                InlineKeyboardMarkup([
                    [InlineKeyboardButton("🟢 Post Ad (go online)",    callback_data="post_ad_do")],
                    [InlineKeyboardButton("🔴 Remove Ad (go offline)", callback_data="remove_ad_confirm")],
                    *back_manager(),
                ])
            )
        else:
            await edit_menu(query,
                f"❌ <code>{result.get('retMsg', result.get('ret_msg',''))}</code>",
                InlineKeyboardMarkup(back_manager())
            )

    # ── 🟢 Post Ad = bring offline ad back ONLINE (LISTING, same ID) ──
    elif data == "post_ad_do":
        mdata     = _s(tuser.id).settings.get("manage_ad_data", {})
        manage_id = _s(tuser.id).settings.get("manage_ad_id", "")
        if not mdata or not manage_id:
            await edit_menu(query, "❌ Fetch Manage Ad details first.", InlineKeyboardMarkup(back_manager()))
            return
        await edit_menu(query, f"⏳ Posting ad `{manage_id}` back online...", InlineKeyboardMarkup(back_manager()))
        result = await asyncio.get_event_loop().run_in_executor(None, partial(put_ad_online, manage_id, mdata, creds=get_user_creds(tuser.id)))
        rc = result.get("retCode", result.get("ret_code", -1))
        rm = result.get("retMsg",  result.get("ret_msg", ""))
        if rc == 0:
            fresh = await asyncio.get_event_loop().run_in_executor(None, partial(get_ad_details, manage_id, creds=get_user_creds(tuser.id)))
            if fresh.get("retCode", -1) == 0:
                _s(tuser.id).settings["manage_ad_data"] = fresh.get("result", mdata)
            await edit_menu(query,
                f"✅ <b>Ad is now Online!</b>\n\n"
                f"🆔 Ad ID: <code>{manage_id}</code> (same — unchanged)\n"
                f"Your ad is live on Bybit P2P.\n\n"
                f"Auto-Update Ad ID: <code>{_s(tuser.id).settings.get('ad_id','not set')}</code> — unchanged.",
                InlineKeyboardMarkup(back_manager())
            )
        else:
            await edit_menu(query,
                f"❌ <b>Failed to post ad online</b>\n\nCode: <code>{rc}</code>\nMessage: <code>{rm}</code>",
                InlineKeyboardMarkup(back_manager())
            )

    # ── 🔴 Remove Ad = take online ad OFFLINE (CANCEL, same ID) ──
    elif data == "remove_ad_confirm":
        manage_id = _s(tuser.id).settings.get("manage_ad_id", "")
        if not manage_id:
            await edit_menu(query,
                "❌ No Manage Ad ID set. Tap 🆔 Set Manage Ad ID first.",
                InlineKeyboardMarkup(back_manager())
            )
            return
        auto_id   = _s(tuser.id).settings.get("ad_id", "")
        same_warn = (
            f"\n\n⚠️ <b>This is also your Auto-Update Ad ID.</b>\n"
            f"Stop auto-price update manually if needed."
        ) if manage_id == auto_id else ""
        await edit_menu(query,
            f"🔴 <b>Remove Ad (go offline)?</b>\n\n"
            f"Manage Ad ID: <code>{manage_id}</code>\n"
            f"Auto-Update Ad ID: <code>{auto_id or 'not set'}</code> (unchanged)\n"
            f"{same_warn}\n\n"
            f"Ad will be paused/taken offline. Same ID — not permanently deleted.\n"
            f"Bring it back online anytime with Post Ad.",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Yes, Take Offline", callback_data="remove_ad_do")],
                [InlineKeyboardButton("❌ Cancel",            callback_data="post_ad_prompt")],
            ])
        )

    elif data == "remove_ad_do":
        mdata     = _s(tuser.id).settings.get("manage_ad_data", {})
        manage_id = _s(tuser.id).settings.get("manage_ad_id", "")
        if not manage_id:
            await edit_menu(query, "❌ No Manage Ad ID set.", InlineKeyboardMarkup(back_manager()))
            return
        await edit_menu(query, f"⏳ Taking ad `{manage_id}` offline...", InlineKeyboardMarkup(back_manager()))
        result = await asyncio.get_event_loop().run_in_executor(None, partial(take_ad_offline, manage_id, mdata, creds=get_user_creds(tuser.id)))
        rc = result.get("retCode", result.get("ret_code", -1))
        rm = result.get("retMsg",  result.get("ret_msg", ""))
        if rc == 0:
            fresh = await asyncio.get_event_loop().run_in_executor(None, partial(get_ad_details, manage_id, creds=get_user_creds(tuser.id)))
            if fresh.get("retCode", -1) == 0:
                _s(tuser.id).settings["manage_ad_data"] = fresh.get("result", mdata)
            await edit_menu(query,
                f"✅ <b>Ad is now Offline (Paused)!</b>\n\n"
                f"🆔 Ad ID: <code>{manage_id}</code> (same — not deleted)\n"
                f"Bring it back online anytime using Post Ad.\n\n"
                f"Auto-Update Ad ID: <code>{_s(tuser.id).settings.get('ad_id','not set')}</code> — unchanged.",
                InlineKeyboardMarkup(back_manager())
            )
        else:
            await edit_menu(query,
                f"❌ <b>Failed to take ad offline</b>\n\nCode: <code>{rc}</code>\nMessage: <code>{rm}</code>",
                InlineKeyboardMarkup(back_manager())
            )


    # ── 🔑 API Setup Section ──
    elif data == "section_apis":
        uid  = query.from_user.id
        bk1  = "✅" if db.get_api(uid, "bybit_key_1")    else "❌"
        bk2  = "✅" if db.get_api(uid, "bybit_key_2")    else "❌"
        # FLW is fully configured only when all 3 keys are saved
        flw_keys = all(db.get_api(uid, k) for k in (
            "flw_public_key", "flw_secret_hash", "flw_secret_key"
        ))
        fk   = "✅" if flw_keys else "❌"
        pk   = "✅" if db.get_api(uid, "paga_principal") else "❌"
        await edit_menu_html(query,
            f"🔑 <b>API Setup</b>\n\n"
            f"Your API keys are stored securely on the server.\n\n"
            f"Bybit Account 1 API: {bk1}\n"
            f"Bybit Account 2 API: {bk2}\n"
            f"Flutterwave API (3 keys): {fk}\n"
            f"Paga API: {pk}\n\n"
            f"⚠️ Keys are stored per user and never shared.\n"
            f"⚠️ FLW and Paga work across both Bybit accounts.",
            InlineKeyboardMarkup([
                [InlineKeyboardButton(f"🔑 {bk1} Set Bybit Account 1 API", callback_data="set_api_bybit_1")],
                [InlineKeyboardButton(f"🔑 {bk2} Set Bybit Account 2 API", callback_data="set_api_bybit_2")],
                [InlineKeyboardButton(f"🟢 {fk} Set Flutterwave API",      callback_data="set_api_flw")],
                [InlineKeyboardButton(f"🟡 {pk} Set Paga API",             callback_data="set_api_paga")],
                [InlineKeyboardButton("🗑 Delete All APIs",                 callback_data="delete_apis")],
                *back_main()
            ])
        )

    elif data == "set_api_bybit":
        # Legacy callback — redirect to account 1
        _btn_state["action"]       = "api_bybit_key_1"
        _btn_state["prev_section"] = "section_apis"
        _btn_state["_api_bybit_slot"] = "1"
        uid = query.from_user.id
        has = bool(db.get_api(uid, "bybit_key_1"))
        await edit_menu(query,
            f"🔑 <b>Set Bybit Account 1 API Key</b>\n\n"
            f"Status: {'✅ Key saved — new key will replace it' if has else '❌ Not set'}\n\n"
            "Send your Bybit API Key for Account 1.",
            InlineKeyboardMarkup(back_section("section_apis"))
        )

    elif data == "set_api_bybit_1":
        _btn_state["action"]          = "api_bybit_key_1"
        _btn_state["prev_section"]    = "section_apis"
        _btn_state["_api_bybit_slot"] = "1"
        uid = query.from_user.id
        has = bool(db.get_api(uid, "bybit_key_1"))
        await edit_menu(query,
            f"🔑 <b>Set Bybit Account 1 API Key</b>\n\n"
            f"Status: {'✅ Key saved — new key will replace it' if has else '❌ Not set'}\n\n"
            "Send your Bybit API Key for Account 1.",
            InlineKeyboardMarkup(back_section("section_apis"))
        )

    elif data == "set_api_bybit_2":
        _btn_state["action"]          = "api_bybit_key_2"
        _btn_state["prev_section"]    = "section_apis"
        _btn_state["_api_bybit_slot"] = "2"
        uid = query.from_user.id
        has = bool(db.get_api(uid, "bybit_key_2"))
        await edit_menu(query,
            f"🔑 <b>Set Bybit Account 2 API Key</b>\n\n"
            f"Status: {'✅ Key saved — new key will replace it' if has else '❌ Not set'}\n\n"
            "Send your Bybit API Key for Account 2.",
            InlineKeyboardMarkup(back_section("section_apis"))
        )

    elif data == "set_api_flw":
        _btn_state["action"]       = "api_flw_public_key"
        _btn_state["prev_section"] = "section_apis"
        uid = query.from_user.id
        has = bool(db.get_api(uid, "flw_secret_key"))
        status_line = "✅ Already configured — new values will replace existing ones" if has else "❌ Not yet configured"
        await edit_menu_html(query,
            f"🟢 <b>Set Flutterwave API</b>\n\n"
            f"Status: {status_line}\n\n"
            f"You will enter <b>3 credentials</b> one at a time:\n"
            f"  1️⃣ FLW_PUBLIC_KEY\n"
            f"  2️⃣ FLW_SECRET_HASH\n"
            f"  3️⃣ FLW_SECRET_KEY\n\n"
            f"<b>Step 1 of 3:</b> Send your <b>FLW_PUBLIC_KEY</b>\n"
            f"<i>(starts with FLWPUBK_ — Flutterwave dashboard → Settings → API)</i>",
            InlineKeyboardMarkup(back_section("section_apis"))
        )

    elif data == "set_api_paga":
        _btn_state["action"]       = "api_paga_api_key"
        _btn_state["prev_section"] = "section_apis"
        uid = query.from_user.id
        has = bool(db.get_api(uid, "paga_principal"))
        await edit_menu(query,
            f"🟡 <b>Set Paga API</b>\n\n"
            f"Status: {'✅ Already configured — new values will replace it' if has else '❌ Not set'}\n\n"
            "Step 1 of 3: Send your *PAGA_API_KEY*\n_(HMAC Hash Key from Paga dashboard)_",
            InlineKeyboardMarkup(back_section("section_apis"))
        )

    elif data == "delete_apis":
        uid_d = query.from_user.id
        bk1 = "✅" if db.get_api(uid_d, "bybit_key_1")    else "—"
        bk2 = "✅" if db.get_api(uid_d, "bybit_key_2")    else "—"
        fk  = "✅" if db.get_api(uid_d, "flw_secret_key")  else "—"
        pk  = "✅" if db.get_api(uid_d, "paga_principal") else "—"
        await edit_menu(query,
            f"🗑 <b>Delete API Keys</b>\n\n"
            f"Choose which keys to delete. This cannot be undone.\n\n"
            f"Bybit Account 1: {bk1}\n"
            f"Bybit Account 2: {bk2}\n"
            f"Flutterwave: {fk}\n"
            f"Paga: {pk}",
            InlineKeyboardMarkup([
                [InlineKeyboardButton(f"🔑 Delete Bybit Acct 1 API {bk1}", callback_data="delete_bybit1_apis")],
                [InlineKeyboardButton(f"🔑 Delete Bybit Acct 2 API {bk2}", callback_data="delete_bybit2_apis")],
                [InlineKeyboardButton(f"🟢 Delete Flutterwave API {fk}",   callback_data="delete_flw_apis")],
                [InlineKeyboardButton(f"🟡 Delete Paga API {pk}",          callback_data="delete_paga_apis")],
                [InlineKeyboardButton("🗑 Delete ALL APIs",                 callback_data="delete_apis_confirm")],
                [InlineKeyboardButton("❌ Cancel",                          callback_data="section_apis")],
            ])
        )

    elif data == "delete_apis_confirm":
        uid_del = query.from_user.id
        db.delete_all_apis(uid_del)
        await edit_menu(query,
            "✅ *All API keys deleted.*\n\n"
            "Your account is still active but API credentials have been removed.\n"
            "Re-enter them anytime via 🔑 Set APIs.",
            InlineKeyboardMarkup([*back_section("section_apis")])
        )

    # ── Granular delete confirmations ──
    elif data == "delete_bybit1_apis":
        uid_d = query.from_user.id
        has   = bool(db.get_api(uid_d, "bybit_key_1"))
        await edit_menu(query,
            f"🔑 <b>Delete Bybit Account 1 API?</b>\n\n"
            f"Status: {'✅ Saved' if has else '❌ Already empty'}\n\n"
            "This permanently removes your Account 1 API key and secret.",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Yes, Delete", callback_data="delete_bybit1_confirm")],
                [InlineKeyboardButton("❌ Cancel",       callback_data="delete_apis")],
            ])
        )

    elif data == "delete_bybit1_confirm":
        uid_del = query.from_user.id
        db.save_api(uid_del, "bybit_key_1",    "")
        db.save_api(uid_del, "bybit_secret_1", "")
        # If this user is currently on slot 1, reset their slot to 0 (no global change)
        if _s(uid_del).selected_slot == 0:
            logger.info(f"[APIs] Bybit Account 1 keys deleted for user {uid_del} (was on slot 1)")
        await edit_menu(query,
            "✅ *Bybit Account 1 API deleted.*\n\nYou can re-add it anytime via 🔑 Set APIs.",
            InlineKeyboardMarkup([*back_section("section_apis")])
        )

    elif data == "delete_bybit2_apis":
        uid_d = query.from_user.id
        has   = bool(db.get_api(uid_d, "bybit_key_2"))
        await edit_menu(query,
            f"🔑 <b>Delete Bybit Account 2 API?</b>\n\n"
            f"Status: {'✅ Saved' if has else '❌ Already empty'}\n\n"
            "This permanently removes your Account 2 API key and secret.",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Yes, Delete", callback_data="delete_bybit2_confirm")],
                [InlineKeyboardButton("❌ Cancel",       callback_data="delete_apis")],
            ])
        )

    elif data == "delete_bybit2_confirm":
        uid_del = query.from_user.id
        db.save_api(uid_del, "bybit_key_2",    "")
        db.save_api(uid_del, "bybit_secret_2", "")
        # If this user is currently on slot 2, reset their slot to 0 (no global change)
        if _s(uid_del).selected_slot == 1:
            _s(uid_del).selected_slot = 0
            logger.info(f"[APIs] Bybit Account 2 keys deleted for user {uid_del} — slot reset to 1")
        await edit_menu(query,
            "✅ *Bybit Account 2 API deleted.*\n\nYou can re-add it anytime via 🔑 Set APIs.",
            InlineKeyboardMarkup([*back_section("section_apis")])
        )

    elif data == "delete_flw_apis":
        uid_d = query.from_user.id
        has   = bool(db.get_api(uid_d, "flw_secret_key"))
        status_str = "✅ Saved" if has else "❌ Already empty"
        await edit_menu_html(query,
            f"🟢 <b>Delete Flutterwave API?</b>\n\n"
            f"Status: {status_str}\n\n"
            "This permanently removes all 3 FLW credentials:\n"
            "FLW_PUBLIC_KEY, FLW_SECRET_HASH, FLW_SECRET_KEY",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Yes, Delete", callback_data="delete_flw_confirm")],
                [InlineKeyboardButton("❌ Cancel",       callback_data="delete_apis")],
            ])
        )

    elif data == "delete_flw_confirm":
        uid_del = query.from_user.id
        for k in ("flw_public_key", "flw_secret_hash", "flw_secret_key"):
            db.save_api(uid_del, k, "")
        logger.info(f"[APIs] All FLW keys deleted for user {uid_del}")
        await edit_menu_html(query,
            "✅ <b>Flutterwave API deleted.</b>\n\nAll 3 credentials removed.\n"
            "You can re-add them anytime via 🔑 Set APIs.",
            InlineKeyboardMarkup([*back_section("section_apis")])
        )

    elif data == "delete_paga_apis":
        uid_d = query.from_user.id
        has   = bool(db.get_api(uid_d, "paga_principal"))
        await edit_menu(query,
            f"🟡 <b>Delete Paga API?</b>\n\n"
            f"Status: {'✅ Saved' if has else '❌ Already empty'}\n\n"
            "This permanently removes your Paga Principal, Credential and API Key.",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Yes, Delete", callback_data="delete_paga_confirm")],
                [InlineKeyboardButton("❌ Cancel",       callback_data="delete_apis")],
            ])
        )

    elif data == "delete_paga_confirm":
        uid_del = query.from_user.id
        for k in ("paga_principal", "paga_credential", "paga_api_key"):
            db.save_api(uid_del, k, "")
        logger.info(f"[APIs] Paga keys deleted for user {uid_del}")
        await edit_menu(query,
            "✅ *Paga API deleted.*\n\nYou can re-add it anytime via 🔑 Set APIs.",
            InlineKeyboardMarkup([*back_section("section_apis")])
        )

    # ── ⬆️ Upgrade Plan ──
    elif data == "upgrade_plan":
        uid   = query.from_user.id
        badge = sub.plan_badge(uid)
        exp   = db.get_plan_expiry_str(uid)
        user_rec = db.get_user(uid)
        pend  = user_rec.get("upgrade_pending", False) if user_rec else False
        if db.is_pro(uid):
            await edit_menu(query,
                f"💎 <b>You are already on Pro!</b>\n\n{exp}\n\nAll features are unlocked.",
                InlineKeyboardMarkup(back_main())
            )
            return
        if pend:
            await edit_menu(query,
                "⏳ *Upgrade request already pending.*\n\n"
                "The admin will review and approve shortly.\n"
                "You will receive a notification when approved.",
                InlineKeyboardMarkup(back_main())
            )
            return
        await edit_menu(query,
            f"⬆️ <b>Upgrade to Pro Plan</b>\n\n"
            f"Current: {badge}\n\n"
            f"Pro unlocks:\n"
            f"  ✅ Auto Price Update bot\n"
            f"  ✅ Order Monitor + Chat Monitor\n"
            f"  ✅ Auto-Pay (Bybit, FLW, Paga)\n"
            f"  ✅ Buyer Protection & Name Match\n"
            f"  ✅ All ad management features\n\n"
            f"Tap <b>Request Upgrade</b> to send a request to the admin.",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Request Upgrade", callback_data="upgrade_request_yes")],
                [InlineKeyboardButton("❌ Cancel",          callback_data="main_menu")],
            ])
        )

    elif data == "upgrade_request_yes":
        _btn_state["action"] = "upgrade_contact"
        await edit_menu(query,
            "📞 <b>One more step before we submit your request</b>\n\n"
            "Send your Telegram username or WhatsApp number (with country code) so the admin can reach you before approving:\n\n"
            "• Telegram: <code>@johnsmith</code>\n"
            "• WhatsApp: <code>+2348012345678</code>",
            InlineKeyboardMarkup(back_main())
        )
                # The background _upgrade_notifier_loop will retry in 30 s

    # ── 🎵 Convert a downloaded video to audio ──
    elif data.startswith("conv_audio_"):
        # IMPORTANT: this callback's message IS the video itself. edit_menu()
        # tries edit_message_caption() first, which succeeds on a video
        # message too — so calling it here would silently rewrite the
        # video's own caption (and any Main Menu button attached that way
        # would then try to turn this same video message into the full
        # dashboard). Everything below sends separate NEW messages instead,
        # and only ever touches the video's reply_markup (never its caption
        # or the video itself) to clear the button once it's been used.
        uid         = query.from_user.id
        sess        = _s(uid)
        download_id = data[len("conv_audio_"):]
        entry       = sess.video_downloads.get(download_id)   # per-user dict — another user's ID can never match here
        if not entry:
            await query.answer(
                f"⌛ This video has expired (files are auto-deleted after {mediadl.FILE_TTL_SECONDS}s). Send the link again.",
                show_alert=True
            )
            return

        await query.answer("🎵 Converting to audio...")
        result = await asyncio.get_event_loop().run_in_executor(None, mediadl.convert_to_audio, entry["file_path"])
        if not result["ok"]:
            msg = (f"⌛ This video has expired (files are auto-deleted after {mediadl.FILE_TTL_SECONDS} seconds).\n\n"
                   "Send the link again to get a fresh copy.") if result["reason"] == "expired" else f"❌ {result['reason']}"
            await context.bot.send_message(chat_id=uid, text=msg)
            return
        try:
            with open(result["audio_path"], "rb") as f:
                await context.bot.send_audio(chat_id=uid, audio=f, caption="🎵 Converted from your video")
            try:
                await query.edit_message_reply_markup(reply_markup=None)   # clear the button only — video/caption untouched
            except Exception:
                pass
        except Exception as e:
            logger.error(f"[MediaDL] send_audio failed: {type(e).__name__}")
            await context.bot.send_message(chat_id=uid, text="❌ Could not send the audio file.")

    # ── 📢 Broadcast — confirm/cancel ──
    elif data == "broadcast_confirm":
        if not is_admin(tuser.id):
            return
        draft = user_state.get("broadcast_draft")
        if not draft:
            await query.answer("❌ No broadcast draft found — start again with /broadcast.", show_alert=True)
            return
        user_state["broadcast_draft"] = None
        await query.answer("📢 Broadcasting — this runs in the background, you'll get a summary when it's done.", show_alert=True)
        try:
            await query.edit_message_reply_markup(reply_markup=None)   # clear buttons only — preview content untouched
        except Exception:
            pass
        asyncio.create_task(_run_broadcast(context.bot, query.message.chat_id, draft))

    elif data == "broadcast_cancel":
        if not is_admin(tuser.id):
            return
        user_state["broadcast_draft"] = None
        await query.answer("❌ Broadcast cancelled.")
        try:
            await query.edit_message_reply_markup(reply_markup=None)
        except Exception:
            pass

    # ── 💬 Contact Support — prompt for a message ──
    elif data == "contact_support":
        _btn_state["action"] = "support_message"
        await edit_menu(query,
            "💬 <b>Contact Support</b>\n\n"
            "Send your message and the admin will get back to you directly.",
            InlineKeyboardMarkup(back_main())
        )

    # ── 💬 Admin taps Reply on a forwarded support message ──
    elif data.startswith("support_reply_"):
        if not is_admin(tuser.id):
            return
        target_uid = data[len("support_reply_"):]
        try:
            target_uid = int(target_uid)
        except ValueError:
            await query.answer("❌ Invalid user ID.", show_alert=True)
            return
        user_state["support_reply_target"] = target_uid
        user_state["action"] = "support_reply_awaiting"
        await query.answer()
        target_user = db.get_user(target_uid)
        target_label = target_user.get("username") or target_user.get("display_name") or str(target_uid) if target_user else str(target_uid)
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"↩️ Type your reply to @{_esc(target_label)} (<code>{target_uid}</code>):",
            parse_mode="HTML"
        )

    # ── 🎬 Video Downloader — prompt for a link ──
    elif data == "video_downloader":
        _btn_state["action"] = "video_link"
        await edit_menu(query,
            "🎬 <b>Social Media Video Downloader</b>\n\n"
            "Send me a video link (YouTube, TikTok, Instagram, Twitter/X, Facebook, etc.) "
            "and I'll download it for you.\n\n"
            f"⏱ Videos are kept for {mediadl.FILE_TTL_SECONDS} seconds after download — "
            "if you want the audio version, tap 🎵 Convert to Audio right away.",
            InlineKeyboardMarkup(back_main())
        )

    # ── 🎁 Referrals ──
    elif data == "referrals":
        uid  = query.from_user.id
        bot_username = (await context.bot.get_me()).username
        code = db.get_or_create_referral_code(uid)
        link = f"https://t.me/{bot_username}?start=ref_{code}"
        bal  = db.get_referral_balance(uid)
        rows = db.get_referrals_for(uid)

        status_icon = {"none": "⏳", "pending": "💵", "approved": "✅"}
        if rows:
            invite_lines = []
            for r in rows[:15]:
                uname = _esc(r.get("referred_username") or "someone")
                icon  = status_icon.get(r.get("reward_status"), "•")
                invite_lines.append(f"{icon} @{uname}")
            invite_block = "\n".join(invite_lines)
            if len(rows) > 15:
                invite_block += f"\n…and {len(rows) - 15} more"
        else:
            invite_block = "No referrals yet — share your link below!"

        upgraded_count = sum(1 for r in rows if r.get("reward_status") in ("pending", "approved"))

        keyboard = [
            [InlineKeyboardButton("💸 Withdraw",            callback_data="withdraw_start")],
            [InlineKeyboardButton("📜 Withdrawal History",  callback_data="withdraw_history")],
        ] + back_main()

        await edit_menu(query,
            f"🎁 <b>Your Referral Program</b>\n\n"
            f"Share your link — when someone joins and later upgrades to "
            f"Pro, you earn <b>₦{REFERRAL_REWARD_NGN:,}</b> once the admin approves it.\n\n"
            f"🔗 <b>Your link:</b>\n<code>{link}</code>\n\n"
            f"👥 Total referred: <b>{len(rows)}</b>\n"
            f"💎 Upgraded to Pro: <b>{upgraded_count}</b>\n"
            f"💰 Available balance: <b>₦{bal['balance']:,}</b>\n"
            f"📈 Total earned (lifetime): <b>₦{bal['total_earned']:,}</b>\n\n"
            f"<b>Your invites:</b>\n{invite_block}\n\n"
            f"⏳ pending upgrade · 💵 commission pending · ✅ paid to balance\n\n"
            f"Minimum withdrawal: ₦{MIN_WITHDRAWAL_NGN:,}",
            InlineKeyboardMarkup(keyboard)
        )

    # ── 💸 Withdraw — start flow (collects bank details, then amount) ──
    elif data == "withdraw_start":
        uid = query.from_user.id
        bal = db.get_referral_balance(uid)

        if bal["balance"] < MIN_WITHDRAWAL_NGN:
            await edit_menu(query,
                f"💸 <b>Withdraw</b>\n\n"
                f"❌ Insufficient balance.\n\n"
                f"💰 Available balance: ₦{bal['balance']:,}\n"
                f"Minimum withdrawal: ₦{MIN_WITHDRAWAL_NGN:,}\n\n"
                f"Keep referring — once your balance reaches the minimum you can withdraw here.",
                InlineKeyboardMarkup(back_main())
            )
        else:
            _btn_state["action"] = "withdraw_acct_number"
            _btn_state["withdraw_temp"] = {}
            await edit_menu(query,
                f"💸 <b>Withdraw — Bank Details</b>\n\n"
                f"💰 Available balance: ₦{bal['balance']:,}\n\n"
                f"Step 1 of 4 — send your <b>bank account number</b>.",
                InlineKeyboardMarkup(back_main())
            )

    # ── 📜 Withdrawal history ──
    elif data == "withdraw_history":
        uid  = query.from_user.id
        rows = db.get_withdrawals_for(uid)
        status_icon = {"pending": "⏳ Pending", "completed": "✅ Completed", "rejected": "❌ Rejected"}
        if not rows:
            body = "No withdrawal requests yet."
        else:
            lines = []
            for w in rows[:15]:
                icon = status_icon.get(w.get("status"), "?")
                line = f"{icon} — ₦{w['amount']:,} — <code>{w['withdrawal_id']}</code> ({_esc(w.get('requested_at',''))})"
                if w.get("status") == "rejected" and w.get("reject_reason"):
                    line += f"\n   Reason: {_esc(w['reject_reason'])}"
                lines.append(line)
            body = "\n".join(lines)
            if len(rows) > 15:
                body += f"\n…and {len(rows) - 15} more"

        await edit_menu(query,
            f"📜 <b>Your Withdrawal History</b>\n\n{body}",
            InlineKeyboardMarkup(back_main())
        )

    # ── 🟢/🔴 Toggle Price Update ──
    elif data == "toggle_refresh":
        sess     = _s(tuser.id)
        slot_idx = sess.editing_slot
        s        = _ad_settings(sess, slot_idx)
        label    = _ad_slot_label(slot_idx)
        if _ad_running(sess, slot_idx):
            _set_ad_running(sess, slot_idx, False)
            _set_ad_task(sess, slot_idx, None)
            _set_ad_current_price(sess, slot_idx, Decimal("0"))
            await edit_menu(query,
                f"🔴 <b>{label} price update stopped.</b>\n\n" + ads_section_text(tuser.id),
                ads_section_keyboard(tuser.id)
            )
        else:
            ad_data = _ad_data_of(sess, slot_idx)
            if not ad_data or not s.get("ad_id"):
                hint_text = next_setup_hint(tuser.id) if slot_idx == -1 else "Set this ad's Ad ID and fetch its details first."
                await edit_menu(query,
                    f"❌ Not ready:\n\n_{hint_text}_",
                    InlineKeyboardMarkup(back_section("section_ads"))
                )
                return
            # ── Interval floor — defense in depth (already enforced when the
            # value was entered, but re-checked here in case of stale state) ──
            ok, err = validate_interval(s.get("interval", 2))
            if not ok:
                await edit_menu(query, err, InlineKeyboardMarkup(back_section("section_ads")))
                return
            # ── Float % range check — defense in depth (already enforced
            # when the value was entered, re-checked here in case of stale
            # state). Matching another active ad's % is allowed — the bot
            # keeps the actual posted PRICES apart live, in the loop
            # itself, not by restricting the % at this stage. ──
            if s.get("mode") == "floating":
                fok, ferr = validate_float_pct(
                    ad_data.get("currencyId","NGN"), ad_data.get("tokenId","USDT"),
                    s.get("float_pct", 0)
                )
                if not fok:
                    await edit_menu(query, ferr, InlineKeyboardMarkup(back_section("section_ads")))
                    return
            # ── Conflict guard: block auto-update while order/chat monitor is running ──
            # Running both simultaneously saturates the shared thread pool and event loop,
            # causing Telegram timeouts for ALL users. Users must choose one or the other.
            if sess.order_monitor_running or sess.chat_monitor_enabled:
                active = []
                if sess.order_monitor_running: active.append("Order Monitor")
                if sess.chat_monitor_enabled:  active.append("Chat Monitor")
                await edit_menu(query,
                    "⚠️ <b>Cannot start Auto-Update</b>\n\n"
                    f"<b>{' and '.join(active)}</b> is currently active.\n\n"
                    "Running Ad Auto-Update together with Order Monitor or Chat Monitor "
                    "overloads the bot and causes delays for all users.\n\n"
                    "Please stop your active monitors first, then start Auto-Update.",
                    InlineKeyboardMarkup(back_section("section_ads"))
                )
                return
            mode     = s.get("mode","fixed")
            interval = s.get("interval",2)
            _reset_ad_failures(sess, slot_idx)
            task = asyncio.create_task(auto_update_loop(context.bot, chat_id, slot_idx))
            _set_ad_task(sess, slot_idx, task)
            await edit_menu(query,
                f"🟢 <b>{label} price update started!</b>\n🔀 <code>{mode.upper()}</code> | ⏱ every <code>{interval}</code> min\n\n"
                + ads_section_text(tuser.id),
                ads_section_keyboard(tuser.id)
            )

    # ── 🔀 Multi-ad: switch which ad the menu is editing ──
    elif data in ("edit_ad_1", "edit_ad_2", "edit_ad_3"):
        sess = _s(tuser.id)
        target = int(data[-1]) - 2   # "edit_ad_1" -> -1, "edit_ad_2" -> 0, "edit_ad_3" -> 1
        if target == -1 or target < len(sess.extra_ad_slots):
            sess.editing_slot = target
        await edit_menu(query, ads_section_text(tuser.id), ads_section_keyboard(tuser.id))

    # ── ➕ Add another ad slot (up to MAX_ADS_PER_USER) ──
    elif data == "add_ad_slot":
        sess = _s(tuser.id)
        if sess.total_ad_slots() >= MAX_ADS_PER_USER:
            await edit_menu(query, f"❌ Maximum {MAX_ADS_PER_USER} ads per account.", InlineKeyboardMarkup(back_section("section_ads")))
            return
        sess.add_ad_slot()
        sess.editing_slot = len(sess.extra_ad_slots) - 1   # jump straight to editing the new one
        await edit_menu(query,
            f"✅ <b>{_ad_slot_label(sess.editing_slot)} added!</b>\n\n"
            "Set its Ad ID, fetch its details, then choose a mode — "
            "it shares this account's Bybit API key and UID with your other ads.\n\n"
            + ads_section_text(tuser.id),
            ads_section_keyboard(tuser.id)
        )

    # ── 🗑 Remove the ad slot currently being edited ──
    elif data == "remove_ad_slot":
        sess = _s(tuser.id)
        slot_idx = sess.editing_slot
        if slot_idx == -1:
            await edit_menu(query, "❌ Ad 1 can't be removed — stop it instead.", InlineKeyboardMarkup(back_section("section_ads")))
            return
        label = _ad_slot_label(slot_idx)
        sess.remove_ad_slot(slot_idx)
        sess.editing_slot = -1   # back to Ad 1
        await edit_menu(query,
            f"🗑 <b>{label} removed.</b>\n\n" + ads_section_text(tuser.id),
            ads_section_keyboard(tuser.id)
        )

    # ── 🗂 All-ads dashboard ──
    elif data == "ads_dashboard":
        await edit_menu(query, ads_dashboard_text(tuser.id), ads_dashboard_keyboard(tuser.id))

    # ── ⏹ Stop one specific ad from the dashboard ──
    elif data in ("stop_ad_1", "stop_ad_2", "stop_ad_3"):
        sess = _s(tuser.id)
        target = int(data[-1]) - 2
        sess.stop_ad_slot(target)
        _set_ad_current_price(sess, target, Decimal("0"))
        await edit_menu(query,
            f"🔴 <b>{_ad_slot_label(target)} stopped.</b>\n\n" + ads_dashboard_text(tuser.id),
            ads_dashboard_keyboard(tuser.id)
        )

    # ── ⏹ Stop all ads in one tap ──
    elif data == "stop_all_ads":
        sess = _s(tuser.id)
        for i in range(-1, len(sess.extra_ad_slots)):
            sess.stop_ad_slot(i)
            _set_ad_current_price(sess, i, Decimal("0"))
        await edit_menu(query,
            "🔴 <b>All ads stopped.</b>\n\n" + ads_dashboard_text(tuser.id),
            ads_dashboard_keyboard(tuser.id)
        )

    # ── ✅ Mark as Paid ──
    elif data.startswith("pay_") and not data.startswith("paywarn_"):
        order_id = data[4:]
        # ── Duplicate action guard ──
        if _is_order_finalized(chat_id, order_id):
            await query.answer("✅ Already processed — no action needed.", show_alert=True)
            return
        async with _get_order_lock(chat_id, order_id):
            if _is_order_finalized(chat_id, order_id):
                await query.answer("✅ Already processed.", show_alert=True)
                return
            await context.bot.send_message(chat_id=chat_id,
                text=f"⏳ Marking order <code>{_esc(order_id)}</code> as paid...", parse_mode="HTML")
            det = await asyncio.get_event_loop().run_in_executor(None, partial(get_order_detail, order_id, creds=get_user_creds(tuser.id)))
            if det.get("retCode",-1) != 0:
                await context.bot.send_message(chat_id=chat_id,
                    text=f"❌ Could not fetch order\n<code>{_esc(det.get('retMsg',''))}</code>", parse_mode="HTML")
                return
            order_detail = det.get("result",{})
            pay_term     = order_detail.get("confirmedPayTerm",{}) or {}
            if not pay_term:
                terms    = order_detail.get("paymentTermList",[])
                pay_term = terms[0] if terms else {}
            payment_type = str(pay_term.get("paymentType",""))
            payment_id   = str(pay_term.get("id",""))
            if not payment_type or not payment_id:
                await context.bot.send_message(chat_id=chat_id,
                    text="❌ No payment info found. Buyer may not have selected payment yet.", parse_mode="HTML")
                return
            result = await asyncio.get_event_loop().run_in_executor(
                None, partial(mark_order_paid, order_id, payment_type, payment_id, creds=get_user_creds(chat_id))
            )
            if result.get("retCode", result.get("ret_code",-1)) == 0:
                _s(tuser.id).paid_order_ids.add(order_id)
                _track_buy_volume(tuser.id, order_id, order_detail)
                # ── Edit original message: remove buttons, show status badge ──
                await _update_order_message_final(context.bot, chat_id, order_id, "Completed", "completed")
                await context.bot.send_message(chat_id=chat_id,
                    text=f"✅ <b>Order marked as paid!</b>\n<code>{_esc(order_id)}</code>", parse_mode="HTML")
            else:
                await context.bot.send_message(chat_id=chat_id,
                    text=f"❌ Failed\n<code>{_esc(result.get('retMsg',''))}</code>", parse_mode="HTML")

    # ── ⚠️ Mark Paid + Warn ──
    elif data.startswith("paywarn_"):
        order_id = data[8:]
        # ── Duplicate action guard ──
        if _is_order_finalized(chat_id, order_id):
            await query.answer("✅ Already processed — no action needed.", show_alert=True)
            return
        async with _get_order_lock(chat_id, order_id):
            if _is_order_finalized(chat_id, order_id):
                await query.answer("✅ Already processed.", show_alert=True)
                return
            await context.bot.send_message(chat_id=chat_id,
                text=f"⏳ Marking paid + sending warning for <code>{_esc(order_id)}</code>...", parse_mode="HTML")
            det = await asyncio.get_event_loop().run_in_executor(None, partial(get_order_detail, order_id, creds=get_user_creds(tuser.id)))
            if det.get("retCode",-1) != 0:
                await context.bot.send_message(chat_id=chat_id,
                    text=f"❌ <code>{_esc(det.get('retMsg',''))}</code>", parse_mode="HTML")
                return
            order_detail = det.get("result",{})
            pay_term     = order_detail.get("confirmedPayTerm",{}) or {}
            if not pay_term:
                terms    = order_detail.get("paymentTermList",[])
                pay_term = terms[0] if terms else {}
            payment_type = str(pay_term.get("paymentType",""))
            payment_id   = str(pay_term.get("id",""))
            if not payment_type or not payment_id:
                await context.bot.send_message(chat_id=chat_id,
                    text="❌ No payment info found.", parse_mode="HTML")
                return
            pr = await asyncio.get_event_loop().run_in_executor(
                None, partial(mark_order_paid, order_id, payment_type, payment_id, creds=get_user_creds(chat_id))
            )
            if pr.get("retCode", pr.get("ret_code",-1)) == 0:
                _s(tuser.id).paid_order_ids.add(order_id)
                _track_buy_volume(tuser.id, order_id, order_detail)
                mr = await asyncio.get_event_loop().run_in_executor(
                    None, partial(send_chat_message, order_id, SELLER_WARN_MSG,
                                      creds=get_user_creds(chat_id))
                )
                warn_ok = mr.get("retCode", mr.get("ret_code",-1)) == 0
                warn_label = "✅ Warning sent to seller" if warn_ok else f"⚠️ Warning failed: <code>{_esc(mr.get('retMsg',''))}</code>"
                # ── Edit original message: remove buttons, show status badge ──
                final_state = "warned" if warn_ok else "completed"
                await _update_order_message_final(context.bot, chat_id, order_id, "Warning Sent", final_state)
                await context.bot.send_message(chat_id=chat_id,
                    text=f"✅ <b>Order paid!</b> <code>{_esc(order_id)}</code>\n{warn_label}", parse_mode="HTML")
            else:
                await context.bot.send_message(chat_id=chat_id,
                    text=f"❌ Failed\n<code>{_esc(pr.get('retMsg',''))}</code>", parse_mode="HTML")

    # ── 🚫 Seller Cancel Review — Accept ──
    elif data.startswith("sc_accept_"):
        order_id = data[len("sc_accept_"):]
        if order_id not in _s(tuser.id).pending_cancel_reviews:
            await query.answer("This cancel request has already been handled.", show_alert=True)
            return
        await query.answer("Processing...")
        # Remove buttons from the notification message
        try:
            await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([]))
        except Exception:
            pass
        await _handle_cancel_review(context.bot, chat_id, order_id, "PASS")

    # ── 🚫 Seller Cancel Review — Reject (show reason buttons) ──
    elif data.startswith("sc_reject_"):
        order_id = data[len("sc_reject_"):]
        if order_id not in _s(tuser.id).pending_cancel_reviews:
            await query.answer("This cancel request has already been handled.", show_alert=True)
            return
        await query.answer()
        # Replace the Accept/Reject buttons with reason selection buttons
        try:
            await query.edit_message_reply_markup(
                reply_markup=_cancel_reject_reason_buttons(order_id)
            )
        except Exception:
            await context.bot.send_message(
                chat_id=chat_id,
                text=(
                    f"Select your reason for rejecting the cancellation of order "
                    f"<code>{_esc(order_id)}</code>:"
                ),
                reply_markup=_cancel_reject_reason_buttons(order_id),
                parse_mode="HTML"
            )

    # ── 🚫 Seller Cancel Review — Reason chosen ──
    elif data.startswith("sc_reason_"):
        # Format: sc_reason_{order_id}_{reason_key}
        rest     = data[len("sc_reason_"):]
        # reason_key is always one of the 3 fixed strings — split from the right
        parts    = rest.rsplit("_", 1)
        if len(parts) != 2:
            await query.answer("Invalid selection.", show_alert=True)
            return
        order_id, reason_key = parts[0], parts[1]
        if order_id not in _s(tuser.id).pending_cancel_reviews:
            await query.answer("This cancel request has already been handled.", show_alert=True)
            return
        await query.answer("Submitting rejection...")
        # Remove reason buttons
        try:
            await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([]))
        except Exception:
            pass
        await _handle_cancel_review(context.bot, chat_id, order_id, "REJECT", reason_key)

    # ── 🔕 Order Status Badge (noop — already finalized) ──
    elif data.startswith("order_status_noop_"):
        await query.answer("This order has already been processed.", show_alert=False)
        return

    # ── 🪙 Release Coin ──
    elif data.startswith("release_"):
        order_id = data[8:]
        await context.bot.send_message(chat_id=chat_id,
            text=f"⏳ Releasing coins for order <code>{_esc(order_id)}</code>...", parse_mode="HTML")
        result   = await asyncio.get_event_loop().run_in_executor(None, partial(release_assets, order_id, creds=get_user_creds(tuser.id)))
        ret_code = result.get("retCode", result.get("ret_code", -1))
        ret_msg  = result.get("retMsg",  result.get("ret_msg",  ""))
        if ret_code == 0:
            _s(tuser.id).released_ids.add(order_id)
            # Remove the release button from the original message
            try:
                await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([]))
            except Exception:
                pass
            await context.bot.send_message(chat_id=chat_id,
                text=f"🪙 <b>Coins released!</b>\n\nOrder: <code>{order_id}</code>\nBuyer has received their coins. ✅",
                parse_mode="HTML")
        else:
            await context.bot.send_message(chat_id=chat_id,
                text=f"❌ <b>Release failed</b>\nCode: <code>{ret_code}</code>\nMessage: <code>{ret_msg}</code>",
                parse_mode="HTML")


# ─────────────────────────────────────────
# 📝 TEXT INPUT HANDLER
# ─────────────────────────────────────────
async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):

    tuser = update.effective_user
    uid   = tuser.id
    _get_or_register_user(tuser)  # ensure user exists in DB

    text = update.message.text.strip()

    # ── Per-user isolated state ──
    # Admin uses the global user_state dict.
    # Non-admin users get their own state via context.user_data so their
    # API key inputs are isolated and don't collide with the admin's state.
    if is_admin(uid):
        _state = user_state
    else:
        if "state" not in context.user_data:
            context.user_data["state"] = {}
        _state = context.user_data["state"]

    action = _state.get("action")
    prev   = _state.get("prev_section", "main_menu")

    async def reply_with_back(msg: str):
        """Reply with success message + back-to-previous button."""
        await update.message.reply_text(msg, parse_mode="HTML", reply_markup=back_prev(prev))

    # ── Bybit API — slot-aware (Account 1 or Account 2) ──
    if action in ("api_bybit_key", "api_bybit_key_1", "api_bybit_key_2"):
        slot = "2" if action == "api_bybit_key_2" else "1"
        val  = text.strip()
        next_action = f"api_bybit_secret_{slot}"
        _state["action"]              = next_action
        _state["prev_section"]        = "section_apis"
        _state["_api_bybit_key_temp"] = val
        _state["_api_bybit_slot"]     = slot
        await update.message.reply_text(
            f"✅ Account {slot} API Key received.\n\n"
            f"Step 2 of 2: Send your Bybit Account {slot} <b>API Secret</b>.",
            parse_mode="HTML"
        )
        return

    elif action in ("api_bybit_secret", "api_bybit_secret_1", "api_bybit_secret_2"):
        uid      = update.effective_user.id
        slot     = _state.pop("_api_bybit_slot", "1")
        key_temp = _state.pop("_api_bybit_key_temp", "")
        db.save_api(uid, f"bybit_key_{slot}",    key_temp)
        db.save_api(uid, f"bybit_secret_{slot}", text.strip())
        # Credentials are loaded per-call via get_user_creds() — no global mutation needed.
        _state["action"] = None
        _save_settings(uid)
        await update.message.reply_text(
            f"✅ <b>Bybit Account {slot} API saved!</b>\n\n"
            f"Key and Secret stored securely.\n"
            f"The bot uses Account {slot} keys when Account {slot} is active.",
            parse_mode="HTML",
            reply_markup=back_prev("section_apis")
        )
        return

    # ── Flutterwave 3-step credential input ──
    # Only FLW_PUBLIC_KEY, FLW_SECRET_HASH, FLW_SECRET_KEY are required.
    # FLW_CLIENT_ID and FLW_CLIENT_SECRET are NOT used by the transfer/webhook system.

    elif action == "api_flw_public_key":
        val = text.strip()
        if not val:
            await update.message.reply_text(
                "❌ FLW_PUBLIC_KEY cannot be empty. Please send the value.",
                parse_mode="HTML"
            )
            return
        _state["action"]                   = "api_flw_secret_hash"
        _state["_api_flw_public_key_temp"] = val
        await update.message.reply_text(
            "✅ <b>FLW_PUBLIC_KEY received.</b>\n\n"
            "<b>Step 2 of 3:</b> Send your <b>FLW_SECRET_HASH</b>\n"
            "<i>(Webhook secret hash — set on Flutterwave dashboard → Webhooks)</i>",
            parse_mode="HTML"
        )
        return

    elif action == "api_flw_secret_hash":
        val = text.strip()
        if not val:
            await update.message.reply_text(
                "❌ FLW_SECRET_HASH cannot be empty. Please send the value.",
                parse_mode="HTML"
            )
            return
        _state["action"]                    = "api_flw_secret_key"
        _state["_api_flw_secret_hash_temp"] = val
        await update.message.reply_text(
            "✅ <b>FLW_SECRET_HASH received.</b>\n\n"
            "<b>Step 3 of 3:</b> Send your <b>FLW_SECRET_KEY</b>\n"
            "<i>(Live secret key — starts with FLWSECK_ — used for transfers and payouts)</i>",
            parse_mode="HTML"
        )
        return

    elif action == "api_flw_secret_key":
        uid = update.effective_user.id
        val = text.strip()
        if not val:
            await update.message.reply_text(
                "❌ FLW_SECRET_KEY cannot be empty. Please send the value.",
                parse_mode="HTML"
            )
            return
        public_key  = _state.pop("_api_flw_public_key_temp",  "")
        secret_hash = _state.pop("_api_flw_secret_hash_temp", "")
        secret_key  = val  # primary key used for all API auth and transfers

        db.save_api(uid, "flw_public_key",  public_key)
        db.save_api(uid, "flw_secret_hash", secret_hash)
        db.save_api(uid, "flw_secret_key",  secret_key)

        _state["action"] = None
        _save_settings(uid)
        await update.message.reply_text(
            "✅ <b>Flutterwave API saved!</b>\n\n"
            "All 3 credentials stored securely per your account:\n"
            "  ✔ FLW_PUBLIC_KEY\n"
            "  ✔ FLW_SECRET_HASH\n"
            "  ✔ FLW_SECRET_KEY\n\n"
            "Use /pingflutterwave to test the connection.",
            parse_mode="HTML",
            reply_markup=back_prev("section_apis")
        )
        return

    elif action == "api_paga_api_key":
        val = text.strip()
        if not val:
            await update.message.reply_text(
                "❌ PAGA_API_KEY cannot be empty. Please send the value.",
                parse_mode="HTML"
            )
            return
        _state["action"]                 = "api_paga_credential"
        _state["_api_paga_api_key_temp"] = val
        await update.message.reply_text(
            "✅ <b>PAGA_API_KEY received.</b>\n\n"
            "<b>Step 2 of 3:</b> Send your <b>PAGA_CREDENTIAL</b>\n"
            "<i>(Live Primary Secret Key from Paga dashboard)</i>",
            parse_mode="HTML"
        )
        return

    elif action == "api_paga_credential":
        val = text.strip()
        if not val:
            await update.message.reply_text(
                "❌ PAGA_CREDENTIAL cannot be empty. Please send the value.",
                parse_mode="HTML"
            )
            return
        _state["action"]                     = "api_paga_principal"
        _state["_api_paga_credential_temp"]  = val
        await update.message.reply_text(
            "✅ <b>PAGA_CREDENTIAL received.</b>\n\n"
            "<b>Step 3 of 3:</b> Send your <b>PAGA_PRINCIPAL</b>\n"
            "<i>(Your Public Key / Principal from Paga dashboard)</i>",
            parse_mode="HTML"
        )
        return

    elif action == "api_paga_principal":
        val = text.strip()
        if not val:
            await update.message.reply_text(
                "❌ PAGA_PRINCIPAL cannot be empty. Please send the value.",
                parse_mode="HTML"
            )
            return
        uid = update.effective_user.id
        db.save_api(uid, "paga_api_key",    _state.pop("_api_paga_api_key_temp", ""))
        db.save_api(uid, "paga_credential", _state.pop("_api_paga_credential_temp", ""))
        db.save_api(uid, "paga_principal",  val)
        _state["action"] = None
        _save_settings(uid)
        await update.message.reply_text(
            "✅ <b>Paga API saved!</b>\n\n"
            "All 3 credentials stored securely per your account:\n"
            "  ✔ PAGA_API_KEY\n"
            "  ✔ PAGA_CREDENTIAL\n"
            "  ✔ PAGA_PRINCIPAL\n\n"
            "Use /pingpaga to test the connection.",
            parse_mode="HTML",
            reply_markup=back_prev("section_apis")
        )
        return

    if action == "manage_ad_id":
        _s(uid).settings["manage_ad_id"] = text.strip()
        _s(uid).settings.pop("manage_ad_data", None)   # clear old manage ad data
        _state["action"] = None
        auto_id = _s(uid).settings.get("ad_id", "not set")
        await reply_with_back(
            f"✅ <b>Manage Ad ID saved!</b>\n\n"
            f"Manage Ad ID: <code>{text.strip()}</code>\n"
            f"Auto-Update Ad ID: <code>{auto_id}</code> (unchanged)\n\n"
            f"Now tap <b>📢 Post/Remove Ad</b> → <b>📋 Fetch Manage Ad</b> to load its details."
        )
        return

    elif action == "chat_reply":
        state    = _s(uid).reply_state.pop(uid, {})
        order_id = state.get("order_id", "")
        nick     = state.get("nick", "counterparty")
        _state["action"] = None
        if not order_id:
            await update.message.reply_text("❌ No active reply state. Tap Reply on a message first.")
            return
        result = await asyncio.get_event_loop().run_in_executor(
            None, partial(send_chat_message, order_id, text, creds=get_user_creds(uid))
        )
        rc = result.get("retCode", result.get("ret_code", -1))
        if rc == 0:
            await update.message.reply_text(
                f"✅ <b>Message sent to {nick}</b>\n\nOrder: <code>{order_id}</code>\n💬 <i>{text[:200]}</i>",
                parse_mode="HTML"
            )
            logger.info(f"[ChatReply] Sent to order {order_id}: {text[:100]}")
        else:
            await update.message.reply_text(
                f"❌ Failed to send message\n<code>{result.get('retMsg', result.get('ret_msg',''))}</code>",
                parse_mode="HTML"
            )
        return

    elif action == "ad_id":
        sess = _s(uid)
        slot_idx = sess.editing_slot
        if slot_idx == -1:
            # Save under BOTH the slot-keyed key and the generic fallback key
            slot_str = _get_user_slot_str(uid)
            sess.settings[f"ad_id_{slot_str}"] = text.strip()
            sess.settings["ad_id"]              = text.strip()
            sess.ad_data.clear()
            _save_settings(uid)   # persisted — Ad 1 lives on disk per account slot
            logger.info(f"[AdID] Saved ad_id for user={uid} slot={slot_str} ad_id={text.strip()!r}")
            label = f"Account {slot_str}"
        else:
            # Ads 2/3 are ephemeral (in-memory only), like the rest of the
            # per-user session state — no disk persistence needed here.
            s = _ad_settings(sess, slot_idx)
            s["ad_id"] = text.strip()
            _ad_data_of(sess, slot_idx).clear()
            logger.info(f"[AdID] Saved ad_id for user={uid} {_ad_slot_label(slot_idx)} ad_id={text.strip()!r}")
            label = _ad_slot_label(slot_idx)
        _state["action"] = None
        hint = next_setup_hint(uid) if slot_idx == -1 else "Now use Fetch Ad Details for this ad."
        await update.message.reply_text(
            f"✅ <b>Ad ID saved for {label}!</b>\n\n"
            f"<code>{_esc(text.strip())}</code>\n\n"
            f"<i>{_esc(hint)}</i>",
            parse_mode="HTML",
            reply_markup=back_prev("section_ads")
        )

    elif action == "bybit_uid":
        # Save under the slot-keyed key ONLY.
        # The generic "bybit_uid" key is synced from the ACTIVE slot's value so it
        # always reflects the current slot without leaking into other slots.
        slot_str = _get_user_slot_str(uid)
        _s(uid).settings[f"bybit_uid_{slot_str}"] = text.strip()
        # Keep generic key in sync with current slot (used by chat monitor etc.)
        _s(uid).settings["bybit_uid"] = text.strip()
        _state["action"] = None
        # Persist to disk immediately so it survives /start, restarts, slot switches
        _save_settings(uid)
        logger.info(f"[UID] Saved bybit_uid for user={uid} slot={slot_str} uid_value={text.strip()!r}")
        hint = next_setup_hint(uid)
        # Return to section_ads with back button pointing to the AD PRICE BOT menu
        try:
            await update.message.reply_text(
                f"✅ <b>UID saved for Account {slot_str}!</b>\n\n"
                f"<code>{_esc(text.strip())}</code>\n\n"
                f"<i>{_esc(hint)}</i>",
                parse_mode="HTML",
                reply_markup=back_prev("section_ads")
            )
        except Exception as _uid_reply_err:
            logger.warning(f"[UID] Reply failed: {_uid_reply_err}")
            await update.message.reply_text(
                f"✅ UID saved: <code>{_esc(text.strip())}</code>",
                parse_mode="HTML"
            )

    elif action == "increment":
        try:
            val = Decimal(text)
            if val <= 0: raise ValueError
            sess = _s(uid)
            slot_idx = sess.editing_slot
            s = _ad_settings(sess, slot_idx)
            s["increment"] = text
            if slot_idx == -1:
                slot_str = _get_user_slot_str(uid)
                sess.settings[f"increment_{slot_str}"] = text
                _save_settings(uid)
            _state["action"] = None
            hint = next_setup_hint(uid) if slot_idx == -1 else ""
            await reply_with_back(f"✅ <b>{_ad_slot_label(slot_idx)} increment saved!</b>\n\n<code>+{_esc(text)}</code> per cycle\n\n<i>{_esc(hint)}</i>")
        except Exception:
            await update.message.reply_text("❌ Send a positive number like `0.05`", parse_mode="HTML")

    elif action == "float_pct":
        sess = _s(uid)
        slot_idx = sess.editing_slot
        ad_data = _ad_data_of(sess, slot_idx)
        token    = ad_data.get("tokenId","USDT").upper()
        currency = ad_data.get("currencyId","NGN").upper()
        ok, err = validate_float_pct(currency, token, text)
        if not ok:
            await update.message.reply_text(err, parse_mode="HTML")
            return
        s = _ad_settings(sess, slot_idx)
        s["float_pct"] = text
        if slot_idx == -1:
            slot_str = _get_user_slot_str(uid)
            sess.settings[f"float_pct_{slot_str}"] = text
            _save_settings(uid)
        _state["action"] = None
        hint = next_setup_hint(uid) if slot_idx == -1 else ""
        await reply_with_back(
            f"✅ <b>{_ad_slot_label(slot_idx)} float % saved!</b>\n\n<code>{text}%</code> for <code>{token}/{currency}</code>\n\n"
            f"_{hint}_"
        )

    elif action == "ngn_usdt_ref":
        try:
            val = float(text)
            if val <= 0: raise ValueError
            sess = _s(uid)
            sess.sync_shared_ref(text)   # applies to every ad slot at once
            if sess.editing_slot == -1:
                slot_str = _get_user_slot_str(uid)
                sess.settings[f"local_usdt_ref_{slot_str}"] = text
                _save_settings(uid)
            _scur = _ad_data_of(sess, sess.editing_slot).get("currencyId","NGN").upper()
            _state["action"] = None
            shared_note = " (applies to all your ads)" if sess.total_ad_slots() > 1 else ""
            await reply_with_back(f"✅ <b>{_esc(_scur)}/USDT ref saved!{shared_note}</b>\n\n<code>{_esc(text)}</code>\n\n<i>{_esc(next_setup_hint(uid) if sess.editing_slot == -1 else '')}</i>")
        except Exception:
            await update.message.reply_text("❌ Send a number like `1580`", parse_mode="HTML")

    elif action == "interval":
        ok, err = validate_interval(text)
        if not ok:
            await update.message.reply_text(err, parse_mode="HTML")
            return
        val = int(text)
        sess = _s(uid)
        slot_idx = sess.editing_slot
        s = _ad_settings(sess, slot_idx)
        s["interval"] = val
        if slot_idx == -1:
            slot_str = _get_user_slot_str(uid)
            sess.settings[f"interval_{slot_str}"] = val
            _save_settings(uid)
        _state["action"] = None
        hint = next_setup_hint(uid) if slot_idx == -1 else ""
        await reply_with_back(f"✅ <b>{_ad_slot_label(slot_idx)} interval saved!</b>\n\nEvery <code>{_esc(str(val))}</code> min\n\n<i>{_esc(hint)}</i>")

    elif action == "sender_name":
        _s(uid).settings["sender_name"] = text.strip()
        _state["action"] = None
        await reply_with_back(
            f"✅ <b>Sender name saved!</b>\n\n<code>{text.strip()}</code>\n\n"
            f"FLW narration: <code>{text.strip()} payment to [receiver]</code>"
        )

    elif action == "sell_custom_msg":
        _s(uid).sell_custom_msg = text
        _state["action"] = None
        preview = text[:80] + "..." if len(text) > 80 else text
        await reply_with_back(
            f"✅ <b>Sell message saved!</b>\n\nPreview: <i>{preview}</i>\n\n"
            f"Will be sent <code>{_s(uid).sell_msg_count}x</code> per sell order."
        )

    elif action == "sell_msg_count":
        try:
            val = int(text)
            if val < 1 or val > 5: raise ValueError
            _s(uid).sell_msg_count = val
            _state["action"] = None
            await reply_with_back(f"✅ <b>Message count saved!</b>\n\nWill send <code>{_esc(str(val))}x</code> per sell order.")
        except Exception:
            await update.message.reply_text("❌ Send a number between `1` and `5`", parse_mode="HTML")

    elif action == "post_ad_qty":
        try:
            val = Decimal(text)
            if val <= 0: raise ValueError
            _s(uid).settings["post_ad_qty"] = text
            _state["action"] = None
            await reply_with_back(
                f"✅ <b>Custom quantity set:</b> <code>{text}</code>\n\n"
                "Now tap *📢 Post Ad (clone)* → *Confirm Post* to post the ad."
            )
        except Exception:
            await update.message.reply_text("❌ Send a positive number like `5000`", parse_mode="HTML")

    elif action == "bp_custom_threshold":
        try:
            val = int(text)
            if val < 1: raise ValueError
            _s(uid).buyer_protection_mins = val
            _state["action"] = None
            await reply_with_back(
                f"✅ <b>Buyer Protection threshold set!</b>\n\n"
                f"Threshold: <code>{val} min</code>\n\n"
                f"Status: {'✅ ON' if _s(uid).buyer_protection_on else '❌ OFF (tap toggle to enable)'}"
            )
        except Exception:
            await update.message.reply_text("❌ Send a whole number like `25`", parse_mode="HTML")

    # ── 💸 Withdraw flow — Step 1/4: bank account number ──
    elif action == "withdraw_acct_number":
        if not text:
            await update.message.reply_text("❌ Please send your bank account number.")
        else:
            _state.setdefault("withdraw_temp", {})["account_number"] = text
            _state["action"] = "withdraw_acct_name"
            await update.message.reply_text(
                "Step 2 of 4 — send the <b>account name</b> (the name on the bank account).",
                parse_mode="HTML"
            )

    # ── 💸 Withdraw flow — Step 2/4: account name ──
    elif action == "withdraw_acct_name":
        if not text:
            await update.message.reply_text("❌ Please send the account name.")
        else:
            _state.setdefault("withdraw_temp", {})["account_name"] = text
            _state["action"] = "withdraw_bank_name"
            await update.message.reply_text(
                "Step 3 of 4 — send the <b>bank name</b> (e.g. GTBank, Access Bank, OPay).",
                parse_mode="HTML"
            )

    # ── 💸 Withdraw flow — Step 3/4: bank name ──
    elif action == "withdraw_bank_name":
        if not text:
            await update.message.reply_text("❌ Please send the bank name.")
        else:
            _state.setdefault("withdraw_temp", {})["bank_name"] = text
            _state["action"] = "withdraw_amount"
            bal = db.get_referral_balance(uid)
            await update.message.reply_text(
                f"Step 4 of 4 — send the <b>amount</b> you want to withdraw (numbers only).\n\n"
                f"💰 Available balance: ₦{bal['balance']:,}\n"
                f"Minimum withdrawal: ₦{MIN_WITHDRAWAL_NGN:,}",
                parse_mode="HTML"
            )

    # ── 💸 Withdraw flow — Step 4/4: amount + submit ──
    elif action == "withdraw_amount":
        try:
            amount = int(text.replace(",", "").strip())
            if amount <= 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text("❌ Send a whole number, e.g. `5000`", parse_mode="HTML")
        else:
            bank = _state.get("withdraw_temp", {})
            result = db.create_withdrawal_request(uid, amount, bank, min_amount=MIN_WITHDRAWAL_NGN)

            if not result["ok"]:
                if result["reason"] == "below_minimum":
                    await update.message.reply_text(
                        f"❌ Minimum withdrawal is ₦{MIN_WITHDRAWAL_NGN:,}. You sent ₦{amount:,}.\n\n"
                        f"Send a new amount, or tap 🏠 Main Menu to cancel.",
                        parse_mode="HTML"
                    )
                elif result["reason"] == "insufficient":
                    await update.message.reply_text(
                        f"❌ <b>Insufficient balance.</b>\n\n"
                        f"💰 Available balance: ₦{result['balance']:,}\n"
                        f"You requested: ₦{amount:,}\n\n"
                        f"Send a smaller amount, or tap 🏠 Main Menu to cancel.",
                        parse_mode="HTML"
                    )
                else:
                    await update.message.reply_text("❌ Could not process withdrawal. Please try again.")
                # Stay in the same step so the user can retry with a different amount
            else:
                _state["action"] = None
                _state.pop("withdraw_temp", None)
                wid = result["withdrawal_id"]

                await update.message.reply_text(
                    f"✅ <b>Withdrawal request submitted!</b>\n\n"
                    f"🆔 Reference: <code>{wid}</code>\n"
                    f"💵 Amount: ₦{amount:,}\n"
                    f"🏦 {_esc(bank.get('bank_name',''))} — {_esc(bank.get('account_number',''))} ({_esc(bank.get('account_name',''))})\n\n"
                    f"Status: ⏳ <b>Pending processing</b>\n"
                    f"💰 New balance: ₦{result['new_balance']:,}\n\n"
                    f"You'll be notified once the admin approves or rejects it. "
                    f"Check 🎁 Referrals -> 📜 Withdrawal History any time.",
                    parse_mode="HTML"
                )

                # Notify admins immediately with full details
                for admin_chat_id in _get_admin_chat_ids():
                    try:
                        await context.bot.send_message(
                            chat_id=admin_chat_id,
                            text=(
                                f"💸 <b>New Withdrawal Request!</b>\n\n"
                                f"🆔 Reference: <code>{wid}</code>\n"
                                f"👤 User: @{_esc(tuser.username or '?')} (<code>{uid}</code>)\n"
                                f"💵 Amount: ₦{amount:,}\n"
                                f"🏦 Bank: {_esc(bank.get('bank_name',''))}\n"
                                f"🔢 Account Number: {_esc(bank.get('account_number',''))}\n"
                                f"🧾 Account Name: {_esc(bank.get('account_name',''))}\n\n"
                                f"✅ Approve: <code>/approvewithdraw {wid}</code>\n"
                                f"❌ Reject: <code>/rejectwithdraw {wid} reason</code>"
                            ),
                            parse_mode="HTML"
                        )
                    except Exception as _notify_err:
                        logger.error(f"[Withdraw] Could not notify admin {admin_chat_id}: {_notify_err}")

    # ── 💬 Contact Support — user's message received ──
    elif action == "support_message":
        _state["action"] = None
        uname_label = tuser.username or tuser.full_name or str(uid)
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Reply to this user", callback_data=f"support_reply_{uid}")]])
        admin_msg = (
            f"📩 <b>New Support Message</b>\n\n"
            f"👤 User: @{_esc(uname_label)}\n"
            f"ID: <code>{uid}</code>\n\n"
            f"💬 Message:\n{_esc(text)}"
        )
        sent_to_any = False
        for admin_id in list(_admin_chat_ids):
            try:
                await context.bot.send_message(chat_id=admin_id, text=admin_msg, parse_mode="HTML", reply_markup=keyboard)
                sent_to_any = True
            except Exception as e:
                logger.error(f"[Support] Could not reach admin {admin_id}: {e}")
        if sent_to_any:
            await update.message.reply_text("✅ Your message has been sent to support. They'll get back to you here soon.")
        else:
            await update.message.reply_text("❌ Could not reach support right now — please try again shortly.")

    # ── 💬 Admin's typed reply to a support message ──
    elif action == "support_reply_awaiting":
        target_uid = _state.get("support_reply_target")
        _state["action"] = None
        _state["support_reply_target"] = None
        if not target_uid:
            await update.message.reply_text("❌ No pending reply target — tap Reply on a support message first.")
            return
        try:
            await context.bot.send_message(
                chat_id=target_uid,
                text=f"💬 <b>Reply from Support:</b>\n\n{_esc(text)}",
                parse_mode="HTML"
            )
            await update.message.reply_text(f"✅ Reply sent to user <code>{target_uid}</code>.", parse_mode="HTML")
        except Exception as e:
            logger.error(f"[Support] Could not deliver reply to {target_uid}: {e}")
            await update.message.reply_text(f"❌ Could not deliver reply — the user may have blocked the bot.")

    # ── ⬆️ Upgrade request — contact details ──
    elif action == "upgrade_contact":
        ok, contact, err = validate_contact(text)
        if not ok:
            await update.message.reply_text(err, parse_mode="HTML")
            return
        _state["action"] = None
        await update.message.reply_text(
            "⏳ <b>Upgrade Request Sent!</b>\n\n"
            "The admin has been notified and will review shortly.\n"
            "You will receive a message once approved.",
            parse_mode="HTML"
        )
        await _submit_upgrade_request(context.bot, uid, tuser.username or "", tuser.full_name or "", contact)

    # ── 📢 Broadcast — text content received ──
    elif action == "broadcast_awaiting_content":
        if text.strip().lower() == "cancel":
            _state["action"] = None
            await update.message.reply_text("❌ Broadcast cancelled.")
            return
        _state["broadcast_draft"] = {"type": "text", "file_id": None, "text": text}
        _state["action"] = None
        await _show_broadcast_preview(update, context)

    # ── 🎬 Video Downloader — link received ──
    elif action == "video_link":
        if not mediadl.looks_like_url(text):
            await update.message.reply_text(
                "❌ That doesn't look like a link. Send a direct video link starting with http:// or https://"
            )
            return
        _state["action"] = None

        status_msg = await update.message.reply_text("🔎 Checking link...")
        probe = await asyncio.get_event_loop().run_in_executor(None, mediadl.probe_video, text)
        if not probe["ok"]:
            await status_msg.edit_text(f"❌ {probe['reason']}")
            return

        await status_msg.edit_text(f"⬇️ Downloading \"{_esc(probe['title'])}\"...")
        result = await asyncio.get_event_loop().run_in_executor(None, mediadl.download_video, text, uid)
        if not result["ok"]:
            await status_msg.edit_text(f"❌ {result['reason']}")
            return

        download_id = result["download_id"]
        # Per-user dict on this user's own session — user A's download_id
        # can never resolve against user B's session, since each user has
        # their own separate SessionState and their own separate dict.
        _s(uid).video_downloads[download_id] = {"file_path": result["file_path"], "dir": result["dir"]}

        if result.get("oversized"):
            # Too big for Telegram's ~50MB bot-upload cap — offer a plain
            # browser download link instead (served by the /download route
            # in app.py, reusing this exact same per-user file). No
            # Convert-to-Audio here since it's not sent through Telegram at all.
            if not PUBLIC_BASE_URL:
                await status_msg.edit_text(
                    "❌ This video is too large for Telegram, and no download link could be generated "
                    "(server URL not configured). Contact the admin."
                )
                mediadl.cleanup_dir(result["dir"])
                _s(uid).video_downloads.pop(download_id, None)
                return

            link = f"{PUBLIC_BASE_URL}/download/{uid}/{download_id}"
            ttl_min = mediadl.WEB_DOWNLOAD_TTL_SECONDS // 60
            await status_msg.edit_text(
                f"🎬 <b>{_esc(result['title'])}</b>\n\n"
                f"This video is too large to send directly in Telegram, so here's a direct download link instead:\n"
                f"{link}\n\n"
                f"⏱ Link works for {ttl_min} minutes — tap it to download the video to your device.",
                parse_mode="HTML"
            )

            async def _cleanup_web_download_later(_uid=uid, _download_id=download_id, _dir_path=result["dir"]):
                await asyncio.sleep(mediadl.WEB_DOWNLOAD_TTL_SECONDS)
                mediadl.cleanup_dir(_dir_path)
                _s(_uid).video_downloads.pop(_download_id, None)
            asyncio.create_task(_cleanup_web_download_later())
            return

        if result.get("has_audio", True):
            keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("🎵 Convert to Audio", callback_data=f"conv_audio_{download_id}")]])
            caption  = f"🎬 {_esc(result['title'])}\n\n⏱ Available for {mediadl.FILE_TTL_SECONDS}s — convert to audio now if you want it."
        else:
            keyboard = None
            caption  = f"🎬 {_esc(result['title'])}\n\n⚠️ This video has no audio track (Convert to Audio isn't available for it)."
        try:
            with open(result["file_path"], "rb") as f:
                await update.message.reply_video(
                    video=f,
                    caption=caption,
                    reply_markup=keyboard
                )
        except Exception as e:
            logger.error(f"[MediaDL] send_video failed: {type(e).__name__}")
            await update.message.reply_text("❌ Downloaded, but couldn't send the video (it may be too large for Telegram).")
            mediadl.cleanup_dir(result["dir"])
            _s(uid).video_downloads.pop(download_id, None)
            await status_msg.delete()
            return

        await status_msg.delete()

        # Auto-delete this download's files ~60s from now no matter what —
        # runs regardless of whether the user converts to audio in time.
        async def _cleanup_video_later(_uid=uid, _download_id=download_id, _dir_path=result["dir"]):
            await asyncio.sleep(mediadl.FILE_TTL_SECONDS)
            mediadl.cleanup_dir(_dir_path)
            _s(_uid).video_downloads.pop(_download_id, None)
        asyncio.create_task(_cleanup_video_later())

    else:
        # ── No active input flow — this is free-text chat, not a bot setting.
        # Route it to the local help agent (no external AI API — see
        # help_agent.py). Typing indicator + a short human-feeling delay
        # apply ONLY here (real chat with the agent) — every other bot
        # feature (buttons, settings input, commands) stays instant.
        # Never uses inline buttons here — unmatched questions get a plain
        # list of things the agent can help with instead.
        await _typing(context, update.effective_chat.id)

        last_topic = _state.get("last_help_topic")
        reply_text, matched_topic = help_agent.answer_question(text, last_topic=last_topic)
        _state["last_help_topic"] = matched_topic  # None clears it — next message starts fresh

        await asyncio.sleep(min(2.0, len(reply_text) / 500))

        try:
            await update.message.reply_text(reply_text, parse_mode="HTML", disable_web_page_preview=True)
        except Exception:
            # Fall back to plain text if HTML somehow fails to parse
            await update.message.reply_text(reply_text, disable_web_page_preview=True)


# ─────────────────────────────────────────
# 🔔 UPGRADE NOTIFIER — background polling
# ─────────────────────────────────────────
# Sends admin notifications for pending upgrade requests every 30 s.
# Completely decoupled from the webhook so it can never crash the bot.
_notified_upgrade_ids: set = set()   # track which requests we already notified about

async def _upgrade_notifier_loop(bot):
    """Poll DB every 30 s for new upgrade requests and notify admins."""
    global _notified_upgrade_ids
    logger.info("[UpgradeNotifier] Background notifier started")
    while True:
        try:
            await asyncio.sleep(30)
            pending = db.get_pending_requests()
            for req in pending:
                uid_r  = req.get("user_id")
                if uid_r in _notified_upgrade_ids:
                    continue   # already notified
                uname_r = req.get("username", "")
                dname_r = req.get("display_name", "")
                contact_r = req.get("contact", "") or "— not provided —"
                msg = (
                    f"🔔 <b>New Upgrade Request!</b>\n\n"
                    f"👤 User ID: <code>{uid_r}</code>\n"
                    f"Username: @{uname_r}\n"
                    f"Name: {dname_r}\n"
                    f"📞 Contact: {contact_r}\n\n"
                    f"Approve: <code>/upgrade {uid_r} 30</code>"
                )
                notified = False
                for admin_id in list(_admin_chat_ids):
                    try:
                        await bot.send_message(
                            chat_id=admin_id, text=msg, parse_mode="HTML"
                        )
                        notified = True
                        logger.info(f"[UpgradeNotifier] Notified admin {admin_id} about uid={uid_r}")
                    except Exception as _e:
                        logger.warning(f"[UpgradeNotifier] Could not reach admin {admin_id}: {_e}")
                if notified:
                    _notified_upgrade_ids.add(uid_r)
                    # Clean up approved IDs no longer pending so re-requests work
                    current_pending_ids = {r.get("user_id") for r in pending}
                    _notified_upgrade_ids &= current_pending_ids
        except Exception as _loop_err:
            logger.error(f"[UpgradeNotifier] Loop error: {_loop_err}")


def _reset_user_session(sess) -> bool:
    """
    Fully deactivate all active features for one user — identical to the
    'Reset Session' button.  Returns True if anything was actually active.
    API keys and saved disk settings are NOT touched.
    """
    was_active = (
        sess.order_monitor_running or sess.refresh_running or
        sess.chat_monitor_enabled  or sess.auto_pay_enabled or
        sess.flw_pay_enabled       or sess.paga_pay_enabled or
        sess.buyer_protection_on   or sess.name_match_enabled or
        sess.sell_msg_enabled      or
        any(slot["running"] for slot in sess.extra_ad_slots)   # Ad 2 / Ad 3
    )
    if not was_active:
        return False

    # Stop all background tasks
    sess.refresh_running       = False
    sess.order_monitor_running = False
    sess.chat_monitor_enabled  = False
    if sess.refresh_task and not sess.refresh_task.done():
        sess.refresh_task.cancel()
    sess.refresh_task = None
    if sess.order_monitor_task and not sess.order_monitor_task.done():
        sess.order_monitor_task.cancel()
    sess.order_monitor_task = None
    if sess.chat_monitor_task and not sess.chat_monitor_task.done():
        sess.chat_monitor_task.cancel()
    sess.chat_monitor_task = None

    # Stop Ad 2 / Ad 3's background tasks ONLY — their configuration (ad_id,
    # mode, float_pct/increment, fetched ad_data) and the slots themselves
    # are intentionally KEPT. This used to wipe sess.extra_ad_slots to []
    # entirely, which deleted Ad 2/Ad 3 outright: the user would find them
    # gone after the hourly reset and have to re-add each slot, re-fetch
    # its ad details, and re-enter its settings from scratch before it
    # could be started again. The hourly reset's job is to stop anything
    # running in the background, not to erase what the user configured.
    for _slot in sess.extra_ad_slots:
        _task = _slot.get("task")
        if _task and not _task.done():
            _task.cancel()
        _slot["running"] = False
        _slot["task"]    = None
        _slot["consecutive_failures"] = 0

    # Deactivate all feature flags
    sess.auto_pay_enabled    = False
    sess.flw_pay_enabled     = False
    sess.paga_pay_enabled    = False
    sess.buyer_protection_on = False
    sess.name_match_enabled  = False
    sess.sell_msg_enabled    = False
    sess.sell_msg_count      = 1

    # Clear volatile order / chat runtime state (dedupe caches, not settings)
    sess.seen_chat_msgs.clear()
    sess.reply_state.clear()
    sess.order_msg_ids.clear()
    sess.my_account_id = ""
    sess.my_nick       = ""
    sess.seen_order_ids = set()
    sess.paid_order_ids = set()
    sess.seen_sell_ids  = set()
    sess.released_ids   = set()
    sess.unpaid_log     = []
    sess.pending_cancel_reviews = {}   # clear pending seller cancel requests
    sess.expecting_cancel_ids  = set()  # clear orders no longer being tracked for cancel review

    # NOTE: sess.ad_data (Ad 1's fetched details), sess.settings (ad_id,
    # mode, increment, float_pct, local_usdt_ref, interval, manage_ad_id,
    # etc.), sess.shared_local_usdt_ref, sess.consecutive_failures, and
    # sess.editing_slot are intentionally left untouched now — this reset
    # only stops what's actively running, it doesn't erase configuration.
    # current_price is left alone too: it's reinitialized from ad_data on
    # the next auto-loop start anyway (see auto_update_loop), so clearing
    # it here bought nothing and only made the "was this cached?" story
    # more confusing.

    logger.info(f"[AutoReset] Background tasks stopped for user {sess.user_id} (settings/ads preserved)")
    return True


async def _session_auto_reset_loop(bot=None):
    """
    Runs every 60 minutes (1 hour).
    For every user with ANY active feature (order monitor, chat monitor,
    auto-pay, sell msg, buyer protection, name match) it:
      1. Fully resets their session — same as pressing the Reset Session button.
      2. Sends a Telegram notification so they can re-enable if still trading.
    Also trims global dicts to prevent unbounded memory growth.
    """
    while True:
        await asyncio.sleep(3600)   # every 60 minutes
        try:
            MAX_IDS  = 500
            notified = 0

            for _sess in get_all_sessions():
                # Always trim memory sets regardless of active state
                if len(_sess.seen_order_ids) > MAX_IDS:
                    _sess.seen_order_ids = set(list(_sess.seen_order_ids)[-MAX_IDS:])
                if len(_sess.paid_order_ids) > MAX_IDS:
                    _sess.paid_order_ids = set(list(_sess.paid_order_ids)[-MAX_IDS:])
                if len(_sess.seen_sell_ids) > MAX_IDS:
                    _sess.seen_sell_ids = set(list(_sess.seen_sell_ids)[-MAX_IDS:])
                if len(_sess.released_ids) > MAX_IDS:
                    _sess.released_ids = set(list(_sess.released_ids)[-MAX_IDS:])
                if len(_sess.seen_chat_msgs) > 30:
                    _keep_keys = list(_sess.seen_chat_msgs.keys())[-30:]
                    _sess.seen_chat_msgs = {k: _sess.seen_chat_msgs[k] for k in _keep_keys}
                if len(_sess.unpaid_log) > 100:
                    _sess.unpaid_log = _sess.unpaid_log[-100:]
                if hasattr(_sess, "order_msg_ids") and len(_sess.order_msg_ids) > 200:
                    _keep = list(_sess.order_msg_ids.items())[-200:]
                    _sess.order_msg_ids = dict(_keep)

                # Reset any user who has at least one active feature, then notify
                was_reset = _reset_user_session(_sess)
                if was_reset and bot:
                    try:
                        await bot.send_message(
                            chat_id=_sess.user_id,
                            text=(
                                "🔄 Scheduled System Reset\n\n"
                                "The bot performs an automatic reset every hour to maintain "
                                "optimal performance and prevent API rate-limit issues.\n\n"
                                "Your running features have been stopped. This includes:\n"
                                "• AD Price Bot (all active ads)\n"
                                "• Order Monitor\n"
                                "• Chat Monitor\n"
                                "• Auto-Pay (Bybit / Flutterwave / Paga)\n"
                                "• Sell Message\n"
                                "• Buyer Protection & Name Match\n\n"
                                "If you are currently trading, please tap /menu and "
                                "re-enable the features you need.\n\n"
                                "✅ Your ad settings — including Ad 2/Ad 3 and fetched ad "
                                "details — API keys, and account settings are all kept "
                                "exactly as you left them. You won't need to set anything "
                                "up again, just restart the features above."
                            )
                        )
                        notified += 1
                    except Exception as _notify_err:
                        logger.debug(f"[AutoReset] Could not notify user {_sess.user_id}: {_notify_err}")

            # Global dict trimming
            global _order_final_states, _order_action_locks, _flw_transfer_registry
            if len(_order_final_states) > 2000:
                _keep_n = list(_order_final_states.items())[-1000:]
                _order_final_states = dict(_keep_n)
            _active_lock_keys = {
                k for k in list(_order_action_locks.keys())
                if k not in _order_final_states
            }
            if len(_order_action_locks) > 500:
                _order_action_locks = {
                    k: v for k, v in _order_action_locks.items()
                    if k in _active_lock_keys
                }
            if len(_flw_transfer_registry) > 200:
                _keep_flw = list(_flw_transfer_registry.items())[-100:]
                _flw_transfer_registry = dict(_keep_flw)
            logger.info(
                f"[AutoReset] Cleanup done | notified={notified} active users | "
                f"final_states={len(_order_final_states)} "
                f"locks={len(_order_action_locks)} "
                f"flw_registry={len(_flw_transfer_registry)}"
            )
        except Exception as e:
            logger.error(f"[AutoReset] Error: {e}")
async def _db_session_cleanup_loop():
    """Clear old disk session files every 12 hours."""
    while True:
        await asyncio.sleep(12 * 3600)
        try:
            count = db.clear_all_old_sessions()
            logger.info(f"[DBCleanup] Cleared {count} stale disk sessions")
        except Exception as e:
            logger.error(f"[DBCleanup] Error: {e}")


# ─────────────────────────────────────────
# 📢 BROADCAST — admin sends a promo to every user (text, photo, or video)
# ─────────────────────────────────────────
async def broadcast_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/broadcast — admin-only. Two-step flow: send the content you want
    broadcast (plain text, or a photo/video with caption), review a
    preview, then confirm before it actually goes out to every user."""
    uid = update.effective_user.id
    if not is_admin(uid):
        return
    user_state["broadcast_draft"] = None
    user_state["action"] = "broadcast_awaiting_content"
    await update.message.reply_text(
        "📢 <b>Broadcast — Step 1/2</b>\n\n"
        "Send the message you want to broadcast to ALL users. You can send:\n"
        "• Plain text\n"
        "• A photo with caption\n"
        "• A video with caption\n\n"
        "Type <code>cancel</code> to cancel.",
        parse_mode="HTML"
    )


async def broadcast_media_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Handles photo/video messages ONLY for the admin broadcast flow.
    Deliberately narrow — every other feature in this bot that reads
    messages (API keys, ad settings, etc.) is plain text, so this quietly
    ignores anything that isn't an admin actively mid-broadcast, with zero
    effect on any other part of the bot.
    """
    uid = update.effective_user.id
    if not is_admin(uid):
        return
    if user_state.get("action") != "broadcast_awaiting_content":
        return

    if update.message.photo:
        file_id, media_type = update.message.photo[-1].file_id, "photo"
    elif update.message.video:
        file_id, media_type = update.message.video.file_id, "video"
    else:
        return

    user_state["broadcast_draft"] = {"type": media_type, "file_id": file_id, "text": update.message.caption or ""}
    user_state["action"] = None
    await _show_broadcast_preview(update, context)


async def _show_broadcast_preview(update: Update, context: ContextTypes.DEFAULT_TYPE):
    draft = user_state.get("broadcast_draft")
    if not draft:
        return
    total_users = len(db.get_all_users())
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton(f"✅ Send to {total_users} users", callback_data="broadcast_confirm")],
        [InlineKeyboardButton("❌ Cancel", callback_data="broadcast_cancel")],
    ])
    note = f"\n\n📊 <b>Step 2/2</b> — this will be sent to {total_users} users. Confirm?"

    if draft["type"] == "text":
        await update.message.reply_text(f"📢 <b>Preview:</b>\n\n{draft['text']}{note}", parse_mode="HTML", reply_markup=keyboard)
    elif draft["type"] == "photo":
        await update.message.reply_photo(photo=draft["file_id"], caption=f"{_esc(draft['text'])}{note}", parse_mode="HTML", reply_markup=keyboard)
    elif draft["type"] == "video":
        await update.message.reply_video(video=draft["file_id"], caption=f"{_esc(draft['text'])}{note}", parse_mode="HTML", reply_markup=keyboard)


async def _run_broadcast(bot, admin_chat_id: int, draft: dict):
    """
    Sends the confirmed broadcast to every user on file, one at a time
    with a small delay to stay comfortably under Telegram's rate limits.
    Blocked/deactivated users and any other per-user send failure are
    counted and skipped — one bad chat_id never stops the rest. On a real
    flood-control response, waits exactly as long as Telegram asks, then
    retries that one user once before moving on.
    """
    users = db.get_all_users()
    sent = failed = 0

    for user in users:
        target_id = user.get("user_id")
        if not target_id:
            continue
        try:
            await _send_broadcast_item(bot, target_id, draft)
            sent += 1
        except RetryAfter as e:
            await asyncio.sleep(e.retry_after)
            try:
                await _send_broadcast_item(bot, target_id, draft)
                sent += 1
            except Exception:
                failed += 1
        except Forbidden:
            failed += 1   # user blocked the bot / deactivated — skip, not fatal
        except Exception as e:
            failed += 1
            logger.debug(f"[Broadcast] failed for {target_id}: {e}")
        await asyncio.sleep(0.05)   # ~20 msg/sec — safely under Telegram's flood limits

    await bot.send_message(
        chat_id=admin_chat_id,
        text=f"📢 <b>Broadcast complete!</b>\n\n✅ Sent: {sent}\n❌ Failed/blocked: {failed}\n👥 Total users: {len(users)}",
        parse_mode="HTML"
    )


async def _send_broadcast_item(bot, chat_id: int, draft: dict):
    if draft["type"] == "text":
        await bot.send_message(chat_id=chat_id, text=draft["text"], parse_mode="HTML")
    elif draft["type"] == "photo":
        await bot.send_photo(chat_id=chat_id, photo=draft["file_id"], caption=draft["text"], parse_mode="HTML")
    elif draft["type"] == "video":
        await bot.send_video(chat_id=chat_id, video=draft["file_id"], caption=draft["text"], parse_mode="HTML")


async def refresh_scammers_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _get_or_register_user(update.effective_user)
    uid = update.effective_user.id
    if not is_admin(uid) and not sub.is_pro(uid):
        await update.message.reply_text(
            "🔒 <b>Pro Plan Required</b>\n\n"
            "The /refreshscammers command is available on the Pro plan only.\n\n"
            "Tap <b>⬆️ Upgrade Plan</b> in /menu to request access from the admin.",
            parse_mode="HTML"
        )
        return
    await update.message.reply_text("⏳ Refreshing scammer list from GitHub...")
    count = await asyncio.get_event_loop().run_in_executor(None, load_scammers)
    updated = get_last_updated()
    if count > 0:
        await update.message.reply_text(
            f"✅ <b>Scammer list refreshed!</b>\n\n"
            f"📋 <code>{count}</code> names loaded\n"
            f"🕐 Updated: <code>{updated}</code>",
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(
            "❌ *Failed to load scammer list.*\n\n"
            "Check that `scammers.txt` exists in your GitHub repo\n"
            "and `SCAMMERS_FILE_URL` is set correctly.",
            parse_mode="HTML"
        )


async def check_name_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Manually check a name against the scammer list. Usage: /checkname John Doe"""
    _get_or_register_user(update.effective_user)
    uid = update.effective_user.id
    if not is_admin(uid) and not sub.is_pro(uid):
        await update.message.reply_text(
            "🔒 <b>Pro Plan Required</b>\n\n"
            "The /checkname command is available on the Pro plan only.\n\n"
            "Tap <b>⬆️ Upgrade Plan</b> in /menu to request access from the admin.",
            parse_mode="HTML"
        )
        return
    name = " ".join(context.args).strip() if context.args else ""
    if not name:
        await update.message.reply_text(
            "Usage: `/checkname John Doe`\n\nChecks a name against your scammer list.",
            parse_mode="HTML"
        )
        return
    fraud = await asyncio.get_event_loop().run_in_executor(None, check_buyer_name, name)
    count = get_scammer_count()
    if fraud["flagged"]:
        match_label = {
            "exact":   "🔴 Exact match",
            "partial": "🟠 Partial match",
            "fuzzy":   "🟡 Similar name",
        }.get(fraud["match_type"], "⚠️ Match")
        await update.message.reply_text(
            f"🚨 <b>FLAGGED!</b>\n\n"
            f"Name: <code>{name}</code>\n"
            f"{match_label}: <code>{fraud['matched_name']}</code>\n"
            f"Similarity: <code>{fraud['similarity']:.0%}</code>\n\n"
            f"<i>(Checked against {count} names)</i>",
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(
            f"✅ <b>Not found</b> — <code>{name}</code> is not in your scammer list.\n\n"
            f"<i>(Checked against {count} names)</i>",
            parse_mode="HTML"
        )

# ─────────────────────────────────────────
# 📊 /userdata — Admin export (overrides admin_commands import)
# Includes total_buy_orders + total_sell_orders from DB and live session.
# ─────────────────────────────────────────
async def cmd_userdata(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Download all user data as Excel. Admin only.

    Buy/sell totals:
      • DB value  — persisted by order monitor each time a new order is seen
      • Live session — get_session(uid).seen_order_ids / seen_sell_ids set sizes
      • Whichever is HIGHER wins, so totals are never under-reported.
    Totals reset naturally when a session clears — no permanent analytics added.
    """
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("❌ Admin only.")
        return

    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        await update.message.reply_text(
            "❌ <b>openpyxl not installed.</b>\n\nRun: <code>pip install openpyxl</code>",
            parse_mode="HTML"
        )
        return

    await update.message.reply_text("⏳ Building user data export...")

    try:
        users = db.get_all_users() or []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        headers = [
            "User ID", "Username", "Display Name", "Plan", "Plan Expires",
            "Upgrade Pending", "Created At", "Last Active",
            "Total BUY Orders", "Total SELL Orders",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, user in enumerate(users, 2):
            uid = int(user.get("user_id") or user.get("id") or 0)

            # ── DB totals (persisted cumulatively by order monitor) ──
            db_buy  = int(user.get("total_buy_orders",  0) or 0)
            db_sell = int(user.get("total_sell_orders", 0) or 0)

            # ── Live session totals (in-memory, current session only) ──
            # get_session(uid) is safe to call for any uid — returns empty session
            # if the user has no active session (sets will be empty → 0 counts).
            try:
                sess      = get_session(uid)
                live_buy  = len(getattr(sess, "seen_order_ids", None) or set())
                live_sell = len(getattr(sess, "seen_sell_ids",  None) or set())
            except Exception:
                live_buy  = 0
                live_sell = 0

            # Take whichever is higher — DB may lag if session hasn't flushed yet,
            # live session resets to 0 after cleanup, so max() covers both cases.
            total_buy  = max(db_buy,  live_buy)
            total_sell = max(db_sell, live_sell)

            row_data = [
                uid,
                user.get("username",        ""),
                user.get("display_name",    "") or user.get("full_name", ""),
                user.get("plan",            "free"),
                user.get("plan_expires",    "") or user.get("plan_expiry", ""),
                user.get("upgrade_pending", False),
                user.get("created_at",      ""),
                user.get("stats", {}).get("last_active", "") or user.get("last_active", "") or user.get("last_seen", ""),
                total_buy,
                total_sell,
            ]
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=val)

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from datetime import datetime as _dt
        fname = f"userdata_{_dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        await update.message.reply_document(
            document=buf,
            filename=fname,
            caption=(
                f"📊 <b>User Data Export</b>\n\n"
                f"👥 {len(users)} users\n"
                f"🕐 Generated: <code>{_dt.now().strftime('%Y-%m-%d %H:%M:%S')}</code>\n\n"
                f"BUY/SELL totals: DB cumulative + live session (max of both)."
            ),
            parse_mode="HTML"
        )

    except Exception as _ude:
        import traceback
        logger.error(f"[userdata] Export error: {_ude}\n{traceback.format_exc()}")
        await update.message.reply_text(
            f"❌ <b>Export failed</b>\n\n<code>{_esc(str(_ude)[:300])}</code>",
            parse_mode="HTML"
        )


def start_bot():
    global _paga_queue, _paga_worker_task

    application = (
        ApplicationBuilder()
        .token(TELEGRAM_TOKEN)
        .updater(None)
        .build()
    )
    # ── User commands ──
    application.add_handler(CommandHandler("start",            start))
    application.add_handler(CommandHandler("menu",             menu_command))
    application.add_handler(CommandHandler("pingbybit",        ping_bybit_command))
    application.add_handler(CommandHandler("pingflutterwave",  ping_flutterwave_command))
    application.add_handler(CommandHandler("pingpaga",         ping_paga_command))
    application.add_handler(CommandHandler("refreshscammers",  refresh_scammers_command))
    application.add_handler(CommandHandler("checkname",        check_name_command))

    # ── Admin-only commands ──
    application.add_handler(CommandHandler("upgrade",       cmd_upgrade))
    application.add_handler(CommandHandler("downgrade",     cmd_downgrade))
    application.add_handler(CommandHandler("requests",      cmd_requests))
    application.add_handler(CommandHandler("listusers",     cmd_listusers))
    application.add_handler(CommandHandler("userdata",      cmd_userdata))
    application.add_handler(CommandHandler("awardref",      cmd_awardref))
    application.add_handler(CommandHandler("addbalance",    cmd_addbalance))
    application.add_handler(CommandHandler("deductbalance", cmd_deductbalance))
    application.add_handler(CommandHandler("referrals",     cmd_referrals))
    application.add_handler(CommandHandler("withdrawals",      cmd_withdrawals))
    application.add_handler(CommandHandler("approvewithdraw",  cmd_approvewithdraw))
    application.add_handler(CommandHandler("rejectwithdraw",   cmd_rejectwithdraw))
    application.add_handler(CommandHandler("broadcast",        broadcast_command))

    application.add_handler(CallbackQueryHandler(button_handler))
    application.add_handler(MessageHandler(filters.PHOTO | filters.VIDEO, broadcast_media_handler))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))

    # ── Global error handler — logs ALL unhandled exceptions with full traceback ──
    async def _global_error_handler(update, context):
        import traceback
        tb = "".join(traceback.format_exception(type(context.error), context.error, context.error.__traceback__))
        logger.error(
            f"[GlobalError] Unhandled exception\n"
            f"  update={update}\n"
            f"  error={context.error}\n"
            f"{tb}"
        )
        # Optionally notify the user something went wrong
        try:
            if update and update.effective_message:
                await update.effective_message.reply_text(
                    "⚠️ An unexpected error occurred. Please try again or use /menu to restart."
                )
        except Exception:
            pass

    application.add_error_handler(_global_error_handler)

    async def _post_init(app):
        global _paga_queue, _paga_worker_task
        _paga_queue       = asyncio.Queue()
        _paga_worker_task = asyncio.create_task(_paga_queue_worker())

        # Pre-load scammer list — run_in_executor returns a Future not a coroutine,
        # so wrap it in an async helper before passing to create_task.
        async def _preload_scammers():
            await asyncio.get_event_loop().run_in_executor(None, load_scammers)
        asyncio.create_task(_preload_scammers())

        # Auto-reset stale sessions every 60 minutes + notify active users
        asyncio.create_task(_session_auto_reset_loop(app.bot))

        # Auto-clear old DB sessions every 12h
        asyncio.create_task(_db_session_cleanup_loop())

        # Background upgrade request notifier (polls DB, notifies admins)
        asyncio.create_task(_upgrade_notifier_loop(app.bot))

        # ── Set admin-scoped bot commands so only current ADMIN_IDS see admin cmds ──
        # This re-syncs on every deploy, so removed admin IDs lose the menu immediately.
        from telegram import BotCommand, BotCommandScopeChat
        admin_commands = [
            BotCommand("upgrade",   "Upgrade a user to Pro"),
            BotCommand("downgrade", "Downgrade a user"),
            BotCommand("requests",  "List upgrade requests"),
            BotCommand("listusers", "List all users"),
            BotCommand("userdata",  "Download user data Excel"),
        ]
        user_commands = [
            BotCommand("start",           "Start the bot"),
            BotCommand("menu",            "Open main menu"),
            BotCommand("pingbybit",       "Test Bybit API"),
            BotCommand("pingflutterwave", "Test Flutterwave API"),
            BotCommand("pingpaga",        "Test Paga API"),
            BotCommand("refreshscammers", "Refresh scammer list"),
            BotCommand("checkname",       "Check a name against scammer list"),
        ]
        # Set user-level commands for everyone (default scope)
        try:
            await app.bot.set_my_commands(user_commands)
        except Exception as _e:
            logger.warning(f"[Init] Could not set default commands: {_e}")
        # Set combined commands for each active admin individually
        for _admin_id in list(ADMIN_IDS):
            try:
                await app.bot.set_my_commands(
                    user_commands + admin_commands,
                    scope=BotCommandScopeChat(chat_id=_admin_id)
                )
                logger.info(f"[Init] Admin commands set for {_admin_id}")
            except Exception as _e:
                logger.warning(f"[Init] Could not set admin commands for {_admin_id}: {_e}")

        logger.info("🟡 Paga queue + session manager + upgrade notifier started")

    application.post_init = _post_init
    logger.info("🤖 Bot handlers registered")
    return application
